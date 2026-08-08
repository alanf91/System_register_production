# VERSÃO 5.20: adiciona painel acumulado diário por setor e metas.
from __future__ import annotations

import hashlib
import json
import hmac
import os
import sqlite3
import re
import shutil
import subprocess
import sys
import unicodedata
from io import BytesIO
from datetime import date, datetime, time, timedelta
from html import escape
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel

NOME_ABA_PADRAO = "5_ACOMPANHAMENTO"
ARQUIVO_PADRAO = "PPCP_ACOMPAN.xlsx"
VERSAO_APP = "SERVIDOR LOCAL"
LOGIN_APONTAMENTO = "pcp"
SENHA_APONTAMENTO = "pcp"
NOME_ABA_PARADAS = "6_Paradas"
CABECALHOS_PARADAS = [
    "Data Parada",
    "Setor",
    "Máquina/Posto",
    "Motivo da Parada",
    "Tempo Total Parada (min)",
    "Observação",
    "Registrado Em",
    "Usuário",
    "Editado Em",
    "Editado Por",
]

VERSAO_SCHEMA_PARADAS = "2"

# Máquinas que devem permanecer disponíveis no apontamento de paradas,
# mesmo quando ainda não aparecem na coluna Máquina/Posto da base de produção.
MAQUINAS_PARADAS_CADASTRADAS = [
    "124 - CNC 02 (NESTING MKL2030 - SERIE YW202412)",
    "23 - COLADEIRA 01 (MOD CBN 50T)",
    "24 - COLADEIRA 02 (MOD CBN 50T)",
    "75 - COPIADORA (VMR DALMAQ FC MOD SERIE 338)",
    "91 - SELADORA MAQUINA PACK N/I",
    "120 - CARTONAGEM (FITSIZE1600-INNOVATOR 2023OCT01)",
    "20 - FURADEIRA 02 (F500 - B)",
    "21 - FURADEIRA 03 (F400 - CS)",
    "22 - FURADEIRA 01 (DL 22-11 12/06)",
    "65 - FURADEIRA 04 (F400 - T)",
    "94 - LIXADEIRA DE CABINE (VMR) PU",
    "28 - CALIBRADEIRA MACLINEA (SP3CCT1 1350) UV",
    "10 - SECCIONADORA 01 (GYBEN SMART SPT N910689)",
    "11 - SECCIONADORA 02 (GYBEN 2 ICON FAST SPT125 H115)",
    "96 - SERRAFITA AUTO (YOWCHERNEG YC-025 2220103)",
    "124 - PADE (VELOX - CNC 5 - EIXOS)",
    "129 - SELADORA EMBALAGEM (MAQUINAPACK - MSA 2200X400)",
]


ARQUIVO_BANCO_OPERACIONAL = "apontamentos_operacionais.db"
ARQUIVO_CATALOGO_FICHAS = "catalogo_fichas.db"
SENHA_PADRAO_OPERADORES = "1234"
MOTIVOS_PARADA_OPERACIONAL = [
    "Manutenção",
    "Falta de peças",
]

# Um usuário por equipamento. A máquina fica vinculada ao login e não pode ser
# trocada pelo operador durante o apontamento.
EQUIPAMENTOS_OPERADORES = [
    {"usuario": "maq124_cnc", "codigo": "124", "setor": "60 USINAGEM AUTOMATIZADA", "maquina": "124 - CNC 02 (NESTING MKL2030 - SERIE YW202412)"},
    {"usuario": "maq023", "codigo": "23", "setor": "9 COLADEIRAS", "maquina": "23 - COLADEIRA 01 (MOD CBN 50T)"},
    {"usuario": "maq024", "codigo": "24", "setor": "9 COLADEIRAS", "maquina": "24 - COLADEIRA 02 (MOD CBN 50T)"},
    {"usuario": "maq075", "codigo": "75", "setor": "56 COPIADORA", "maquina": "75 - COPIADORA (VMR DALMAQ FC MOD SERIE 338)"},
    {"usuario": "maq091", "codigo": "91", "setor": "8 EMBALAGEM", "maquina": "91 - SELADORA MAQUINA PACK N/I"},
    {"usuario": "maq120", "codigo": "120", "setor": "8 EMBALAGEM", "maquina": "120 - CARTONAGEM (FITSIZE1600-INNOVATOR 2023OCT01)"},
    {"usuario": "maq020", "codigo": "20", "setor": "10 FURADEIRAS", "maquina": "20 - FURADEIRA 02 (F500 - B)"},
    {"usuario": "maq021", "codigo": "21", "setor": "10 FURADEIRAS", "maquina": "21 - FURADEIRA 03 (F400 - CS)"},
    {"usuario": "maq022", "codigo": "22", "setor": "10 FURADEIRAS", "maquina": "22 - FURADEIRA 01 (DL 22-11 12/06)"},
    {"usuario": "maq065", "codigo": "65", "setor": "10 FURADEIRAS", "maquina": "65 - FURADEIRA 04 (F400 - T)"},
    {"usuario": "maq094", "codigo": "94", "setor": "36 PINTURA P.U.", "maquina": "94 - LIXADEIRA DE CABINE (VMR) PU"},
    {"usuario": "maq028", "codigo": "28", "setor": "6 PINTURA UV", "maquina": "28 - CALIBRADEIRA MACLINEA (SP3CCT1 1350) UV"},
    {"usuario": "maq010", "codigo": "10", "setor": "50 SECCIONADORA 2", "maquina": "10 - SECCIONADORA 01 (GYBEN SMART SPT N910689)"},
    {"usuario": "maq011", "codigo": "11", "setor": "50 SECCIONADORA 2", "maquina": "11 - SECCIONADORA 02 (GYBEN 2 ICON FAST SPT125 H115)"},
    {"usuario": "maq096", "codigo": "96", "setor": "60 USINAGEM AUTOMATIZADA", "maquina": "96 - SERRAFITA AUTO (YOWCHERNEG YC-025 2220103)"},
    {"usuario": "maq124_pade", "codigo": "124", "setor": "60 USINAGEM AUTOMATIZADA", "maquina": "124 - PADE (VELOX - CNC 5 - EIXOS)"},
    {"usuario": "maq129", "codigo": "129", "setor": "8 EMBALAGEM", "maquina": "129 - SELADORA EMBALAGEM (MAQUINAPACK - MSA 2200X400)"},
]



# -----------------------------------------------------------------------------
# Utilidades gerais
# -----------------------------------------------------------------------------
def normalizar(texto: Any) -> str:
    """Normaliza textos para comparar nomes de colunas sem acento/caixa."""
    if texto is None:
        return ""
    texto = str(texto).strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = re.sub(r"[^a-z0-9%]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def texto_limpo(valor: Any) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def numero_limpo(valor: Any) -> Any:
    if valor is None or valor == "":
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return int(valor)
    return valor


def valor_para_data(valor: Any) -> date | None:
    """Converte datas reais ou serial Excel para date."""
    if valor in (None, ""):
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if isinstance(valor, (int, float)) and 20000 <= valor <= 60000:
        try:
            return from_excel(valor).date()
        except Exception:
            return None
    try:
        convertido = pd.to_datetime(valor, dayfirst=True, errors="coerce")
        if pd.notna(convertido):
            return convertido.date()
    except Exception:
        pass
    return None


def data_br(valor: Any) -> str:
    d = valor_para_data(valor)
    return d.strftime("%d/%m/%Y") if d else ""


def to_num(serie: pd.Series) -> pd.Series:
    return pd.to_numeric(serie, errors="coerce").fillna(0)


def fmt_num(valor: Any, casas: int = 0) -> str:
    try:
        numero = float(valor)
    except Exception:
        numero = 0
    texto = f"{numero:,.{casas}f}"
    return texto.replace(",", "X").replace(".", ",").replace("X", ".")


def dataframe_para_csv_excel(
    df: pd.DataFrame,
    separador: str = ";",
    decimal: str = ",",
) -> bytes:
    """Gera CSV UTF-8 com BOM em bytes para o Excel reconhecer acentos corretamente.

    O parâmetro ``encoding`` do pandas é ignorado quando ``to_csv`` devolve uma
    string. Por isso a conversão explícita para bytes é necessária.
    """
    texto = df.to_csv(
        index=False,
        sep=separador,
        decimal=decimal,
        lineterminator="\r\n",
    )
    return texto.encode("utf-8-sig")


def dataframe_para_excel(
    df: pd.DataFrame,
    nome_aba: str = "Dados",
    linhas_alerta: pd.Series | list[bool] | tuple[bool, ...] | None = None,
) -> bytes:
    """Gera um XLSX nativo, preservando acentos e aplicando formatação básica.

    Quando ``linhas_alerta`` é informado, as linhas correspondentes recebem
    fonte vermelha, negrito e preenchimento vermelho-claro.
    """
    nome_seguro = re.sub(r"[\\/*?:\[\]]", " ", texto_limpo(nome_aba))[:31] or "Dados"
    saida = BytesIO()

    alertas: list[bool] | None = None
    if linhas_alerta is not None:
        alertas = [bool(valor) for valor in list(linhas_alerta)]
        if len(alertas) != len(df):
            raise ValueError(
                "A quantidade de marcações de alerta deve ser igual à quantidade de linhas do DataFrame."
            )

    with pd.ExcelWriter(saida, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=nome_seguro)
        ws = writer.book[nome_seguro]

        preenchimento_cabecalho = PatternFill("solid", fgColor="1F4E78")
        for celula in ws[1]:
            celula.fill = preenchimento_cabecalho
            celula.font = Font(color="FFFFFF", bold=True)
            celula.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = False
        ws.row_dimensions[1].height = 34

        for indice_coluna, coluna in enumerate(df.columns, start=1):
            cabecalho = texto_limpo(coluna)
            cabecalho_norm = normalizar(cabecalho)
            amostra = df[coluna].head(1000).astype(object)
            amostra = amostra.where(pd.notna(amostra), "")
            valores = amostra.astype(str).tolist()
            maior_texto = max([len(cabecalho), *(len(v) for v in valores)], default=10)
            largura_maxima = 55 if any(
                termo in cabecalho_norm
                for termo in ["descricao", "equipamento", "maquina", "lotes", "observacao"]
            ) else 30
            ws.column_dimensions[get_column_letter(indice_coluna)].width = min(
                max(maior_texto + 2, 11), largura_maxima
            )

            for linha in range(2, ws.max_row + 1):
                celula = ws.cell(row=linha, column=indice_coluna)
                celula.alignment = Alignment(vertical="center", wrap_text=False)

                if isinstance(celula.value, (datetime, date)):
                    celula.number_format = "dd/mm/yyyy"
                elif isinstance(celula.value, (int, float)) and not isinstance(celula.value, bool):
                    if "%" in cabecalho:
                        # Os percentuais do relatório já estão na escala 0 a 100.
                        celula.number_format = '0.0"%"'
                    elif any(termo in cabecalho_norm for termo in ["metrica", "comprimento", "largura", "fator"]):
                        celula.number_format = "#,##0.000"
                    elif any(termo in cabecalho_norm for termo in ["dias", "linhas", "quantidade de"]):
                        celula.number_format = "0"
                    else:
                        celula.number_format = "#,##0.00"

        if alertas is not None:
            preenchimento_alerta = PatternFill("solid", fgColor="FDE9E7")
            fonte_alerta = Font(color="C00000", bold=True)
            for indice_dataframe, em_alerta in enumerate(alertas, start=2):
                if not em_alerta:
                    continue
                for celula in ws[indice_dataframe]:
                    celula.fill = preenchimento_alerta
                    celula.font = fonte_alerta

    saida.seek(0)
    return saida.getvalue()


def mascara_programado_nao_atendido(df: pd.DataFrame) -> pd.Series:
    """Identifica setores/equipamentos que ficaram abaixo da quantidade programada.

    Na matriz diária, utiliza o total do período selecionado. Quando os totais
    não existem, verifica os pares de colunas diárias de peças. Linhas sem
    programação não são marcadas como atraso.
    """
    if df.empty:
        return pd.Series(False, index=df.index, dtype=bool)

    tolerancia = 1e-9
    if {"Total Programado (peças)", "Total Apontado (peças)"}.issubset(df.columns):
        programado = pd.to_numeric(df["Total Programado (peças)"], errors="coerce").fillna(0)
        apontado = pd.to_numeric(df["Total Apontado (peças)"], errors="coerce").fillna(0)
        return (programado > tolerancia) & (apontado + tolerancia < programado)

    if {"Programado Peças", "Apontado Peças"}.issubset(df.columns):
        programado = pd.to_numeric(df["Programado Peças"], errors="coerce").fillna(0)
        apontado = pd.to_numeric(df["Apontado Peças"], errors="coerce").fillna(0)
        return (programado > tolerancia) & (apontado + tolerancia < programado)

    if {"Programado", "Apontado"}.issubset(df.columns):
        programado = pd.to_numeric(df["Programado"], errors="coerce").fillna(0)
        apontado = pd.to_numeric(df["Apontado"], errors="coerce").fillna(0)
        return (programado > tolerancia) & (apontado + tolerancia < programado)

    resultado = pd.Series(False, index=df.index, dtype=bool)
    for coluna_programado in [
        coluna for coluna in df.columns if str(coluna).endswith("— Prog. peças")
    ]:
        coluna_apontado = str(coluna_programado).replace("— Prog. peças", "— Apont. peças")
        if coluna_apontado not in df.columns:
            continue
        programado = pd.to_numeric(df[coluna_programado], errors="coerce").fillna(0)
        apontado = pd.to_numeric(df[coluna_apontado], errors="coerce").fillna(0)
        resultado |= (programado > tolerancia) & (apontado + tolerancia < programado)
    return resultado


def estilizar_programado_nao_atendido(df: pd.DataFrame):
    """Aplica destaque vermelho às linhas abaixo do programado no Streamlit."""
    alertas = mascara_programado_nao_atendido(df)
    estilos = pd.DataFrame("", index=df.index, columns=df.columns)
    estilos.loc[alertas, :] = (
        "color: #C62828; background-color: #FFEBEE; font-weight: 700;"
    )
    return df.style.apply(lambda _dados: estilos, axis=None)


def chave_ordenacao_texto(texto: Any) -> list[Any]:
    """Ordena textos com números de forma natural: 6 antes de 10."""
    partes = re.split(r"(\d+)", texto_limpo(texto))
    return [int(p) if p.isdigit() else p.lower() for p in partes]


def opcoes_unicas(df: pd.DataFrame, coluna: str) -> list[str]:
    """Retorna opções únicas limpas, sem depender dos filtros atuais."""
    if coluna not in df.columns:
        return []
    vistos: set[str] = set()
    opcoes: list[str] = []
    for valor in df[coluna].dropna().tolist():
        texto = texto_limpo(valor)
        if texto and texto not in vistos:
            vistos.add(texto)
            opcoes.append(texto)
    return sorted(opcoes, key=chave_ordenacao_texto)


def combinar_opcoes_maquinas(*listas: list[str]) -> list[str]:
    """Combina máquinas da planilha e do cadastro fixo, removendo duplicidades."""
    opcoes: list[str] = []
    vistos: set[str] = set()
    for lista in listas:
        for valor in lista:
            texto = texto_limpo(valor)
            chave = normalizar(texto)
            if texto and chave not in vistos:
                vistos.add(chave)
                opcoes.append(texto)
    return opcoes


# -----------------------------------------------------------------------------
# Leitura e gravação da planilha
# -----------------------------------------------------------------------------
def localizar_colunas(ws) -> dict[str, int]:
    """Localiza colunas pelo cabeçalho da linha 1."""
    cabecalhos = {normalizar(ws.cell(row=1, column=c).value): c for c in range(1, ws.max_column + 1)}

    candidatos = {
        "data_programada": ["data programada", "programada", "data planejamento", "data planejada"],
        # Data Entrega é usada especificamente para medir atraso contra a data de hoje.
        # Se a planilha não tiver esta coluna, o sistema usa Data Programada como fallback.
        "data_entrega": ["data entrega", "data de entrega", "entrega", "dt entrega", "prazo entrega", "prazo de entrega"],
        "lote": ["op lote", "lote", "op/lote"],
        "cliente": ["cliente pedido", "cliente/pedido"],
        "produto": ["produto equipamento", "produto/equipamento"],
        "codigo": ["codigo peca", "cod peca", "codigo"],
        "descricao": ["descricao peca", "descricao"],
        "operacao": ["operacao", "operação"],
        "setor": ["setor"],
        "maquina": ["maquina posto", "maquina/posto", "máquina/posto"],
        "qtde_programada": ["qtde programada", "quantidade programada"],
        "setup_padrao": ["setup padrao min", "setup padrão min", "setup padrao"],
        "tempo_unitario": ["tempo unit padrao min peca", "tempo unit padrão min peça", "tempo unit padrao", "tempo ciclo", "tempo de ciclo"],
        "tempo_programado": ["tempo programado min", "tempo programado"],
        "data_realizada": ["data realizada"],
        "qtde_realizada": ["qtde realizada", "quantidade realizada"],
        "inicio_real": ["inicio real", "início real"],
        "fim_real": ["fim real"],
        "tempo_realizado": ["tempo realizado min", "tempo realizado"],
        "refugo_retrabalho": ["refugo retrabalho", "refugo/retrabalho"],
        "percentual": ["% realizado", "percentual realizado"],
        "status": ["status"],
        "atraso": ["atraso?", "atraso"],
    }

    colunas: dict[str, int] = {}
    for chave, nomes in candidatos.items():
        for nome in nomes:
            nome_norm = normalizar(nome)
            if nome_norm in cabecalhos:
                colunas[chave] = cabecalhos[nome_norm]
                break

    # Compatibilidade: algumas bases antigas têm apenas Data Entrega ou apenas Data Programada.
    if "data_programada" not in colunas and "data_entrega" in colunas:
        colunas["data_programada"] = colunas["data_entrega"]
    if "data_entrega" not in colunas and "data_programada" in colunas:
        colunas["data_entrega"] = colunas["data_programada"]

    obrigatorias = [
        "data_programada",
        "lote",
        "setor",
        "qtde_programada",
        "data_realizada",
        "qtde_realizada",
    ]
    faltando = [c for c in obrigatorias if c not in colunas]
    if faltando:
        encontrados = ", ".join(str(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1))
        raise ValueError(
            "Não localizei estas colunas obrigatórias: "
            + ", ".join(faltando)
            + "\n\nCabeçalhos encontrados: "
            + encontrados
        )
    return colunas


def valor_celula(ws, linha: int, colunas: dict[str, int], chave: str) -> Any:
    col = colunas.get(chave)
    if not col:
        return ""
    return ws.cell(linha, col).value


def float_seguro(valor: Any) -> float:
    if valor in (None, ""):
        return 0.0
    try:
        return float(valor)
    except Exception:
        try:
            return float(str(valor).strip().replace(",", "."))
        except Exception:
            return 0.0


def carregar_tempos_cadastro(wb_valores) -> dict[str, tuple[float, float]]:
    """Lê os tempos técnicos da aba de cadastro quando as fórmulas do acompanhamento não têm cache."""
    nome_aba_cadastro = "01_Cad_Pecas"
    if nome_aba_cadastro not in wb_valores.sheetnames:
        return {}

    ws_cad = wb_valores[nome_aba_cadastro]
    cabecalhos = {
        normalizar(ws_cad.cell(row=1, column=c).value): c
        for c in range(1, ws_cad.max_column + 1)
    }
    col_codigo = cabecalhos.get(normalizar("Codigo Peca"))
    col_setup = cabecalhos.get(normalizar("Tempo Setup Padrao (min)"))
    col_unitario = cabecalhos.get(normalizar("Tempo Unit Padrao (min/peca)"))
    if not col_codigo or not col_setup or not col_unitario:
        return {}

    lookup: dict[str, tuple[float, float]] = {}
    for linha in range(2, ws_cad.max_row + 1):
        codigo = texto_limpo(ws_cad.cell(linha, col_codigo).value)
        if not codigo:
            continue
        setup = float_seguro(ws_cad.cell(linha, col_setup).value)
        unitario = float_seguro(ws_cad.cell(linha, col_unitario).value)
        # MATCH do Excel retorna a primeira ocorrência; setdefault preserva o mesmo comportamento.
        lookup.setdefault(codigo, (setup, unitario))
    return lookup


def carregar_dados(caminho: Path, nome_aba: str) -> tuple[pd.DataFrame, dict[str, int]]:
    # Abre duas leituras: uma preserva fórmulas/cabeçalhos e outra traz os valores calculados
    # que o Excel deixou gravados. Isso é importante para Tempo Programado (min), que é fórmula.
    wb = load_workbook(caminho, data_only=False)
    wb_valores = load_workbook(caminho, data_only=True)
    if nome_aba not in wb.sheetnames:
        wb.close()
        wb_valores.close()
        raise ValueError(f'A aba "{nome_aba}" não existe. Abas encontradas: {", ".join(wb.sheetnames)}')
    ws = wb[nome_aba]
    ws_valores = wb_valores[nome_aba]
    col = localizar_colunas(ws)
    tempos_cadastro = carregar_tempos_cadastro(wb_valores)

    linhas = []
    for linha_excel in range(2, ws.max_row + 1):
        lote = ws_valores.cell(linha_excel, col["lote"]).value
        setor = ws_valores.cell(linha_excel, col["setor"]).value
        qtd_prog = ws_valores.cell(linha_excel, col["qtde_programada"]).value
        if lote in (None, "") and setor in (None, "") and qtd_prog in (None, ""):
            continue

        data_prog_valor = ws_valores.cell(linha_excel, col["data_programada"]).value
        data_entrega_valor = ws_valores.cell(linha_excel, col["data_entrega"]).value if "data_entrega" in col else data_prog_valor
        data_real_valor = ws_valores.cell(linha_excel, col["data_realizada"]).value
        data_prog = valor_para_data(data_prog_valor)
        data_entrega = valor_para_data(data_entrega_valor)
        data_real = valor_para_data(data_real_valor)
        qtde_real = ws_valores.cell(linha_excel, col["qtde_realizada"]).value

        codigo_valor = valor_celula(ws_valores, linha_excel, col, "codigo")
        setup_valor = valor_celula(ws_valores, linha_excel, col, "setup_padrao")
        tempo_unitario_valor = valor_celula(ws_valores, linha_excel, col, "tempo_unitario")
        tempo_programado_valor = valor_celula(ws_valores, linha_excel, col, "tempo_programado")

        # O openpyxl não calcula fórmulas. Se a planilha veio sem cache calculado,
        # busca Setup e Tempo Unitário diretamente no cadastro técnico de peças.
        tempos_tecnicos = tempos_cadastro.get(texto_limpo(codigo_valor))
        if tempos_tecnicos:
            if setup_valor in (None, ""):
                setup_valor = tempos_tecnicos[0]
            if tempo_unitario_valor in (None, ""):
                tempo_unitario_valor = tempos_tecnicos[1]

        tempo_calculado = float_seguro(setup_valor) + (
            float_seguro(tempo_unitario_valor) * float_seguro(qtd_prog)
        )
        if float_seguro(tempo_programado_valor) <= 0 and tempo_calculado > 0:
            tempo_programado_valor = tempo_calculado

        linhas.append(
            {
                "Selecionar": False,
                "Linha Excel": linha_excel,
                "Data Programada": data_br(data_prog_valor),
                "Data Programada Valor": data_prog,
                "Data Entrega": data_br(data_entrega_valor),
                "Data Entrega Valor": data_entrega,
                "OP/Lote": texto_limpo(lote),
                "Cliente/Pedido": texto_limpo(valor_celula(ws_valores, linha_excel, col, "cliente")),
                "Produto/Equipamento": texto_limpo(valor_celula(ws_valores, linha_excel, col, "produto")),
                "Código Peça": texto_limpo(codigo_valor),
                "Descrição Peça": texto_limpo(valor_celula(ws_valores, linha_excel, col, "descricao")),
                "Operação": texto_limpo(valor_celula(ws_valores, linha_excel, col, "operacao")),
                "Setor": texto_limpo(setor),
                "Máquina/Posto": texto_limpo(valor_celula(ws_valores, linha_excel, col, "maquina")),
                "Qtde Programada": numero_limpo(qtd_prog),
                "Setup Padrão (min)": numero_limpo(setup_valor),
                "Tempo Unit Padrão (min/peça)": numero_limpo(tempo_unitario_valor),
                "Tempo Programado (min)": numero_limpo(tempo_programado_valor),
                "Data Realizada": data_br(data_real_valor),
                "Data Realizada Valor": data_real,
                "Qtde Realizada": numero_limpo(qtde_real),
                "Início Real": texto_limpo(valor_celula(ws_valores, linha_excel, col, "inicio_real")),
                "Fim Real": texto_limpo(valor_celula(ws_valores, linha_excel, col, "fim_real")),
                "Tempo Realizado (min)": numero_limpo(valor_celula(ws_valores, linha_excel, col, "tempo_realizado")),
                "Refugo/Retrabalho": numero_limpo(valor_celula(ws_valores, linha_excel, col, "refugo_retrabalho")),
                "Situação": "Realizado" if data_real is not None or qtde_real not in (None, "") else "Pendente",
            }
        )
    wb.close()
    wb_valores.close()
    return pd.DataFrame(linhas), col


@st.cache_data(show_spinner=False)
def carregar_dados_cacheado(
    caminho_texto: str,
    nome_aba: str,
    modificacao_ns: int,
    tamanho_arquivo: int,
) -> tuple[pd.DataFrame, dict[str, int]]:
    """
    Mantém a planilha em memória entre as interações do Streamlit.

    Os parâmetros de modificação e tamanho fazem parte da chave do cache.
    Quando o Excel é salvo ou substituído, o cache é renovado automaticamente.
    """
    del modificacao_ns, tamanho_arquivo
    return carregar_dados(Path(caminho_texto), nome_aba)


def obter_dados_cacheados(caminho: Path, nome_aba: str) -> tuple[pd.DataFrame, dict[str, int]]:
    estatistica = caminho.stat()
    return carregar_dados_cacheado(
        str(caminho.resolve()),
        nome_aba,
        estatistica.st_mtime_ns,
        estatistica.st_size,
    )


def salvar_realizacao(
    caminho: Path,
    nome_aba: str,
    linhas_excel: list[int],
    data_realizada: date,
    preencher_qtde: bool,
    atualizar_status: bool,
    criar_backup: bool,
) -> Path | None:
    if criar_backup:
        backup = caminho.with_name(f"{caminho.stem}_backup_{datetime.now():%Y%m%d_%H%M%S}{caminho.suffix}")
        shutil.copy2(caminho, backup)
    else:
        backup = None

    wb = load_workbook(caminho, data_only=False)
    ws = wb[nome_aba]
    col = localizar_colunas(ws)

    data_excel = datetime.combine(data_realizada, time.min)
    letra_qtd_prog = get_column_letter(col["qtde_programada"])
    letra_qtd_real = get_column_letter(col["qtde_realizada"])

    for linha in linhas_excel:
        cel_data = ws.cell(linha, col["data_realizada"])
        cel_data.value = data_excel
        cel_data.number_format = "dd/mm/yyyy"

        if preencher_qtde:
            qtde_programada = ws.cell(linha, col["qtde_programada"]).value
            cel_qtd = ws.cell(linha, col["qtde_realizada"])
            cel_qtd.value = qtde_programada
            cel_qtd.number_format = "0"

        if "percentual" in col:
            ws.cell(linha, col["percentual"]).value = (
                f'=IF(OR(${letra_qtd_prog}{linha}="",${letra_qtd_real}{linha}=""),"",${letra_qtd_real}{linha}/${letra_qtd_prog}{linha})'
            )
            ws.cell(linha, col["percentual"]).number_format = "0%"

        if atualizar_status and "status" in col:
            ws.cell(linha, col["status"]).value = "Concluído"

    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass

    wb.save(caminho)
    wb.close()
    return backup


# -----------------------------------------------------------------------------
# Leitura e gravação das paradas
# -----------------------------------------------------------------------------
def dataframe_paradas_vazio() -> pd.DataFrame:
    return pd.DataFrame(
        columns=[
            "Linha Excel",
            "Data Parada",
            "Data Parada Valor",
            "Setor",
            "Máquina/Posto",
            "Motivo da Parada",
            "Tempo Total Parada (min)",
            "Observação",
            "Registrado Em",
            "Usuário",
            "Editado Em",
            "Editado Por",
        ]
    )


def garantir_colunas_dataframe_paradas(df: pd.DataFrame) -> pd.DataFrame:
    """Garante o esquema atual mesmo quando o cache ou a planilha são de versão anterior."""
    base = df.copy() if isinstance(df, pd.DataFrame) else dataframe_paradas_vazio()
    colunas_padrao = dataframe_paradas_vazio().columns.tolist()
    for coluna in colunas_padrao:
        if coluna not in base.columns:
            if coluna == "Tempo Total Parada (min)":
                base[coluna] = 0.0
            else:
                base[coluna] = ""
    return base


def localizar_aba_paradas(wb, criar: bool = False):
    """Localiza a aba de paradas aceitando nomes como 6_Paradas ou 6 Paradas."""
    nomes_preferidos = {
        normalizar(NOME_ABA_PARADAS),
        normalizar("6 Paradas"),
        normalizar("Paradas"),
    }
    for nome in wb.sheetnames:
        if normalizar(nome) in nomes_preferidos:
            return wb[nome]
    if criar:
        return wb.create_sheet(NOME_ABA_PARADAS)
    return None


def aliases_colunas_paradas() -> dict[str, list[str]]:
    return {
        "Data Parada": ["data parada", "data da parada", "data"],
        "Setor": ["setor"],
        "Máquina/Posto": ["maquina posto", "maquina", "máquina/posto", "posto"],
        "Motivo da Parada": ["motivo da parada", "motivo parada", "motivo"],
        "Tempo Total Parada (min)": [
            "tempo total parada min",
            "tempo total de parada min",
            "tempo parada min",
            "tempo total parada",
            "tempo de parada",
            "tempo",
        ],
        "Observação": ["observacao", "observação", "detalhes", "comentario"],
        "Registrado Em": ["registrado em", "data registro", "data hora registro"],
        "Usuário": ["usuario", "usuário", "login"],
        "Editado Em": ["editado em", "ultima edicao em", "última edição em"],
        "Editado Por": ["editado por", "ultima edicao por", "última edição por"],
    }


def mapear_colunas_paradas(ws, criar_faltantes: bool = False) -> dict[str, int]:
    """Mapeia os cabeçalhos da aba e, ao salvar, acrescenta colunas ausentes."""
    cabecalhos_existentes = {
        normalizar(ws.cell(row=1, column=c).value): c
        for c in range(1, max(ws.max_column, 1) + 1)
        if texto_limpo(ws.cell(row=1, column=c).value)
    }
    mapa: dict[str, int] = {}
    aliases = aliases_colunas_paradas()

    for cabecalho_padrao in CABECALHOS_PARADAS:
        candidatos = [cabecalho_padrao, *aliases.get(cabecalho_padrao, [])]
        for candidato in candidatos:
            encontrado = cabecalhos_existentes.get(normalizar(candidato))
            if encontrado:
                mapa[cabecalho_padrao] = encontrado
                break

    if criar_faltantes:
        proxima_coluna = max(ws.max_column, 0) + 1
        if not cabecalhos_existentes:
            proxima_coluna = 1
        for cabecalho_padrao in CABECALHOS_PARADAS:
            if cabecalho_padrao not in mapa:
                while texto_limpo(ws.cell(row=1, column=proxima_coluna).value):
                    proxima_coluna += 1
                ws.cell(row=1, column=proxima_coluna).value = cabecalho_padrao
                mapa[cabecalho_padrao] = proxima_coluna
                proxima_coluna += 1

        preenchimento = PatternFill("solid", fgColor="0B2F6B")
        for cabecalho, coluna in mapa.items():
            celula = ws.cell(row=1, column=coluna)
            celula.value = cabecalho
            celula.fill = preenchimento
            celula.font = Font(color="FFFFFF", bold=True)
            celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(max(mapa.values()))}{max(ws.max_row, 1)}"
        larguras = {
            "Data Parada": 14,
            "Setor": 25,
            "Máquina/Posto": 28,
            "Motivo da Parada": 34,
            "Tempo Total Parada (min)": 22,
            "Observação": 40,
            "Registrado Em": 20,
            "Usuário": 14,
            "Editado Em": 20,
            "Editado Por": 14,
        }
        for cabecalho, coluna in mapa.items():
            ws.column_dimensions[get_column_letter(coluna)].width = larguras.get(cabecalho, 18)
        ws.row_dimensions[1].height = 32

    return mapa


def minutos_seguro(valor: Any) -> float:
    """Converte minutos numéricos, horário Excel ou texto HH:MM em minutos."""
    if valor in (None, ""):
        return 0.0
    if isinstance(valor, time):
        return float(valor.hour * 60 + valor.minute + valor.second / 60)
    if isinstance(valor, timedelta):
        return max(0.0, valor.total_seconds() / 60)
    if isinstance(valor, str):
        texto = valor.strip().lower().replace("h", ":").replace("min", "")
        if ":" in texto:
            partes = texto.split(":", 1)
            try:
                return max(0.0, float(partes[0] or 0) * 60 + float(partes[1] or 0))
            except Exception:
                pass
    return max(0.0, float_seguro(valor))


def formatar_tempo_minutos(valor: Any) -> str:
    minutos = int(round(minutos_seguro(valor)))
    horas, minutos_restantes = divmod(minutos, 60)
    if horas and minutos_restantes:
        return f"{horas}h {minutos_restantes}min"
    if horas:
        return f"{horas}h"
    return f"{minutos_restantes}min"


def carregar_paradas(caminho: Path) -> pd.DataFrame:
    wb = load_workbook(caminho, data_only=True, read_only=True)
    ws = localizar_aba_paradas(wb, criar=False)
    if ws is None:
        wb.close()
        return dataframe_paradas_vazio()

    colunas = mapear_colunas_paradas(ws, criar_faltantes=False)
    obrigatorias = [
        "Data Parada",
        "Setor",
        "Máquina/Posto",
        "Motivo da Parada",
        "Tempo Total Parada (min)",
    ]
    if any(c not in colunas for c in obrigatorias):
        wb.close()
        return dataframe_paradas_vazio()

    registros: list[dict[str, Any]] = []
    for linha in range(2, ws.max_row + 1):
        def valor(cabecalho: str) -> Any:
            coluna = colunas.get(cabecalho)
            return ws.cell(row=linha, column=coluna).value if coluna else ""

        data_valor = valor("Data Parada")
        setor = texto_limpo(valor("Setor"))
        maquina = texto_limpo(valor("Máquina/Posto"))
        motivo = texto_limpo(valor("Motivo da Parada"))
        tempo_min = minutos_seguro(valor("Tempo Total Parada (min)"))
        if not any([data_valor, setor, maquina, motivo, tempo_min]):
            continue

        data_parada = valor_para_data(data_valor)
        registros.append(
            {
                "Linha Excel": linha,
                "Data Parada": data_br(data_valor),
                "Data Parada Valor": data_parada,
                "Setor": setor or "Sem setor",
                "Máquina/Posto": maquina or "Sem máquina informada",
                "Motivo da Parada": motivo or "Sem motivo informado",
                "Tempo Total Parada (min)": tempo_min,
                "Observação": texto_limpo(valor("Observação")),
                "Registrado Em": valor("Registrado Em"),
                "Usuário": texto_limpo(valor("Usuário")),
                "Editado Em": valor("Editado Em"),
                "Editado Por": texto_limpo(valor("Editado Por")),
            }
        )

    wb.close()
    return pd.DataFrame(registros) if registros else dataframe_paradas_vazio()


@st.cache_data(show_spinner=False)
def carregar_paradas_cacheado(
    caminho_texto: str,
    modificacao_ns: int,
    tamanho_arquivo: int,
    versao_schema: str,
) -> pd.DataFrame:
    # A versão do esquema participa da chave do cache. Assim, novas colunas não
    # reutilizam DataFrames antigos armazenados pelo Streamlit.
    del modificacao_ns, tamanho_arquivo, versao_schema
    return garantir_colunas_dataframe_paradas(carregar_paradas(Path(caminho_texto)))


def obter_paradas_cacheadas(caminho: Path) -> pd.DataFrame:
    estatistica = caminho.stat()
    paradas = carregar_paradas_cacheado(
        str(caminho.resolve()),
        estatistica.st_mtime_ns,
        estatistica.st_size,
        VERSAO_SCHEMA_PARADAS,
    )
    return garantir_colunas_dataframe_paradas(paradas)


def salvar_parada(
    caminho: Path,
    data_parada: date,
    setor: str,
    maquina: str,
    motivo: str,
    tempo_total_min: float,
    observacao: str,
    criar_backup: bool,
) -> Path | None:
    if criar_backup:
        backup = caminho.with_name(
            f"{caminho.stem}_backup_paradas_{datetime.now():%Y%m%d_%H%M%S}{caminho.suffix}"
        )
        shutil.copy2(caminho, backup)
    else:
        backup = None

    wb = load_workbook(caminho, data_only=False)
    ws = localizar_aba_paradas(wb, criar=True)
    colunas = mapear_colunas_paradas(ws, criar_faltantes=True)
    linha = max(ws.max_row + 1, 2)

    valores = {
        "Data Parada": datetime.combine(data_parada, time.min),
        "Setor": setor.strip(),
        "Máquina/Posto": maquina.strip(),
        "Motivo da Parada": motivo.strip(),
        "Tempo Total Parada (min)": float(tempo_total_min),
        "Observação": observacao.strip(),
        "Registrado Em": datetime.now(),
        "Usuário": LOGIN_APONTAMENTO,
    }
    for cabecalho, valor in valores.items():
        ws.cell(row=linha, column=colunas[cabecalho]).value = valor

    ws.cell(row=linha, column=colunas["Data Parada"]).number_format = "dd/mm/yyyy"
    ws.cell(row=linha, column=colunas["Tempo Total Parada (min)"]).number_format = "0"
    ws.cell(row=linha, column=colunas["Registrado Em"]).number_format = "dd/mm/yyyy hh:mm"
    for coluna in colunas.values():
        ws.cell(row=linha, column=coluna).alignment = Alignment(vertical="top", wrap_text=True)
    ws.auto_filter.ref = f"A1:{get_column_letter(max(colunas.values()))}{linha}"

    wb.save(caminho)
    wb.close()
    return backup


def editar_parada(
    caminho: Path,
    linha_excel: int,
    data_parada: date,
    setor: str,
    maquina: str,
    motivo: str,
    tempo_total_min: float,
    observacao: str,
    criar_backup: bool = False,
) -> Path | None:
    """Atualiza uma parada existente na mesma linha da aba 6_Paradas."""
    if int(linha_excel) < 2:
        raise ValueError("Linha inválida para edição.")

    if criar_backup:
        backup = caminho.with_name(
            f"{caminho.stem}_backup_edicao_parada_{datetime.now():%Y%m%d_%H%M%S}{caminho.suffix}"
        )
        shutil.copy2(caminho, backup)
    else:
        backup = None

    wb = load_workbook(caminho, data_only=False)
    ws = localizar_aba_paradas(wb, criar=False)
    if ws is None:
        wb.close()
        raise ValueError(f"A aba {NOME_ABA_PARADAS} não foi encontrada.")

    colunas = mapear_colunas_paradas(ws, criar_faltantes=True)
    linha = int(linha_excel)
    if linha > ws.max_row:
        wb.close()
        raise ValueError("O registro selecionado não existe mais na planilha.")

    campos_identificacao = [
        "Data Parada",
        "Setor",
        "Máquina/Posto",
        "Motivo da Parada",
        "Tempo Total Parada (min)",
    ]
    if not any(
        texto_limpo(ws.cell(row=linha, column=colunas[campo]).value)
        for campo in campos_identificacao
    ):
        wb.close()
        raise ValueError("O registro selecionado está vazio ou foi removido.")

    valores = {
        "Data Parada": datetime.combine(data_parada, time.min),
        "Setor": setor.strip(),
        "Máquina/Posto": maquina.strip(),
        "Motivo da Parada": motivo.strip(),
        "Tempo Total Parada (min)": float(tempo_total_min),
        "Observação": observacao.strip(),
        "Editado Em": datetime.now(),
        "Editado Por": LOGIN_APONTAMENTO,
    }
    for cabecalho, valor in valores.items():
        ws.cell(row=linha, column=colunas[cabecalho]).value = valor

    ws.cell(row=linha, column=colunas["Data Parada"]).number_format = "dd/mm/yyyy"
    ws.cell(row=linha, column=colunas["Tempo Total Parada (min)"]).number_format = "0"
    ws.cell(row=linha, column=colunas["Editado Em"]).number_format = "dd/mm/yyyy hh:mm"
    for coluna in colunas.values():
        ws.cell(row=linha, column=coluna).alignment = Alignment(vertical="top", wrap_text=True)
    ws.auto_filter.ref = f"A1:{get_column_letter(max(colunas.values()))}{max(ws.max_row, 1)}"

    wb.save(caminho)
    wb.close()
    return backup


# -----------------------------------------------------------------------------
# Tela de apontamento
# -----------------------------------------------------------------------------
def filtrar_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    st.sidebar.header("Filtros do apontamento")

    # As opções de setor agora vêm da base completa, e não apenas do resultado
    # já filtrado. Isso evita sumir setores como "6 PINTURA UV" quando outros
    # filtros estiverem ativos ou quando o setor aparecer mais abaixo na planilha.
    todos_setores = opcoes_unicas(df, "Setor")
    todas_maquinas = opcoes_unicas(df, "Máquina/Posto")

    with st.sidebar.expander("Ver setores encontrados"):
        st.caption(f"{len(todos_setores)} setores encontrados na planilha carregada.")
        if todos_setores:
            st.write(", ".join(todos_setores))
        else:
            st.write("Nenhum setor encontrado.")

    situacao = st.sidebar.radio("Situação", ["Pendentes", "Realizados", "Todos"], horizontal=True)
    filtrado = df.copy()
    if situacao == "Pendentes":
        filtrado = filtrado[filtrado["Situação"] == "Pendente"]
    elif situacao == "Realizados":
        filtrado = filtrado[filtrado["Situação"] == "Realizado"]

    busca_setor = st.sidebar.text_input("Buscar setor digitando", placeholder="Ex.: PINTURA, UV, 6")
    setores_para_exibir = todos_setores
    if busca_setor:
        busca_norm = normalizar(busca_setor)
        setores_para_exibir = [s for s in todos_setores if busca_norm in normalizar(s)]
        filtrado = filtrado[filtrado["Setor"].astype(str).apply(lambda x: busca_norm in normalizar(x))]

    setor_sel = st.sidebar.multiselect(
        "Filtrar por setor",
        setores_para_exibir,
        placeholder="Selecione um ou mais setores",
    )
    if setor_sel:
        filtrado = filtrado[filtrado["Setor"].isin(setor_sel)]

    busca_lote = st.sidebar.text_input("Buscar lote", placeholder="Ex.: 10176")
    if busca_lote:
        filtrado = filtrado[filtrado["OP/Lote"].str.contains(busca_lote, case=False, na=False)]

    # A lista de máquinas também vem da base completa para não desaparecer ao trocar filtros.
    maquinas_para_exibir = todas_maquinas
    if setor_sel:
        maquinas_para_exibir = opcoes_unicas(df[df["Setor"].isin(setor_sel)], "Máquina/Posto")
    elif busca_setor:
        busca_norm = normalizar(busca_setor)
        maquinas_para_exibir = opcoes_unicas(df[df["Setor"].astype(str).apply(lambda x: busca_norm in normalizar(x))], "Máquina/Posto")

    maquinas_sel = st.sidebar.multiselect("Máquina/Posto", maquinas_para_exibir)
    if maquinas_sel:
        filtrado = filtrado[filtrado["Máquina/Posto"].isin(maquinas_sel)]

    texto = st.sidebar.text_input("Buscar peça/produto/cliente", placeholder="Digite parte do texto")
    if texto:
        campos = ["Cliente/Pedido", "Produto/Equipamento", "Código Peça", "Descrição Peça", "Operação"]
        mascara = pd.Series(False, index=filtrado.index)
        for campo in campos:
            mascara = mascara | filtrado[campo].astype(str).str.contains(texto, case=False, na=False)
        filtrado = filtrado[mascara]

    return filtrado


def renderizar_apontamento(df: pd.DataFrame, caminho: Path, nome_aba: str) -> None:
    st.title("Acompanhamento de Produção - PPCP")
    st.caption(f"Sistema versão {VERSAO_APP}")
    st.caption(
        "Modo rápido: marque quantas linhas precisar e clique em Salvar. "
        "A página não é recarregada a cada caixa selecionada."
    )

    total = len(df)
    pendentes = int((df["Situação"] == "Pendente").sum())
    realizados = int((df["Situação"] == "Realizado").sum())

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total de linhas", total)
    c2.metric("Pendentes", pendentes)
    c3.metric("Realizados", realizados)
    c4.metric("Arquivo", caminho.name)

    filtrado = filtrar_dataframe(df).copy()

    if filtrado.empty:
        st.warning("Nenhum item encontrado para os filtros selecionados.")
        return

    # Paginação reduz a quantidade de células que o navegador precisa renderizar.
    col_pagina1, col_pagina2 = st.columns([1, 1])
    opcoes_por_pagina = [100, 200, 300, 500, 1000]
    itens_por_pagina = col_pagina1.selectbox(
        "Itens por página",
        opcoes_por_pagina,
        index=1,
        help="Quanto menor este valor, mais rápida fica a seleção das linhas.",
        key="apontamento_itens_por_pagina",
    )

    total_paginas = max(1, (len(filtrado) + itens_por_pagina - 1) // itens_por_pagina)
    pagina_guardada = int(st.session_state.get("apontamento_pagina", 1))
    if pagina_guardada < 1 or pagina_guardada > total_paginas:
        st.session_state["apontamento_pagina"] = 1
    pagina_atual = col_pagina2.number_input(
        "Página",
        min_value=1,
        max_value=total_paginas,
        step=1,
        key="apontamento_pagina",
    )

    inicio = (int(pagina_atual) - 1) * itens_por_pagina
    fim = min(inicio + itens_por_pagina, len(filtrado))
    pagina_df = filtrado.iloc[inicio:fim].copy().reset_index(drop=True)

    st.info(
        f"Exibindo itens **{inicio + 1} a {fim}** de **{len(filtrado)}** filtrados "
        f"— página **{int(pagina_atual)} de {total_paginas}**."
    )

    mapa_linhas_excel_pagina = pagina_df["Linha Excel"].astype(int).reset_index(drop=True)
    todas_linhas_filtradas = filtrado["Linha Excel"].astype(int).tolist()

    colunas_visiveis = [
        "Selecionar",
        "Data Programada",
        "Data Entrega",
        "OP/Lote",
        "Setor",
        "Máquina/Posto",
        "Produto/Equipamento",
        "Código Peça",
        "Descrição Peça",
        "Qtde Programada",
        "Data Realizada",
        "Qtde Realizada",
        "Situação",
    ]
    tabela = pagina_df[[c for c in colunas_visiveis if c in pagina_df.columns]].copy()
    tabela["Selecionar"] = False

    # Dentro do formulário, as marcações ficam no navegador e só são enviadas
    # quando o usuário clica em Salvar. Isso elimina uma execução completa por clique.
    with st.form("form_apontamento_rapido", clear_on_submit=False):
        st.subheader("Lançamento rápido")
        col_data, col_qtd, col_status, col_backup = st.columns([1, 1.2, 1.2, 1])
        data_realizada = col_data.date_input(
            "Data realizada",
            value=date.today(),
            format="DD/MM/YYYY",
        )
        preencher_qtde = col_qtd.checkbox(
            "Qtde realizada = Qtde programada",
            value=True,
        )
        atualizar_status = col_status.checkbox(
            "Marcar status como Concluído",
            value=True,
        )
        criar_backup = col_backup.checkbox(
            "Criar backup",
            value=False,
        )

        salvar_todos_filtrados = st.checkbox(
            "Salvar todos os itens filtrados, incluindo as outras páginas",
            value=False,
            help=(
                "Quando marcado, o sistema ignora as caixas individuais e salva "
                "todos os itens encontrados pelos filtros atuais."
            ),
        )

        st.caption(
            ""
            ""
        )

        editada = st.data_editor(
            tabela,
            use_container_width=True,
            hide_index=True,
            height=520,
            disabled=[c for c in tabela.columns if c != "Selecionar"],
            column_config={
                "Selecionar": st.column_config.CheckboxColumn("Selecionar"),
                "Descrição Peça": st.column_config.TextColumn("Descrição Peça", width="large"),
                "Produto/Equipamento": st.column_config.TextColumn("Produto/Equipamento", width="medium"),
            },
        )

        salvar = st.form_submit_button(
            "Salvar apontamentos selecionados",
            type="primary",
            use_container_width=True,
        )

    if salvar:
        if salvar_todos_filtrados:
            linhas_selecionadas = todas_linhas_filtradas
        else:
            indices_selecionados = editada.index[editada["Selecionar"]].tolist()
            linhas_selecionadas = (
                mapa_linhas_excel_pagina.loc[indices_selecionados].astype(int).tolist()
                if indices_selecionados
                else []
            )

        if not linhas_selecionadas:
            st.warning("Nenhuma linha foi selecionada. Marque pelo menos um item e tente novamente.")
        else:
            try:
                with st.spinner(f"Salvando {len(linhas_selecionadas)} apontamento(s)..."):
                    backup = salvar_realizacao(
                        caminho=caminho,
                        nome_aba=nome_aba,
                        linhas_excel=linhas_selecionadas,
                        data_realizada=data_realizada,
                        preencher_qtde=preencher_qtde,
                        atualizar_status=atualizar_status,
                        criar_backup=criar_backup,
                    )
                # Força a próxima leitura a refletir o Excel recém-salvo.
                carregar_dados_cacheado.clear()
                st.success(f"Salvo com sucesso em {len(linhas_selecionadas)} linha(s).")
                if backup:
                    st.info(f"Backup criado: {backup.name}")
                st.rerun()
            except PermissionError:
                st.error("Não consegui salvar. Feche a planilha no Excel e tente novamente.")
            except Exception as exc:
                st.error(f"Erro ao salvar: {exc}")

    st.divider()
    with open(caminho, "rb") as f:
        st.download_button(
            "Baixar planilha atualizada",
            data=f.read(),
            file_name=caminho.name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    with st.expander("Como usar o modo rápido"):
        st.markdown(
            """
1. Use os filtros de setor, lote, máquina ou texto.
2. Escolha quantos itens deseja exibir por página.
3. Marque as linhas necessárias sem esperar recarregamentos.
4. Para salvar todo o resultado do filtro, marque **Salvar todos os itens filtrados**.
5. Clique em **Salvar apontamentos selecionados**.

O formulário envia todas as marcações de uma vez. Por isso, a quantidade selecionada não é recalculada a cada clique.
            """
        )

# -----------------------------------------------------------------------------
# Tela de produção por setor
# -----------------------------------------------------------------------------
def preparar_base_producao_setor(df: pd.DataFrame) -> pd.DataFrame:
    """
    Prepara os campos usados na comparação por setor.

    Regra do realizado:
    - usa a Qtde Realizada quando ela foi informada;
    - para apontamentos antigos com Data Realizada preenchida e quantidade vazia,
      considera a Qtde Programada como realizada.
    """
    base = df.copy()
    base["Setor"] = base["Setor"].fillna("").astype(str).str.strip().replace("", "Sem setor")
    base["Qtde Programada Num"] = to_num(base["Qtde Programada"]).clip(lower=0)

    qtde_real_original = base["Qtde Realizada"].copy()
    base["Qtde Realizada Num"] = to_num(qtde_real_original).clip(lower=0)
    base["Data Programada DT"] = pd.to_datetime(base["Data Programada Valor"], errors="coerce")
    base["Data Realizada DT"] = pd.to_datetime(base["Data Realizada Valor"], errors="coerce")

    texto_qtde_real = qtde_real_original.fillna("").astype(str).str.strip()
    # Algumas planilhas antigas ou fórmulas retornam 0 em vez de célula vazia.
    # Quando existe Data Realizada e a quantidade é vazia ou zero, considera-se
    # a Qtde Programada como apontada. Isso evita que um setor concluído apareça
    # com apontamento 0 apenas por causa da forma como o Excel gravou a célula.
    base["Qtde Realizada Vazia?"] = texto_qtde_real.eq("")
    base["Qtde Realizada Ausente ou Zero?"] = (
        base["Qtde Realizada Vazia?"] | (base["Qtde Realizada Num"] <= 0)
    )
    base["Realizado por fallback?"] = (
        base["Data Realizada DT"].notna()
        & base["Qtde Realizada Ausente ou Zero?"]
        & (base["Qtde Programada Num"] > 0)
    )
    base["Qtde Realizada Considerada Num"] = base["Qtde Realizada Num"]
    base.loc[
        base["Realizado por fallback?"],
        "Qtde Realizada Considerada Num",
    ] = base.loc[
        base["Realizado por fallback?"],
        "Qtde Programada Num",
    ]
    return base


def intervalo_datas_disponiveis(base: pd.DataFrame) -> tuple[date, date] | None:
    datas = pd.concat(
        [base["Data Programada DT"], base["Data Realizada DT"]],
        ignore_index=True,
    ).dropna()
    if datas.empty:
        return None
    return datas.min().date(), datas.max().date()


def escolher_filtros_producao_setor(
    base: pd.DataFrame,
) -> tuple[list[str], date | None, date | None, str]:
    st.subheader("Filtros")
    setores = opcoes_unicas(base, "Setor")

    col_setor, col_visao = st.columns([1.4, 1])
    setores_sel = col_setor.multiselect(
        "Setor",
        setores,
        placeholder="Vazio = todos os setores",
        key="producao_setores",
    )
    modo = col_visao.radio(
        "Visualização",
        ["Dia específico", "Período personalizado", "Período completo"],
        horizontal=False,
        key="producao_modo_data",
    )

    intervalo = intervalo_datas_disponiveis(base)
    if intervalo is None:
        st.warning("Não existem datas programadas ou realizadas válidas na base.")
        return setores_sel, None, None, modo

    data_min, data_max = intervalo
    inicio: date
    fim: date

    if modo == "Dia específico":
        hoje = date.today()
        data_padrao = hoje if data_min <= hoje <= data_max else data_max
        dia = st.date_input(
            "Dia analisado",
            value=data_padrao,
            min_value=data_min,
            max_value=data_max,
            format="DD/MM/YYYY",
            key="producao_dia",
        )
        inicio = dia
        fim = dia
    elif modo == "Período personalizado":
        periodo = st.date_input(
            "Período analisado",
            value=(data_min, data_max),
            min_value=data_min,
            max_value=data_max,
            format="DD/MM/YYYY",
            key="producao_periodo",
        )
        if isinstance(periodo, (tuple, list)) and len(periodo) == 2:
            inicio, fim = periodo
        elif isinstance(periodo, (tuple, list)) and len(periodo) == 1:
            inicio = fim = periodo[0]
        else:
            inicio = fim = periodo
    else:
        inicio, fim = data_min, data_max
        st.caption(f"Período completo da base: {inicio:%d/%m/%Y} a {fim:%d/%m/%Y}")

    return setores_sel, inicio, fim, modo


def entre_datas(serie: pd.Series, inicio: date, fim: date) -> pd.Series:
    return serie.notna() & (serie.dt.date >= inicio) & (serie.dt.date <= fim)


def montar_resumo_producao_setor(
    base: pd.DataFrame,
    setores_sel: list[str],
    inicio: date,
    fim: date,
) -> pd.DataFrame:
    """
    Compara o programado e o realizado das MESMAS linhas programadas no período.

    Exemplo: ao selecionar 17/07/2026, primeiro são localizadas todas as linhas
    cuja Data Programada é 17/07/2026. Depois:
    - Programado = soma da Qtde Programada dessas linhas;
    - Realizado = soma da Qtde Realizada apontada nessas mesmas linhas.
    """
    setores_base = setores_sel if setores_sel else opcoes_unicas(base, "Setor")
    base_setores = base[base["Setor"].isin(setores_base)].copy()
    base_periodo = base_setores[
        entre_datas(base_setores["Data Programada DT"], inicio, fim)
    ].copy()

    programado = (
        base_periodo.groupby("Setor")["Qtde Programada Num"]
        .sum()
        .rename("Programado")
    )
    realizado = (
        base_periodo.groupby("Setor")["Qtde Realizada Considerada Num"]
        .sum()
        .rename("Realizado")
    )

    resumo = pd.DataFrame({"Setor": setores_base}).set_index("Setor")
    resumo = resumo.join(programado).join(realizado).fillna(0).reset_index()
    resumo["Saldo"] = resumo["Programado"] - resumo["Realizado"]
    resumo["% Realizado"] = 0.0
    mask_programado = resumo["Programado"] > 0
    resumo.loc[mask_programado, "% Realizado"] = (
        resumo.loc[mask_programado, "Realizado"]
        / resumo.loc[mask_programado, "Programado"]
        * 100
    )

    def situacao(linha: pd.Series) -> str:
        programado_val = float(linha["Programado"])
        realizado_val = float(linha["Realizado"])
        percentual = float(linha["% Realizado"])
        if programado_val <= 0 and realizado_val <= 0:
            return "Sem programação"
        if realizado_val >= programado_val and programado_val > 0:
            return "Programado atingido"
        if percentual >= 90:
            return "Próximo do programado"
        return "Abaixo do programado"

    resumo["Situação"] = resumo.apply(situacao, axis=1)
    return resumo.sort_values(["Programado", "Realizado", "Setor"], ascending=[False, False, True])


def montar_producao_diaria(
    base: pd.DataFrame,
    setores_sel: list[str],
    inicio: date,
    fim: date,
) -> pd.DataFrame:
    """
    Monta o resultado diário pela Data Programada.

    Em cada dia, o realizado representa quanto foi apontado nas linhas que estavam
    programadas para aquele mesmo dia, mesmo que a Data Realizada seja diferente.
    """
    setores_base = setores_sel if setores_sel else opcoes_unicas(base, "Setor")
    base_setores = base[base["Setor"].isin(setores_base)].copy()
    periodo = base_setores[
        entre_datas(base_setores["Data Programada DT"], inicio, fim)
    ].copy()

    programado_dia = (
        periodo.groupby(periodo["Data Programada DT"].dt.date)["Qtde Programada Num"]
        .sum()
        .rename("Programado")
    )
    realizado_dia = (
        periodo.groupby(periodo["Data Programada DT"].dt.date)["Qtde Realizada Considerada Num"]
        .sum()
        .rename("Realizado")
    )
    programado_dia.index = pd.to_datetime(programado_dia.index)
    realizado_dia.index = pd.to_datetime(realizado_dia.index)

    calendario = pd.date_range(inicio, fim, freq="D")
    diario = pd.DataFrame(index=calendario)
    diario.index.name = "Data"
    diario = diario.join(programado_dia.rename_axis("Data"), how="left")
    diario = diario.join(realizado_dia.rename_axis("Data"), how="left")
    diario = diario.fillna(0).reset_index()
    diario["Saldo"] = diario["Programado"] - diario["Realizado"]
    diario["% Realizado"] = 0.0
    mask = diario["Programado"] > 0
    diario.loc[mask, "% Realizado"] = (
        diario.loc[mask, "Realizado"] / diario.loc[mask, "Programado"] * 100
    )
    return diario


def montar_detalhamento_periodo(
    base: pd.DataFrame,
    setores_sel: list[str],
    inicio: date,
    fim: date,
    tipo: str,
) -> pd.DataFrame:
    """
    Detalha as linhas programadas no período.

    As duas abas usam a Data Programada para garantir que o programado e o realizado
    sejam auditáveis sobre exatamente o mesmo conjunto de linhas.
    """
    setores_base = setores_sel if setores_sel else opcoes_unicas(base, "Setor")
    dados = base[base["Setor"].isin(setores_base)].copy()
    dados = dados[entre_datas(dados["Data Programada DT"], inicio, fim)]
    dados = dados.sort_values(["Data Programada DT", "Setor", "OP/Lote"])

    if tipo == "programado":
        return dados[
            [
                "Data Programada",
                "Setor",
                "OP/Lote",
                "Produto/Equipamento",
                "Código Peça",
                "Descrição Peça",
                "Qtde Programada",
            ]
        ]

    detalhe = dados[
        [
            "Data Programada",
            "Data Realizada",
            "Setor",
            "OP/Lote",
            "Produto/Equipamento",
            "Código Peça",
            "Descrição Peça",
            "Qtde Programada",
            "Qtde Realizada",
            "Qtde Realizada Considerada Num",
            "Realizado por fallback?",
        ]
    ].copy()
    detalhe = detalhe.rename(
        columns={
            "Qtde Realizada Considerada Num": "Qtde Realizada Considerada",
            "Realizado por fallback?": "Quantidade assumida pela data?",
        }
    )
    return detalhe


# Tela Produção por Setor removida na versão 5.19.


# -----------------------------------------------------------------------------
# Tela pública: Programado x apontado por dia, equipamento e unidade de medição
# -----------------------------------------------------------------------------
# Unidade produtiva definida conforme o cadastro fornecido pelo PPCP.
# A identificação principal é feita pelo setor; o código/nome da máquina é usado
# como apoio para bases antigas que possuem nomes de setor diferentes.
METRICA_POR_SETOR = {
    "pintura p u": "uni",
    "usinagem": "uni",
    "usinagem automatizada": "uni",
    "seccionadora 1": "M²",
    "seccionadora 2": "M²",
    "montagem cadeira": "uni",
    "montagem de cadeiras": "uni",
    "cola de espuma": "uni",
    "tapecaria": "uni",
    "furadeiras": "uni",
    "coladeiras": "m linear",
    "cnc": "uni",
    "copiadora": "uni",
    "fabricacao embalagem": "uni",
    "cartonagem": "uni",
    "pintura uv": "M²",
    "uv": "M²",
    "costura": "uni",
    "corte de tecido": "uni",
    "embalagem": "uni",
    "vidro": "uni",
    "serra fita": "uni",
    "pade velox": "uni",
}

METRICA_POR_CODIGO_EQUIPAMENTO = {
    "50": "uni",
    "58": "uni",
    "96": "uni",
    "123": "uni",
    "10": "M²",
    "11": "M²",
    "59": "uni",
    "60": "uni",
    "62": "uni",
    "20": "uni",
    "21": "uni",
    "22": "uni",
    "65": "uni",
    "23": "m linear",
    "24": "m linear",
    "57": "m linear",
    "122": "m linear",
    "56": "uni",
    "75": "uni",
    "120": "uni",
    "54": "M²",
    "64": "uni",
    "63": "uni",
    "91": "uni",
}


def remover_codigo_inicial(texto: Any) -> str:
    """Remove códigos numéricos iniciais como '46 SECCIONADORA 1'."""
    valor = normalizar(texto)
    return re.sub(r"^\d+\s+", "", valor).strip()


def codigo_inicial(texto: Any) -> str:
    encontrado = re.match(r"^\s*(\d{1,4})\b", texto_limpo(texto))
    return encontrado.group(1) if encontrado else ""


def unidade_metrica_linha(setor: Any, maquina: Any) -> tuple[str, str]:
    """Retorna unidade e origem da classificação da métrica."""
    setor_sem_codigo = remover_codigo_inicial(setor)
    maquina_sem_codigo = remover_codigo_inicial(maquina)

    for chave in (setor_sem_codigo, maquina_sem_codigo):
        if chave in METRICA_POR_SETOR:
            return METRICA_POR_SETOR[chave], "Cadastro do setor"

    # Compatibilidade com descrições como SECCIONADORA 10, COLADEIRA 23 e UV.
    combinado = f"{setor_sem_codigo} {maquina_sem_codigo}".strip()
    if "seccionadora" in combinado or combinado == "uv" or "pintura uv" in combinado:
        return "M²", "Regra do setor"
    if "coladeira" in combinado:
        return "m linear", "Regra do setor"

    for valor in (maquina, setor):
        codigo = codigo_inicial(valor)
        if codigo in METRICA_POR_CODIGO_EQUIPAMENTO:
            return METRICA_POR_CODIGO_EQUIPAMENTO[codigo], "Código do equipamento"

    # Os setores não listados permanecem em unidade, mas são identificados na tela.
    return "uni", "Padrão não cadastrado"


def numero_decimal_flexivel(texto: Any) -> float | None:
    valor = texto_limpo(texto).replace(" ", "").replace(",", ".")
    try:
        numero = float(valor)
    except Exception:
        return None
    return numero if numero > 0 else None


def dimensoes_texto_metros(texto: Any) -> tuple[float | None, float | None]:
    """Extrai comprimento e largura em metros de medida ou descrição.

    Exemplos aceitos:
    - 1,605/0,905/0,018 (medidas já em metros);
    - 1605X905X18, 751 X 70 X 12 e 800X50 (medidas em milímetros);
    - 1,20X0,80 (medidas em metros).
    """
    original = texto_limpo(texto).upper()
    if not original:
        return None, None

    # O campo Medida das fichas usa normalmente comprimento/largura/espessura.
    barra = re.search(
        r"(?<!\d)(\d+(?:[\.,]\d+)?)\s*/\s*(\d+(?:[\.,]\d+)?)(?:\s*/\s*\d+(?:[\.,]\d+)?)?",
        original,
    )
    if barra:
        comprimento = numero_decimal_flexivel(barra.group(1))
        largura = numero_decimal_flexivel(barra.group(2))
        if comprimento and largura:
            # Valores superiores a 20 dificilmente estão em metros nesta base.
            if comprimento > 20:
                comprimento /= 1000
            if largura > 20:
                largura /= 1000
            return comprimento, largura

    # Remove observações intermediárias como 430X(108)51X25 e 360X56(911)X25.
    sem_parenteses = re.sub(r"\([^)]*\)", "", original)
    vezes = re.search(
        r"(?<!\d)(\d+(?:[\.,]\d+)?)\s*[X×]\s*(\d+(?:[\.,]\d+)?)(?:\s*[X×]\s*\d+(?:[\.,]\d+)?)?",
        sem_parenteses,
    )
    if not vezes:
        return None, None

    comprimento = numero_decimal_flexivel(vezes.group(1))
    largura = numero_decimal_flexivel(vezes.group(2))
    if not comprimento or not largura:
        return None, None

    # Inteiros como 1605X905 são milímetros. Decimais pequenos como 1,20X0,80
    # já representam metros.
    if comprimento > 20:
        comprimento /= 1000
    if largura > 20:
        largura /= 1000
    return comprimento, largura


@st.cache_data(show_spinner=False)
def carregar_medidas_catalogo_cacheado(
    caminho_texto: str,
    modificacao_ns: int,
    tamanho_arquivo: int,
) -> dict[str, str]:
    """Cria um lookup Código Peça -> Medida usando o catálogo SQLite."""
    del modificacao_ns, tamanho_arquivo
    caminho = Path(caminho_texto)
    if not caminho.exists():
        return {}

    conexao = sqlite3.connect(caminho, timeout=5)
    try:
        existe = conexao.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='fichas_catalogo'"
        ).fetchone()
        if not existe:
            return {}
        linhas = conexao.execute(
            """
            SELECT codigo_peca, medida, COUNT(*) AS ocorrencias
            FROM fichas_catalogo
            WHERE TRIM(COALESCE(codigo_peca, '')) <> ''
              AND TRIM(COALESCE(medida, '')) <> ''
            GROUP BY codigo_peca, medida
            ORDER BY codigo_peca, ocorrencias DESC
            """
        ).fetchall()
    except sqlite3.Error:
        return {}
    finally:
        conexao.close()

    lookup: dict[str, str] = {}
    for codigo, medida, _ in linhas:
        chave = texto_limpo(codigo)
        if chave and chave not in lookup:
            lookup[chave] = texto_limpo(medida)
    return lookup


def obter_medidas_catalogo() -> dict[str, str]:
    caminho = caminho_catalogo_fichas()
    if not caminho.exists():
        return {}
    estatistica = caminho.stat()
    return carregar_medidas_catalogo_cacheado(
        str(caminho.resolve()),
        estatistica.st_mtime_ns,
        estatistica.st_size,
    )


def enriquecer_metricas_programacao(base: pd.DataFrame) -> pd.DataFrame:
    """Adiciona unidade, dimensões e valores métricos em cada linha."""
    resultado = base.copy()
    lookup_catalogo = obter_medidas_catalogo()

    descricoes = resultado["Descrição Peça"].fillna("").astype(str)
    codigos = resultado["Código Peça"].fillna("").astype(str).map(texto_limpo)

    dimensoes_descricao = {
        descricao: dimensoes_texto_metros(descricao)
        for descricao in descricoes.unique().tolist()
    }

    unidades: list[str] = []
    origens_unidade: list[str] = []
    comprimentos: list[float] = []
    larguras: list[float] = []
    origens_medida: list[str] = []
    medidas_catalogo: list[str] = []

    for setor, maquina, descricao, codigo in zip(
        resultado["Setor"],
        resultado["Máquina/Posto"],
        descricoes,
        codigos,
    ):
        unidade, origem_unidade = unidade_metrica_linha(setor, maquina)
        unidades.append(unidade)
        origens_unidade.append(origem_unidade)

        comprimento, largura = dimensoes_descricao.get(descricao, (None, None))
        origem_medida = "Descrição da peça" if comprimento and largura else ""
        medida_catalogo = lookup_catalogo.get(codigo, "")
        if (not comprimento or not largura) and medida_catalogo:
            comprimento, largura = dimensoes_texto_metros(medida_catalogo)
            if comprimento and largura:
                origem_medida = "Catálogo de fichas"

        comprimentos.append(float(comprimento or 0))
        larguras.append(float(largura or 0))
        origens_medida.append(origem_medida or "Não identificada")
        medidas_catalogo.append(medida_catalogo)

    resultado["Unidade Métrica"] = unidades
    resultado["Origem Unidade Métrica"] = origens_unidade
    resultado["Comprimento (m)"] = comprimentos
    resultado["Largura (m)"] = larguras
    resultado["Origem Medida"] = origens_medida
    resultado["Medida Catálogo"] = medidas_catalogo

    precisa_comprimento = resultado["Unidade Métrica"].isin(["M²", "m linear"])
    precisa_largura = resultado["Unidade Métrica"].eq("M²")
    resultado["Medida Válida?"] = True
    resultado.loc[
        precisa_comprimento & (resultado["Comprimento (m)"] <= 0),
        "Medida Válida?",
    ] = False
    resultado.loc[
        precisa_largura & (resultado["Largura (m)"] <= 0),
        "Medida Válida?",
    ] = False

    resultado["Fator Métrico Unitário"] = 1.0
    mask_m2 = resultado["Unidade Métrica"].eq("M²") & resultado["Medida Válida?"]
    mask_ml = resultado["Unidade Métrica"].eq("m linear") & resultado["Medida Válida?"]
    resultado.loc[mask_m2, "Fator Métrico Unitário"] = (
        resultado.loc[mask_m2, "Comprimento (m)"]
        * resultado.loc[mask_m2, "Largura (m)"]
    )
    resultado.loc[mask_ml, "Fator Métrico Unitário"] = resultado.loc[
        mask_ml, "Comprimento (m)"
    ]
    resultado.loc[precisa_comprimento & ~resultado["Medida Válida?"], "Fator Métrico Unitário"] = 0.0

    resultado["Métrica Programada Num"] = (
        resultado["Qtde Programada Num"].clip(lower=0)
        * resultado["Fator Métrico Unitário"]
    )
    resultado["Métrica Apontada Num"] = (
        resultado["Qtde Apontada Num"].clip(lower=0)
        * resultado["Fator Métrico Unitário"]
    )
    return resultado


def preparar_base_programado_equipamento(df: pd.DataFrame) -> pd.DataFrame:
    """Padroniza a base da comparação diária usando somente o apontamento manual."""
    base = preparar_base_producao_setor(df)
    base["Máquina/Posto"] = (
        base["Máquina/Posto"]
        .fillna("")
        .astype(str)
        .str.strip()
        .replace("", "Sem máquina/posto")
    )
    base["OP/Lote"] = base["OP/Lote"].fillna("").astype(str).str.strip()

    # Regra desta tela: apontado é exatamente a Qtde Realizada digitada na planilha.
    # Uma Data Realizada sem quantidade não transforma automaticamente o programado
    # em realizado, evitando aumentar o total apontado indevidamente.
    base["Qtde Apontada Num"] = base["Qtde Realizada Num"].clip(lower=0)
    return enriquecer_metricas_programacao(base)




def _codigo_op_relatorio(valor: Any) -> str:
    texto = texto_limpo(valor)
    if not texto:
        return ""
    encontrado = re.search(r"(?<!\d)(\d{8})(?!\d)", texto)
    if encontrado:
        return encontrado.group(1)
    somente_digitos = re.sub(r"\D", "", texto)
    return somente_digitos if len(somente_digitos) == 8 else ""


def _codigo_equipamento_relatorio(valor: Any) -> str:
    texto = texto_limpo(valor)
    if not texto:
        return ""
    encontrado = re.match(r"^\s*(\d{1,4})(?:\s|[-–—])", texto)
    return encontrado.group(1).lstrip("0") or "0" if encontrado else ""


def _assinatura_apontamento_relatorio(linha: pd.Series) -> tuple[Any, ...] | None:
    data_valor = pd.to_datetime(linha.get("Data Realizada DT"), errors="coerce")
    if pd.isna(data_valor):
        return None
    codigo_op = _codigo_op_relatorio(linha.get("OP/Lote"))
    codigo_equipamento = _codigo_equipamento_relatorio(linha.get("Máquina/Posto"))
    if not codigo_op or not codigo_equipamento:
        return None
    quantidade = pd.to_numeric(
        pd.Series([linha.get("Qtde Apontada Num", linha.get("Qtde Realizada Considerada Num", 0))]),
        errors="coerce",
    ).fillna(0).iloc[0]
    return (
        data_valor.date(),
        normalizar(linha.get("Setor")),
        codigo_equipamento,
        codigo_op,
        round(float(quantidade), 6),
    )


def data_hora_operacional(valor: Any) -> pd.Timestamp | pd.NaT:
    """Converte datas do SQLite aceitando ISO e formatos brasileiros antigos."""
    if valor is None:
        return pd.NaT
    if isinstance(valor, pd.Timestamp):
        convertido = valor
    elif isinstance(valor, datetime):
        convertido = pd.Timestamp(valor)
    else:
        texto = texto_limpo(valor)
        if not texto:
            return pd.NaT

        convertido = pd.NaT
        formatos = (
            "%Y-%m-%d %H:%M:%S.%f",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%dT%H:%M:%S.%f",
            "%Y-%m-%dT%H:%M:%S",
            "%Y-%m-%d",
            "%d/%m/%Y %H:%M:%S",
            "%d/%m/%Y %H:%M",
            "%d/%m/%Y",
        )
        for formato in formatos:
            try:
                convertido = pd.Timestamp(datetime.strptime(texto, formato))
                break
            except ValueError:
                continue

        if pd.isna(convertido):
            convertido = pd.to_datetime(texto, errors="coerce", dayfirst=True)

    if pd.isna(convertido):
        return pd.NaT
    try:
        if convertido.tzinfo is not None:
            convertido = convertido.tz_localize(None)
    except (TypeError, AttributeError):
        pass
    return pd.Timestamp(convertido)


def _diagnostico_operacional_vazio(caminho: Path) -> dict[str, Any]:
    return {
        "caminho": str(caminho),
        "total_fechados_banco": 0,
        "datas_validas": 0,
        "datas_invalidas": 0,
        "registros_periodo": 0,
        "quantidades_positivas": 0,
        "quantidades_zero": 0,
        "erro": "",
    }


def carregar_apontamentos_operacionais_programado(
    inicio: date,
    fim: date,
) -> pd.DataFrame:
    """
    Transforma ordens finalizadas do SQLite em linhas compatíveis com o relatório.

    A filtragem do período é feita no pandas, depois da conversão da data. Isso evita
    que registros antigos gravados como DD/MM/AAAA sejam descartados pelo date() do SQLite.
    """
    caminho = caminho_banco_operacional()
    diagnostico = _diagnostico_operacional_vazio(caminho)
    vazio = pd.DataFrame()
    vazio.attrs["diagnostico_operacional"] = diagnostico
    if not caminho.exists():
        diagnostico["erro"] = "Arquivo do banco operacional não encontrado."
        return vazio

    conexao: sqlite3.Connection | None = None
    try:
        # mode=ro impede que a tela de relatório crie silenciosamente um banco vazio.
        uri = caminho.resolve().as_uri() + "?mode=ro"
        conexao = sqlite3.connect(uri, uri=True, timeout=30)
        conexao.execute("PRAGMA busy_timeout = 30000")
        existe = conexao.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='ordens_operacionais'"
        ).fetchone()
        if not existe:
            diagnostico["erro"] = "A tabela ordens_operacionais não existe no banco consultado."
            return vazio

        colunas_existentes = {
            str(linha[1])
            for linha in conexao.execute("PRAGMA table_info(ordens_operacionais)").fetchall()
        }

        def coluna_sql(nome: str) -> str:
            return nome if nome in colunas_existentes else f"NULL AS {nome}"

        consulta = f"""
            SELECT
                id,
                {coluna_sql('equipamento_id')},
                {coluna_sql('usuario')},
                setor,
                maquina,
                codigo_barras,
                {coluna_sql('lote_identificado')},
                {coluna_sql('codigo_peca_identificado')},
                {coluna_sql('descricao_peca')},
                {coluna_sql('quantidade_programada')},
                fechada_em,
                {coluna_sql('quantidade_boa')},
                {coluna_sql('refugo')},
                {coluna_sql('retrabalho')},
                {coluna_sql('rota_codigo_equipamento')},
                {coluna_sql('rota_equipamento')},
                {coluna_sql('rota_operacao')},
                {coluna_sql('medida_ficha')},
                {coluna_sql('status')}
            FROM ordens_operacionais
            WHERE fechada_em IS NOT NULL
              AND TRIM(CAST(fechada_em AS TEXT)) <> ''
            ORDER BY id
        """
        dados = pd.read_sql_query(consulta, conexao)
    except (sqlite3.Error, pd.errors.DatabaseError, OSError) as exc:
        diagnostico["erro"] = f"Falha ao ler o banco operacional: {exc}"
        return vazio
    finally:
        if conexao is not None:
            conexao.close()

    diagnostico["total_fechados_banco"] = int(len(dados))
    if dados.empty:
        return vazio

    dados["Data Realizada DT"] = dados["fechada_em"].map(data_hora_operacional)
    diagnostico["datas_validas"] = int(dados["Data Realizada DT"].notna().sum())
    diagnostico["datas_invalidas"] = int(dados["Data Realizada DT"].isna().sum())

    dados = dados[dados["Data Realizada DT"].notna()].copy()
    if dados.empty:
        vazio.attrs["diagnostico_operacional"] = diagnostico
        return vazio

    data_movimento = dados["Data Realizada DT"].dt.date
    dados = dados[(data_movimento >= inicio) & (data_movimento <= fim)].copy()
    diagnostico["registros_periodo"] = int(len(dados))
    if dados.empty:
        vazio.attrs["diagnostico_operacional"] = diagnostico
        return vazio

    quantidade_boa_original = pd.to_numeric(
        dados["quantidade_boa"], errors="coerce"
    ).fillna(0).clip(lower=0)
    quantidade_programada_operacional = pd.to_numeric(
        dados["quantidade_programada"], errors="coerce"
    ).fillna(0).clip(lower=0)
    fallback_quantidade_operacional = (
        (quantidade_boa_original <= 0)
        & (quantidade_programada_operacional > 0)
        & dados["Data Realizada DT"].notna()
    )
    quantidade_boa = quantidade_boa_original.copy()
    quantidade_boa.loc[fallback_quantidade_operacional] = (
        quantidade_programada_operacional.loc[fallback_quantidade_operacional]
    )
    diagnostico["quantidades_positivas"] = int((quantidade_boa > 0).sum())
    diagnostico["quantidades_zero"] = int((quantidade_boa <= 0).sum())

    refugo = pd.to_numeric(dados["refugo"], errors="coerce").fillna(0).clip(lower=0)
    retrabalho = pd.to_numeric(dados["retrabalho"], errors="coerce").fillna(0).clip(lower=0)

    linhas = pd.DataFrame(index=dados.index)
    linhas["Linha Excel"] = dados["id"].map(lambda valor: f"SQL-{int(valor)}")
    linhas["ID Apontamento Operacional"] = dados["id"]
    linhas["ID Equipamento Operacional"] = dados["equipamento_id"]
    linhas["Usuário Operacional"] = dados["usuario"].fillna("").astype(str).map(texto_limpo)
    linhas["Data Programada"] = ""
    linhas["Data Programada Valor"] = pd.NaT
    linhas["Data Programada DT"] = pd.NaT
    linhas["Data Realizada DT"] = dados["Data Realizada DT"]
    linhas["Data Realizada Valor"] = dados["Data Realizada DT"]
    linhas["Data Realizada"] = dados["Data Realizada DT"].dt.strftime("%d/%m/%Y")
    linhas["Fechada Em Original"] = dados["fechada_em"].fillna("").astype(str)
    linhas["Setor"] = dados["setor"].fillna("").astype(str).map(texto_limpo).replace("", "Sem setor")

    maquina_operacional = dados["maquina"].fillna("").astype(str).map(texto_limpo)
    rota_equipamento = dados["rota_equipamento"].fillna("").astype(str).map(texto_limpo)
    codigo_equipamento = dados["rota_codigo_equipamento"].fillna("").astype(str).map(texto_limpo)
    maquina_final: list[str] = []
    for maquina, rota_nome, codigo in zip(maquina_operacional, rota_equipamento, codigo_equipamento):
        nome = maquina or rota_nome or "Sem máquina/posto"
        if codigo and not re.match(rf"^\s*0*{re.escape(codigo)}(?:\s|[-–—])", nome):
            nome = f"{codigo} - {nome}"
        maquina_final.append(nome)
    linhas["Máquina/Posto"] = maquina_final
    linhas["Código Equipamento Operacional"] = codigo_equipamento.values

    lote_identificado = dados["lote_identificado"].fillna("").astype(str).map(texto_limpo)
    codigo_barras = dados["codigo_barras"].fillna("").astype(str).map(texto_limpo)
    linhas["OP/Lote"] = lote_identificado.where(lote_identificado.ne(""), codigo_barras)
    linhas["Código de Barras Operacional"] = codigo_barras
    linhas["Produto/Equipamento"] = ""
    linhas["Código Peça"] = dados["codigo_peca_identificado"].fillna("").astype(str).map(texto_limpo)
    linhas["Descrição Peça"] = dados["descricao_peca"].fillna("").astype(str).map(texto_limpo)
    linhas["Operação"] = dados["rota_operacao"].fillna("").astype(str).map(texto_limpo)
    linhas["Qtde Programada"] = 0.0
    linhas["Qtde Programada Num"] = 0.0
    linhas["Qtde Realizada"] = quantidade_boa.values
    linhas["Qtde Realizada Num"] = quantidade_boa.values
    linhas["Qtde Realizada Considerada Num"] = quantidade_boa.values
    linhas["Qtde Apontada Num"] = quantidade_boa.values
    linhas["Qtde Realizada Vazia?"] = False
    linhas["Realizado por fallback?"] = fallback_quantidade_operacional.values
    linhas["Fallback Operacional por Quantidade Zero?"] = fallback_quantidade_operacional.values
    linhas["Qtde Boa Original Operacional"] = quantidade_boa_original.values
    linhas["Situação"] = "Realizado"
    linhas["Fonte Apontamento"] = "Operacional (SQLite)"
    linhas["Refugo"] = refugo.values
    linhas["Retrabalho"] = retrabalho.values
    linhas["Medida Catálogo"] = dados["medida_ficha"].fillna("").astype(str).map(texto_limpo)

    # A medida gravada na ficha é incorporada à descrição somente para o cálculo dimensional.
    descricao_calculo = linhas["Descrição Peça"].fillna("").astype(str)
    medida_ficha = dados["medida_ficha"].fillna("").astype(str).map(texto_limpo)
    linhas["Descrição Peça"] = [
        f"{descricao} {medida}".strip() if medida else descricao
        for descricao, medida in zip(descricao_calculo, medida_ficha)
    ]
    linhas = enriquecer_metricas_programacao(linhas.reset_index(drop=True))
    linhas.attrs["diagnostico_operacional"] = diagnostico
    return linhas

def combinar_apontamentos_planilha_operacional(
    base_planilha: pd.DataFrame,
    inicio: date,
    fim: date,
) -> tuple[pd.DataFrame, int, int, dict[str, Any]]:
    """Acrescenta o SQLite ao relatório, evita duplicatas e devolve diagnóstico."""
    planilha = base_planilha.copy()
    if "Fonte Apontamento" not in planilha.columns:
        planilha["Fonte Apontamento"] = "Planilha"
    if "ID Apontamento Operacional" not in planilha.columns:
        planilha["ID Apontamento Operacional"] = pd.NA
    if "Refugo" not in planilha.columns:
        planilha["Refugo"] = 0.0
    if "Retrabalho" not in planilha.columns:
        planilha["Retrabalho"] = 0.0

    operacional = carregar_apontamentos_operacionais_programado(inicio, fim)
    diagnostico = operacional.attrs.get(
        "diagnostico_operacional",
        _diagnostico_operacional_vazio(caminho_banco_operacional()),
    )
    if operacional.empty:
        return planilha, 0, 0, diagnostico

    apontado_planilha = planilha[
        planilha["Data Realizada DT"].notna()
        & (pd.to_numeric(planilha["Qtde Apontada Num"], errors="coerce").fillna(0) > 0)
    ].copy()
    assinaturas_planilha = {
        assinatura
        for _, linha in apontado_planilha.iterrows()
        if (assinatura := _assinatura_apontamento_relatorio(linha)) is not None
    }

    assinaturas_operacionais = operacional.apply(_assinatura_apontamento_relatorio, axis=1)
    duplicadas = assinaturas_operacionais.map(
        lambda assinatura: assinatura is not None and assinatura in assinaturas_planilha
    )
    quantidade_duplicadas = int(duplicadas.sum())
    operacional = operacional[~duplicadas].copy()
    quantidade_incluidas = len(operacional)

    combinado = pd.concat([planilha, operacional], ignore_index=True, sort=False)
    for coluna in ["Refugo", "Retrabalho"]:
        combinado[coluna] = pd.to_numeric(combinado[coluna], errors="coerce").fillna(0)
    combinado["Fonte Apontamento"] = combinado["Fonte Apontamento"].fillna("Planilha")
    return combinado, quantidade_incluidas, quantidade_duplicadas, diagnostico

def intervalo_programado_equipamento(base: pd.DataFrame) -> tuple[date, date] | None:
    """Obtém os limites da planilha e do SQLite usando a mesma conversão robusta."""
    datas = pd.concat(
        [base["Data Programada DT"], base["Data Realizada DT"]],
        ignore_index=True,
    ).dropna()
    limites: list[date] = []
    if not datas.empty:
        limites.extend([datas.min().date(), datas.max().date()])

    caminho = caminho_banco_operacional()
    if caminho.exists():
        conexao: sqlite3.Connection | None = None
        try:
            uri = caminho.resolve().as_uri() + "?mode=ro"
            conexao = sqlite3.connect(uri, uri=True, timeout=15)
            conexao.execute("PRAGMA busy_timeout = 15000")
            existe = conexao.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name='ordens_operacionais'"
            ).fetchone()
            if existe:
                valores = conexao.execute(
                    """
                    SELECT fechada_em
                    FROM ordens_operacionais
                    WHERE fechada_em IS NOT NULL
                      AND TRIM(CAST(fechada_em AS TEXT)) <> ''
                    """
                ).fetchall()
                datas_operacionais = pd.Series(
                    [data_hora_operacional(linha[0]) for linha in valores],
                    dtype="datetime64[ns]",
                ).dropna()
                if not datas_operacionais.empty:
                    limites.extend(
                        [datas_operacionais.min().date(), datas_operacionais.max().date()]
                    )
        except (sqlite3.Error, OSError):
            pass
        finally:
            if conexao is not None:
                conexao.close()

    if not limites:
        return None
    return min(limites), max(limites)

def escolher_periodo_programado_equipamento(
    base: pd.DataFrame,
) -> tuple[date | None, date | None, str]:
    intervalo = intervalo_programado_equipamento(base)
    if intervalo is None:
        st.warning("Não existem Datas Programadas ou Datas Realizadas válidas na base.")
        return None, None, ""

    data_min, data_max = intervalo
    dia_anterior = date.today() - timedelta(days=1)
    referencia_padrao = (
        dia_anterior
        if data_min <= dia_anterior <= data_max
        else data_max
    )

    modo = st.radio(
        "Período",
        ["Dia específico", "Período personalizado", "Período completo"],
        horizontal=True,
        key="prog_equip_modo_periodo",
    )

    if modo == "Semana de trabalho":
        referencia = st.date_input(
            "Escolha uma data da semana",
            value=referencia_padrao,
            min_value=data_min,
            max_value=data_max,
            format="DD/MM/YYYY",
            key="prog_equip_semana_referencia",
        )
        inicio = referencia - timedelta(days=referencia.weekday())
        fim = inicio + timedelta(days=4)
        st.caption(
            f"Semana considerada: segunda-feira {inicio:%d/%m/%Y} "
            f"a sexta-feira {fim:%d/%m/%Y}."
        )
    elif modo == "Dia específico":
        inicio = fim = st.date_input(
            "Data analisada",
            value=referencia_padrao,
            min_value=data_min,
            max_value=data_max,
            format="DD/MM/YYYY",
            key="prog_equip_dia",
        )
    elif modo == "Período personalizado":
        periodo = st.date_input(
            "Datas analisadas",
            value=(data_min, data_max),
            min_value=data_min,
            max_value=data_max,
            format="DD/MM/YYYY",
            key="prog_equip_periodo",
        )
        if isinstance(periodo, (tuple, list)) and len(periodo) == 2:
            inicio, fim = periodo
        elif isinstance(periodo, (tuple, list)) and len(periodo) == 1:
            inicio = fim = periodo[0]
        else:
            inicio = fim = periodo
    else:
        inicio, fim = data_min, data_max
        st.caption(f"Período completo da base: {inicio:%d/%m/%Y} a {fim:%d/%m/%Y}.")

    if inicio > fim:
        inicio, fim = fim, inicio
    return inicio, fim, modo


def _codigo_inicial_relatorio(valor: Any) -> str:
    texto = texto_limpo(valor)
    encontrado = re.match(r"^\s*0*(\d{1,4})(?:\s|[-–—]|$)", texto)
    return (encontrado.group(1).lstrip("0") or "0") if encontrado else ""


def _maquina_generica_do_setor(setor: Any, maquina: Any) -> bool:
    setor_norm = normalizar(setor)
    maquina_norm = normalizar(maquina)
    return (
        not maquina_norm
        or maquina_norm in {"sem maquina posto", "sem maquina", "sem posto"}
        or maquina_norm == setor_norm
    )


def _tokens_equipamento_relatorio(valor: Any) -> set[str]:
    """Retorna palavras úteis para distinguir máquinas com o mesmo código."""
    ignorar = {
        "maquina", "posto", "mod", "modelo", "serie", "equipamento", "de", "da",
        "do", "e", "n", "i", "auto", "automatica", "automatico",
    }
    return {
        token
        for token in normalizar(valor).split()
        if len(token) >= 3 and not token.isdigit() and token not in ignorar
    }


def _equipamento_cadastrado_relatorio(
    usuario: Any = "",
    setor: Any = "",
    maquina: Any = "",
    codigo: Any = "",
) -> dict[str, str] | None:
    """
    Localiza o cadastro oficial do equipamento.

    A tabela operacional pode guardar nomes antigos ou resumidos. O cadastro fixo dos
    logins é a fonte mais confiável para descobrir em qual setor o apontamento deve cair.
    """
    usuario_norm = normalizar(usuario)
    setor_norm = normalizar(setor)
    maquina_norm = normalizar(maquina)
    codigo_texto = texto_limpo(codigo) or _codigo_inicial_relatorio(maquina)
    codigo_texto = codigo_texto.lstrip("0") or ("0" if codigo_texto else "")

    registros = EQUIPAMENTOS_OPERADORES

    if usuario_norm:
        por_usuario = [r for r in registros if normalizar(r.get("usuario")) == usuario_norm]
        if len(por_usuario) == 1:
            return por_usuario[0]

    if maquina_norm:
        por_maquina_exata = [
            r for r in registros if normalizar(r.get("maquina")) == maquina_norm
        ]
        if len(por_maquina_exata) == 1:
            return por_maquina_exata[0]

        # Aceita versões resumidas, desde que a parte textual seja suficientemente específica.
        por_nome_contido = []
        for registro in registros:
            cadastrado_norm = normalizar(registro.get("maquina"))
            if not cadastrado_norm:
                continue
            if maquina_norm in cadastrado_norm or cadastrado_norm in maquina_norm:
                por_nome_contido.append(registro)
        if len(por_nome_contido) == 1:
            return por_nome_contido[0]

    candidatos_codigo = []
    if codigo_texto:
        candidatos_codigo = [
            r
            for r in registros
            if (texto_limpo(r.get("codigo")).lstrip("0") or "0") == codigo_texto
        ]
        if len(candidatos_codigo) == 1:
            return candidatos_codigo[0]

    # Há códigos repetidos no cadastro, como 124. Nesse caso, usa palavras como CNC/PADE.
    if candidatos_codigo and maquina_norm:
        tokens_origem = _tokens_equipamento_relatorio(maquina)
        pontuados: list[tuple[int, dict[str, str]]] = []
        for registro in candidatos_codigo:
            pontos = len(tokens_origem & _tokens_equipamento_relatorio(registro.get("maquina")))
            if setor_norm and normalizar(registro.get("setor")) == setor_norm:
                pontos += 3
            pontuados.append((pontos, registro))
        pontuados.sort(key=lambda item: item[0], reverse=True)
        if pontuados and pontuados[0][0] > 0:
            if len(pontuados) == 1 or pontuados[0][0] > pontuados[1][0]:
                return pontuados[0][1]

    return None


def direcionar_apontamentos_para_programacao_do_dia(
    apontado_base: pd.DataFrame,
    programado_base: pd.DataFrame,
) -> pd.DataFrame:
    """
    Consolida o apontado na linha programada do mesmo dia.

    O erro anterior acontecia porque o SQLite registra a máquina física, por exemplo
    "20 - FURADEIRA 02", enquanto a programação pode possuir apenas a linha genérica
    "10 FURADEIRAS". Agora o cadastro do login/equipamento é usado para descobrir o
    setor oficial antes da consolidação.
    """
    if apontado_base.empty:
        return apontado_base.copy()

    resultado = apontado_base.copy()
    resultado["Data"] = resultado["Data Realizada DT"].dt.date
    if programado_base.empty:
        resultado["Regra Consolidação Apontado"] = "Sem programação no período"
        return resultado

    programacao = programado_base.copy()
    programacao["Data"] = programacao["Data Programada DT"].dt.date
    programacao = programacao[
        ["Data", "Setor", "Máquina/Posto", "Unidade Métrica"]
    ].drop_duplicates()

    por_dia_setor: dict[tuple[date, str], list[dict[str, str]]] = {}
    por_dia_codigo_setor: dict[tuple[date, str], list[dict[str, str]]] = {}
    por_dia_maquina: dict[tuple[date, str], list[dict[str, str]]] = {}
    por_dia_codigo_maquina: dict[tuple[date, str], list[dict[str, str]]] = {}
    por_dia: dict[date, list[dict[str, str]]] = {}

    for _, linha in programacao.iterrows():
        item = {
            "setor": texto_limpo(linha["Setor"]) or "Sem setor",
            "maquina": texto_limpo(linha["Máquina/Posto"]) or "Sem máquina/posto",
            "unidade": texto_limpo(linha["Unidade Métrica"]) or "uni",
        }
        data_item = linha["Data"]
        setor_norm = normalizar(item["setor"])
        maquina_norm = normalizar(item["maquina"])
        codigo_setor = _codigo_inicial_relatorio(item["setor"])
        codigo_maquina = _codigo_inicial_relatorio(item["maquina"])

        por_dia.setdefault(data_item, []).append(item)
        por_dia_setor.setdefault((data_item, setor_norm), []).append(item)
        por_dia_maquina.setdefault((data_item, maquina_norm), []).append(item)
        if codigo_setor:
            por_dia_codigo_setor.setdefault((data_item, codigo_setor), []).append(item)
        if codigo_maquina:
            por_dia_codigo_maquina.setdefault((data_item, codigo_maquina), []).append(item)

    setores_destino: list[str] = []
    maquinas_destino: list[str] = []
    unidades_destino: list[str] = []
    regras_destino: list[str] = []
    setores_cadastro: list[str] = []
    maquinas_cadastro: list[str] = []

    for _, linha in resultado.iterrows():
        data_item = linha["Data"]
        setor_origem = texto_limpo(linha.get("Setor")) or "Sem setor"
        maquina_origem = texto_limpo(linha.get("Máquina/Posto")) or "Sem máquina/posto"
        unidade_origem = texto_limpo(linha.get("Unidade Métrica")) or "uni"
        usuario_origem = texto_limpo(linha.get("Usuário Operacional", linha.get("usuario", "")))
        codigo_origem = texto_limpo(
            linha.get("Código Equipamento Operacional", linha.get("rota_codigo_equipamento", ""))
        )

        cadastro = _equipamento_cadastrado_relatorio(
            usuario=usuario_origem,
            setor=setor_origem,
            maquina=maquina_origem,
            codigo=codigo_origem,
        )
        setor_cadastro = texto_limpo(cadastro.get("setor")) if cadastro else ""
        maquina_cadastro = texto_limpo(cadastro.get("maquina")) if cadastro else ""
        setores_cadastro.append(setor_cadastro)
        maquinas_cadastro.append(maquina_cadastro)

        # Tenta primeiro o setor oficial do cadastro; depois o setor gravado na ordem.
        chaves_setor: list[str] = []
        for valor_setor in [setor_cadastro, setor_origem]:
            chave = normalizar(valor_setor)
            if chave and chave not in chaves_setor:
                chaves_setor.append(chave)

        candidatos: list[dict[str, str]] = []
        origem_candidatos = ""
        for chave_setor in chaves_setor:
            encontrados = por_dia_setor.get((data_item, chave_setor), [])
            if encontrados:
                candidatos = encontrados
                origem_candidatos = "Setor cadastrado" if setor_cadastro and chave_setor == normalizar(setor_cadastro) else "Mesmo setor"
                break

        if not candidatos:
            for valor_setor in [setor_cadastro, setor_origem]:
                codigo_setor = _codigo_inicial_relatorio(valor_setor)
                encontrados = (
                    por_dia_codigo_setor.get((data_item, codigo_setor), [])
                    if codigo_setor else []
                )
                if encontrados:
                    candidatos = encontrados
                    origem_candidatos = "Mesmo código de setor"
                    break

        escolhido: dict[str, str] | None = None
        regra = "Sem programação correspondente"

        # Dentro do setor correto, procura a máquina exata ou seu código.
        nomes_maquina = [maquina_origem, maquina_cadastro]
        for nome_maquina in nomes_maquina:
            maquina_norm = normalizar(nome_maquina)
            if not maquina_norm:
                continue
            exatos = [c for c in candidatos if normalizar(c["maquina"]) == maquina_norm]
            if len(exatos) == 1:
                escolhido = exatos[0]
                regra = f"{origem_candidatos} + mesma máquina".strip(" +")
                break

        if escolhido is None and candidatos:
            codigos_maquina = []
            for nome_maquina in nomes_maquina:
                codigo_maquina = _codigo_inicial_relatorio(nome_maquina)
                if codigo_maquina and codigo_maquina not in codigos_maquina:
                    codigos_maquina.append(codigo_maquina)
            for codigo_maquina in codigos_maquina:
                por_codigo = [
                    c for c in candidatos
                    if not _maquina_generica_do_setor(c["setor"], c["maquina"])
                    and _codigo_inicial_relatorio(c["maquina"]) == codigo_maquina
                ]
                if len(por_codigo) == 1:
                    escolhido = por_codigo[0]
                    regra = f"{origem_candidatos} + mesmo código de máquina".strip(" +")
                    break

        # A programação consolidada normalmente usa a linha genérica do setor.
        if escolhido is None and candidatos:
            genericos = [
                c for c in candidatos
                if _maquina_generica_do_setor(c["setor"], c["maquina"])
            ]
            if len(genericos) == 1:
                escolhido = genericos[0]
                regra = (
                    "Cadastro do equipamento → linha genérica do setor"
                    if setor_cadastro else "Máquina genérica do setor"
                )

        if escolhido is None and len(candidatos) == 1:
            escolhido = candidatos[0]
            regra = (
                "Cadastro do equipamento → única linha programada no setor"
                if setor_cadastro else "Única máquina programada no setor"
            )

        # Último recurso: procura a máquina no dia inteiro, mesmo se o setor foi gravado errado.
        if escolhido is None:
            for nome_maquina in nomes_maquina:
                maquina_norm = normalizar(nome_maquina)
                exatos_globais = (
                    por_dia_maquina.get((data_item, maquina_norm), []) if maquina_norm else []
                )
                if len(exatos_globais) == 1:
                    escolhido = exatos_globais[0]
                    regra = "Máquina localizada no dia, independente do setor gravado"
                    break

        if escolhido is None:
            codigos_globais = []
            for nome_maquina in nomes_maquina:
                codigo_maquina = _codigo_inicial_relatorio(nome_maquina)
                if codigo_maquina and codigo_maquina not in codigos_globais:
                    codigos_globais.append(codigo_maquina)
            for codigo_maquina in codigos_globais:
                por_codigo_global = por_dia_codigo_maquina.get((data_item, codigo_maquina), [])
                if len(por_codigo_global) == 1:
                    escolhido = por_codigo_global[0]
                    regra = "Código da máquina localizado no dia"
                    break

        if escolhido:
            setores_destino.append(escolhido["setor"])
            maquinas_destino.append(escolhido["maquina"])
            unidades_destino.append(escolhido["unidade"])
        else:
            setores_destino.append(setor_cadastro or setor_origem)
            maquinas_destino.append(maquina_origem)
            unidades_destino.append(unidade_origem)
        regras_destino.append(regra)

    resultado["Setor Origem Apontamento"] = resultado["Setor"]
    resultado["Máquina Origem Apontamento"] = resultado["Máquina/Posto"]
    resultado["Setor Cadastro Equipamento"] = setores_cadastro
    resultado["Máquina Cadastro Equipamento"] = maquinas_cadastro
    resultado["Setor"] = setores_destino
    resultado["Máquina/Posto"] = maquinas_destino
    resultado["Unidade Métrica"] = unidades_destino
    resultado["Regra Consolidação Apontado"] = regras_destino
    return resultado

def montar_programado_dia_equipamento(
    base: pd.DataFrame,
    inicio: date,
    fim: date,
    setores_sel: list[str],
    maquinas_sel: list[str],
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Compara, em cada dia, o programado com o apontamento manual efetivo.

    Regras fixas:
    - Programado: Data Programada dentro do período.
    - Apontado: Data Realizada dentro do período e Qtde Realizada maior que zero.
    - Fonte: somente a planilha de acompanhamento; o SQLite operacional não participa.
    """
    base_filtros = base.copy()

    mascara_programado = (
        base_filtros["Data Programada DT"].notna()
        & (base_filtros["Data Programada DT"].dt.date >= inicio)
        & (base_filtros["Data Programada DT"].dt.date <= fim)
    )
    programado_base = base_filtros[mascara_programado].copy()

    mascara_apontado = (
        base_filtros["Data Realizada DT"].notna()
        & (base_filtros["Data Realizada DT"].dt.date >= inicio)
        & (base_filtros["Data Realizada DT"].dt.date <= fim)
        & (pd.to_numeric(base_filtros["Qtde Apontada Num"], errors="coerce").fillna(0) > 0)
    )
    apontado_base = base_filtros[mascara_apontado].copy()

    if setores_sel:
        programado_base = programado_base[programado_base["Setor"].isin(setores_sel)]
        apontado_base = apontado_base[apontado_base["Setor"].isin(setores_sel)]
    if maquinas_sel:
        programado_base = programado_base[
            programado_base["Máquina/Posto"].isin(maquinas_sel)
        ]
        apontado_base = apontado_base[
            apontado_base["Máquina/Posto"].isin(maquinas_sel)
        ]

    programado_base["Data"] = programado_base["Data Programada DT"].dt.date
    programado_base["Data Referência"] = "Data Programada"
    apontado_base["Data"] = apontado_base["Data Realizada DT"].dt.date
    apontado_base["Data Referência"] = "Data Realizada"

    linhas_periodo = pd.concat(
        [
            programado_base.assign(Tipo_Movimento="Programado"),
            apontado_base.assign(Tipo_Movimento="Apontado"),
        ],
        ignore_index=True,
        sort=False,
    )

    chaves = ["Data", "Setor", "Máquina/Posto", "Unidade Métrica"]

    if programado_base.empty:
        programado = pd.DataFrame(
            columns=chaves
            + [
                "Programado",
                "Métrica Programada",
                "Linhas Programadas",
                "Lotes Programados",
                "Peças Programadas Distintas",
                "Linhas sem Medida Programadas",
            ]
        )
    else:
        programado = (
            programado_base.groupby(chaves, as_index=False, dropna=False)
            .agg(
                Programado=("Qtde Programada Num", "sum"),
                **{
                    "Métrica Programada": ("Métrica Programada Num", "sum"),
                    "Linhas Programadas": ("Linha Excel", "count"),
                    "Lotes Programados": ("OP/Lote", juntar_lotes),
                    "Peças Programadas Distintas": ("Código Peça", "nunique"),
                    "Linhas sem Medida Programadas": (
                        "Medida Válida?",
                        lambda s: int((~s).sum()),
                    ),
                },
            )
        )

    if apontado_base.empty:
        apontado = pd.DataFrame(
            columns=chaves
            + [
                "Apontado",
                "Métrica Apontada",
                "Linhas Apontadas",
                "Lotes Apontados",
                "Peças Apontadas Distintas",
                "Linhas sem Medida Apontadas",
            ]
        )
    else:
        apontado = (
            apontado_base.groupby(chaves, as_index=False, dropna=False)
            .agg(
                Apontado=("Qtde Apontada Num", "sum"),
                **{
                    "Métrica Apontada": ("Métrica Apontada Num", "sum"),
                    "Linhas Apontadas": ("Linha Excel", "count"),
                    "Lotes Apontados": ("OP/Lote", juntar_lotes),
                    "Peças Apontadas Distintas": ("Código Peça", "nunique"),
                    "Linhas sem Medida Apontadas": (
                        "Medida Válida?",
                        lambda s: int((~s).sum()),
                    ),
                },
            )
        )

    diario = programado.merge(apontado, on=chaves, how="outer")
    if diario.empty:
        return linhas_periodo, diario, pd.DataFrame()

    colunas_numericas = [
        "Programado",
        "Apontado",
        "Métrica Programada",
        "Métrica Apontada",
        "Linhas Programadas",
        "Linhas Apontadas",
        "Peças Programadas Distintas",
        "Peças Apontadas Distintas",
        "Linhas sem Medida Programadas",
        "Linhas sem Medida Apontadas",
    ]
    for coluna in colunas_numericas:
        if coluna not in diario.columns:
            diario[coluna] = 0
        diario[coluna] = pd.to_numeric(diario[coluna], errors="coerce").fillna(0)

    for coluna in ["Lotes Programados", "Lotes Apontados"]:
        if coluna not in diario.columns:
            diario[coluna] = ""
        diario[coluna] = diario[coluna].fillna("")

    diario["Saldo Peças"] = diario["Programado"] - diario["Apontado"]
    diario["Saldo Métrica"] = diario["Métrica Programada"] - diario["Métrica Apontada"]
    diario["% Atendimento Peças"] = 0.0
    diario["% Atendimento Métrica"] = 0.0
    tem_programado = diario["Programado"] > 0
    tem_metrica = diario["Métrica Programada"] > 0
    diario.loc[tem_programado, "% Atendimento Peças"] = (
        diario.loc[tem_programado, "Apontado"]
        / diario.loc[tem_programado, "Programado"]
        * 100
    )
    diario.loc[tem_metrica, "% Atendimento Métrica"] = (
        diario.loc[tem_metrica, "Métrica Apontada"]
        / diario.loc[tem_metrica, "Métrica Programada"]
        * 100
    )
    diario = diario.sort_values(["Data", "Setor", "Máquina/Posto"])

    equipamentos = (
        diario[["Setor", "Máquina/Posto", "Unidade Métrica"]]
        .drop_duplicates()
        .sort_values(["Setor", "Máquina/Posto"])
    )
    indices = ["Setor", "Máquina/Posto", "Unidade Métrica"]
    matriz = equipamentos.set_index(indices)

    quantidade_dias = (fim - inicio).days + 1
    if quantidade_dias <= 31:
        datas_exibicao = [d.date() for d in pd.date_range(inicio, fim, freq="D")]
    else:
        datas_exibicao = sorted(diario["Data"].dropna().unique().tolist())

    colunas_dias: dict[str, pd.Series] = {}
    for data_movimento in datas_exibicao:
        movimento_data = diario[diario["Data"] == data_movimento].set_index(indices)
        data_texto = data_movimento.strftime("%d/%m/%Y")
        colunas_dias[f"{data_texto} — Prog. peças"] = movimento_data["Programado"]
        colunas_dias[f"{data_texto} — Apont. peças"] = movimento_data["Apontado"]
        colunas_dias[f"{data_texto} — Lotes programados"] = movimento_data[
            "Lotes Programados"
        ]
        colunas_dias[f"{data_texto} — Lotes apontados"] = movimento_data[
            "Lotes Apontados"
        ]
        colunas_dias[f"{data_texto} — Prog. métrica"] = movimento_data[
            "Métrica Programada"
        ]
        colunas_dias[f"{data_texto} — Apont. métrica"] = movimento_data[
            "Métrica Apontada"
        ]
    if colunas_dias:
        matriz = matriz.join(pd.DataFrame(colunas_dias), how="left")

    col_lotes = [c for c in matriz.columns if "— Lotes " in c]
    col_numericas_matriz = [c for c in matriz.columns if c not in col_lotes]
    if col_numericas_matriz:
        matriz[col_numericas_matriz] = matriz[col_numericas_matriz].fillna(0)
    if col_lotes:
        matriz[col_lotes] = matriz[col_lotes].fillna("")

    # Consolida os blocos adicionados antes de calcular os totais. Isso evita
    # fragmentação interna do DataFrame em períodos com muitas datas/colunas.
    matriz = matriz.copy()

    col_prog_pecas = [c for c in matriz.columns if c.endswith("— Prog. peças")]
    col_apont_pecas = [c for c in matriz.columns if c.endswith("— Apont. peças")]
    col_prog_metrica = [c for c in matriz.columns if c.endswith("— Prog. métrica")]
    col_apont_metrica = [c for c in matriz.columns if c.endswith("— Apont. métrica")]

    matriz["Total Programado (peças)"] = matriz[col_prog_pecas].sum(axis=1)
    matriz["Total Apontado (peças)"] = matriz[col_apont_pecas].sum(axis=1)
    matriz["Saldo (peças)"] = (
        matriz["Total Programado (peças)"] - matriz["Total Apontado (peças)"]
    )
    matriz["% Atendimento Peças"] = 0.0
    tem_programado_matriz = matriz["Total Programado (peças)"] > 0
    matriz.loc[tem_programado_matriz, "% Atendimento Peças"] = (
        matriz.loc[tem_programado_matriz, "Total Apontado (peças)"]
        / matriz.loc[tem_programado_matriz, "Total Programado (peças)"]
        * 100
    )
    matriz["Total Programado (métrica)"] = matriz[col_prog_metrica].sum(axis=1)
    matriz["Total Apontado (métrica)"] = matriz[col_apont_metrica].sum(axis=1)
    matriz["Saldo (métrica)"] = (
        matriz["Total Programado (métrica)"] - matriz["Total Apontado (métrica)"]
    )
    matriz["% Atendimento Métrica"] = 0.0
    tem_metrica_matriz = matriz["Total Programado (métrica)"] > 0
    matriz.loc[tem_metrica_matriz, "% Atendimento Métrica"] = (
        matriz.loc[tem_metrica_matriz, "Total Apontado (métrica)"]
        / matriz.loc[tem_metrica_matriz, "Total Programado (métrica)"]
        * 100
    )
    matriz = matriz.reset_index().sort_values(
        ["Total Programado (peças)", "Total Apontado (peças)", "Setor", "Máquina/Posto"],
        ascending=[False, False, True, True],
    )
    return linhas_periodo, diario, matriz


def renderizar_programado_dia_equipamento(df: pd.DataFrame, caminho: Path) -> None:
    st.title("PROGRAMADO X REALIZADO")
    st.caption(
        ""
        ""
    )

    base = preparar_base_programado_equipamento(df)
    inicio, fim, modo = escolher_periodo_programado_equipamento(base)
    if inicio is None or fim is None:
        return

    st.info(
        ""
        ""
        ""
    )

    st.subheader("Filtros")
    col_setor, col_maquina = st.columns(2)
    setores = opcoes_unicas(base, "Setor")
    setores_sel = col_setor.multiselect(
        "Setor", setores, placeholder="Vazio = todos os setores", key="prog_equip_setores"
    )

    base_maquinas = base.copy()
    if setores_sel:
        base_maquinas = base_maquinas[base_maquinas["Setor"].isin(setores_sel)]
    maquinas = opcoes_unicas(base_maquinas, "Máquina/Posto")
    maquinas_sel = col_maquina.multiselect(
        "Equipamento",
        maquinas,
        placeholder="Vazio = todos os equipamentos",
        key="prog_equip_maquinas",
    )

    linhas_periodo, diario, matriz = montar_programado_dia_equipamento(
        base, inicio, fim, setores_sel, maquinas_sel
    )

    if diario.empty:
        st.warning("Não existe programação nem apontamento manual para os filtros selecionados.")
        return

    total_programado = float(diario["Programado"].sum())
    total_apontado = float(diario["Apontado"].sum())
    saldo = total_programado - total_apontado
    percentual = (total_apontado / total_programado * 100) if total_programado > 0 else 0.0
    equipamentos = int(diario[["Setor", "Máquina/Posto"]].drop_duplicates().shape[0])

    st.divider()
    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("Programado (peças)", fmt_num(total_programado, 2))
    k2.metric("Apontado na data", fmt_num(total_apontado, 2))
    k3.metric("Saldo (peças)", fmt_num(saldo, 2))
    k4.metric("% apontado", f"{percentual:.1f}%".replace(".", ","))
    k5.metric("Equipamentos", fmt_num(equipamentos))

    aba_matriz, aba_diario, aba_equipamento, aba_detalhe = st.tabs(
        ["Matriz diária", "Consolidado diário", "Total por equipamento", "Linhas consideradas"]
    )

    with aba_matriz:
        st.subheader("Programado, apontado e lotes por equipamento em cada dia")
        st.caption(
            "Cada data possui as quantidades programada e apontada, os lotes programados "
            "e apontados, além das métricas produtivas."
        )
        configuracao_colunas: dict[str, Any] = {
            "Setor": st.column_config.TextColumn("Setor", width="medium"),
            "Máquina/Posto": st.column_config.TextColumn("Equipamento/Máquina", width="large"),
            "Unidade Métrica": st.column_config.TextColumn("Unidade", width="small"),
            "Total Programado (peças)": st.column_config.NumberColumn("Prog. peças", format="%.2f"),
            "Total Apontado (peças)": st.column_config.NumberColumn("Apont. peças", format="%.2f"),
            "Saldo (peças)": st.column_config.NumberColumn("Saldo peças", format="%.2f"),
            "% Atendimento Peças": st.column_config.NumberColumn("% peças", format="%.1f%%"),
            "Total Programado (métrica)": st.column_config.NumberColumn("Prog. métrica", format="%.3f"),
            "Total Apontado (métrica)": st.column_config.NumberColumn("Apont. métrica", format="%.3f"),
            "Saldo (métrica)": st.column_config.NumberColumn("Saldo métrica", format="%.3f"),
            "% Atendimento Métrica": st.column_config.NumberColumn("% métrica", format="%.1f%%"),
        }
        for coluna in matriz.columns:
            if coluna.endswith("— Prog. peças") or coluna.endswith("— Apont. peças"):
                configuracao_colunas[coluna] = st.column_config.NumberColumn(coluna, format="%.2f")
            elif "— Lotes " in coluna:
                configuracao_colunas[coluna] = st.column_config.TextColumn(coluna, width="large")
            elif coluna.endswith("— Prog. métrica") or coluna.endswith("— Apont. métrica"):
                configuracao_colunas[coluna] = st.column_config.NumberColumn(coluna, format="%.3f")

        linhas_abaixo_programado = mascara_programado_nao_atendido(matriz)
        quantidade_abaixo = int(linhas_abaixo_programado.sum())
        if quantidade_abaixo > 0:
            st.error(
                f"{quantidade_abaixo} equipamento(s) ficaram abaixo da quantidade programada "
                "no período selecionado e estão destacados em vermelho."
            )
        else:
            st.success("Todos os equipamentos com programação atingiram a quantidade programada.")

        st.dataframe(
            estilizar_programado_nao_atendido(matriz),
            use_container_width=True,
            hide_index=True,
            height=600,
            column_config=configuracao_colunas,
        )
        excel_matriz = dataframe_para_excel(
            matriz, "Matriz diária", linhas_alerta=linhas_abaixo_programado
        )
        csv_matriz = dataframe_para_csv_excel(matriz)
        coluna_excel, coluna_csv = st.columns(2)
        coluna_excel.download_button(
            "Baixar matriz em Excel",
            data=excel_matriz,
            file_name=f"programado_apontado_lotes_{inicio:%Y%m%d}_{fim:%Y%m%d}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_prog_equip_matriz_xlsx",
            use_container_width=True,
        )
        coluna_csv.download_button(
            "Baixar matriz em CSV",
            data=csv_matriz,
            file_name=f"programado_apontado_lotes_{inicio:%Y%m%d}_{fim:%Y%m%d}.csv",
            mime="text/csv; charset=utf-8",
            key="download_prog_equip_matriz_csv",
            use_container_width=True,
        )

    with aba_diario:
        exibicao_diaria = diario.copy()
        exibicao_diaria["Data"] = pd.to_datetime(exibicao_diaria["Data"]).dt.strftime("%d/%m/%Y")
        st.dataframe(
            estilizar_programado_nao_atendido(exibicao_diaria),
            use_container_width=True,
            hide_index=True,
            height=580,
            column_config={
                "Máquina/Posto": st.column_config.TextColumn("Equipamento/Máquina", width="large"),
                "Unidade Métrica": st.column_config.TextColumn("Unidade", width="small"),
                "Programado": st.column_config.NumberColumn("Prog. peças", format="%.2f"),
                "Apontado": st.column_config.NumberColumn("Apont. peças", format="%.2f"),
                "Lotes Programados": st.column_config.TextColumn("Lotes programados", width="large"),
                "Lotes Apontados": st.column_config.TextColumn("Lotes apontados", width="large"),
                "Saldo Peças": st.column_config.NumberColumn("Saldo peças", format="%.2f"),
                "% Atendimento Peças": st.column_config.NumberColumn("% peças", format="%.1f%%"),
                "Métrica Programada": st.column_config.NumberColumn("Prog. métrica", format="%.3f"),
                "Métrica Apontada": st.column_config.NumberColumn("Apont. métrica", format="%.3f"),
                "Saldo Métrica": st.column_config.NumberColumn("Saldo métrica", format="%.3f"),
                "% Atendimento Métrica": st.column_config.NumberColumn("% métrica", format="%.1f%%"),
            },
        )

    with aba_equipamento:
        resumo = (
            diario.groupby(["Setor", "Máquina/Posto", "Unidade Métrica"], as_index=False)
            .agg(
                **{
                    "Programado Peças": ("Programado", "sum"),
                    "Apontado Peças": ("Apontado", "sum"),
                    "Programado Métrica": ("Métrica Programada", "sum"),
                    "Apontado Métrica": ("Métrica Apontada", "sum"),
                    "Dias Programados": ("Programado", lambda s: int((s > 0).sum())),
                    "Dias Apontados": ("Apontado", lambda s: int((s > 0).sum())),
                }
            )
        )
        resumo["Saldo Peças"] = resumo["Programado Peças"] - resumo["Apontado Peças"]
        resumo["% Atendimento Peças"] = 0.0
        tem_prog = resumo["Programado Peças"] > 0
        resumo.loc[tem_prog, "% Atendimento Peças"] = (
            resumo.loc[tem_prog, "Apontado Peças"]
            / resumo.loc[tem_prog, "Programado Peças"]
            * 100
        )
        resumo["Saldo Métrica"] = resumo["Programado Métrica"] - resumo["Apontado Métrica"]

        programados_lotes = linhas_periodo[linhas_periodo["Tipo_Movimento"].eq("Programado")]
        apontados_lotes = linhas_periodo[linhas_periodo["Tipo_Movimento"].eq("Apontado")]
        lotes_prog = (
            programados_lotes.groupby(["Setor", "Máquina/Posto"])["OP/Lote"]
            .agg(juntar_lotes)
            .rename("Lotes Programados")
            .reset_index()
            if not programados_lotes.empty
            else pd.DataFrame(columns=["Setor", "Máquina/Posto", "Lotes Programados"])
        )
        lotes_apont = (
            apontados_lotes.groupby(["Setor", "Máquina/Posto"])["OP/Lote"]
            .agg(juntar_lotes)
            .rename("Lotes Apontados")
            .reset_index()
            if not apontados_lotes.empty
            else pd.DataFrame(columns=["Setor", "Máquina/Posto", "Lotes Apontados"])
        )
        resumo = resumo.merge(lotes_prog, on=["Setor", "Máquina/Posto"], how="left")
        resumo = resumo.merge(lotes_apont, on=["Setor", "Máquina/Posto"], how="left")
        resumo[["Lotes Programados", "Lotes Apontados"]] = resumo[
            ["Lotes Programados", "Lotes Apontados"]
        ].fillna("")
        resumo = resumo.sort_values(
            ["Programado Peças", "Apontado Peças"], ascending=[False, False]
        )

        st.dataframe(
            estilizar_programado_nao_atendido(resumo),
            use_container_width=True,
            hide_index=True,
            height=580,
            column_config={
                "Máquina/Posto": st.column_config.TextColumn("Equipamento/Máquina", width="large"),
                "Unidade Métrica": st.column_config.TextColumn("Unidade", width="small"),
                "Programado Peças": st.column_config.NumberColumn("Prog. peças", format="%.2f"),
                "Apontado Peças": st.column_config.NumberColumn("Apont. peças", format="%.2f"),
                "Saldo Peças": st.column_config.NumberColumn("Saldo peças", format="%.2f"),
                "% Atendimento Peças": st.column_config.NumberColumn("% peças", format="%.1f%%"),
                "Lotes Programados": st.column_config.TextColumn("Lotes programados", width="large"),
                "Lotes Apontados": st.column_config.TextColumn("Lotes apontados", width="large"),
                "Dias Programados": st.column_config.NumberColumn("Dias programados", format="%d"),
                "Dias Apontados": st.column_config.NumberColumn("Dias apontados", format="%d"),
            },
        )

    with aba_detalhe:
        detalhe_base = linhas_periodo.copy()
        detalhe_base["Qtde Apontada Considerada"] = detalhe_base["Qtde Apontada Num"]
        detalhe_base["Métrica Programada"] = detalhe_base["Métrica Programada Num"]
        detalhe_base["Métrica Apontada"] = detalhe_base["Métrica Apontada Num"]
        detalhe_base = detalhe_base.sort_values(
            ["Data", "Setor", "Máquina/Posto", "OP/Lote", "Tipo_Movimento"],
            na_position="last",
        )
        colunas_detalhe = [
            "Tipo_Movimento", "Data", "Data Referência", "Data Programada", "Data Realizada",
            "Setor", "Máquina/Posto", "Unidade Métrica", "OP/Lote",
            "Produto/Equipamento", "Código Peça", "Descrição Peça", "Operação",
            "Qtde Programada", "Qtde Realizada", "Qtde Apontada Considerada",
            "Comprimento (m)", "Largura (m)", "Métrica Programada", "Métrica Apontada",
        ]
        detalhe = detalhe_base[[c for c in colunas_detalhe if c in detalhe_base.columns]].copy()
        if "Data" in detalhe.columns:
            detalhe["Data"] = pd.to_datetime(detalhe["Data"]).dt.strftime("%d/%m/%Y")
        st.caption(f"{len(detalhe)} linha(s) de programação ou apontamento manual no período.")
        st.dataframe(
            detalhe,
            use_container_width=True,
            hide_index=True,
            height=600,
            column_config={
                "Máquina/Posto": st.column_config.TextColumn("Equipamento/Máquina", width="large"),
                "Descrição Peça": st.column_config.TextColumn("Descrição Peça", width="large"),
                "Métrica Programada": st.column_config.NumberColumn("Prog. métrica", format="%.3f"),
                "Métrica Apontada": st.column_config.NumberColumn("Apont. métrica", format="%.3f"),
            },
        )

    st.divider()
    st.caption(f"Fonte dos dados: {caminho.name} — apontamentos manuais da planilha.")


def fmt_qtd(valor: Any) -> str:
    """Formata quantidade sem casas desnecessárias e com padrão brasileiro."""
    try:
        numero = float(valor)
    except Exception:
        numero = 0.0
    casas = 0 if numero.is_integer() else 2
    return fmt_num(numero, casas)


def juntar_lotes(valores: Any) -> str:
    """Junta os lotes únicos, limpos e ordenados naturalmente."""
    lotes: list[str] = []
    vistos: set[str] = set()
    for valor in valores:
        lote = texto_limpo(valor)
        if lote and lote not in vistos:
            vistos.add(lote)
            lotes.append(lote)
    return ", ".join(sorted(lotes, key=chave_ordenacao_texto))



# -----------------------------------------------------------------------------
# Tela pública: acumulado diário por setor e comparação com metas
# -----------------------------------------------------------------------------
# Metas diárias fornecidas pelo PPCP. A ordem abaixo também define a ordem das
# colunas no painel, reproduzindo o modelo visual enviado.
METAS_DIARIAS_SETOR = [
    {
        "chave": "seccionadora 2",
        "nome": "SECCIONADORA 2",
        "meta": 750.0,
        "unidade": "M²",
        "aliases": ["seccionadora 2"],
    },
    {
        "chave": "seccionadora 1",
        "nome": "SECCIONADORA 1",
        "meta": 750.0,
        "unidade": "M²",
        "aliases": ["seccionadora 1"],
    },
    {
        "chave": "usinagem automatizada",
        "nome": "USINAGEM AUTOMATIZADA",
        "meta": 4150.0,
        "unidade": "uni",
        "aliases": ["usinagem automatizada"],
    },
    {
        "chave": "usinagem",
        "nome": "USINAGEM",
        "meta": 3900.0,
        "unidade": "uni",
        "aliases": ["usinagem"],
    },
    {
        "chave": "cnc",
        "nome": "CNC",
        "meta": 740.0,
        "unidade": "uni",
        "aliases": ["cnc"],
    },
    {
        "chave": "pintura p u",
        "nome": "PINTURA P.U.",
        "meta": 2060.0,
        "unidade": "uni",
        "aliases": ["pintura p u", "pintura pu"],
    },
    {
        "chave": "furadeiras",
        "nome": "FURADEIRAS",
        "meta": 10590.0,
        "unidade": "uni",
        "aliases": ["furadeiras"],
    },
    {
        "chave": "coladeiras",
        "nome": "COLADEIRAS",
        "meta": 3350.0,
        "unidade": "m linear",
        "aliases": ["coladeiras"],
    },
    {
        "chave": "embalagem",
        "nome": "EMBALAGEM",
        "meta": 1150.0,
        "unidade": "uni",
        "aliases": ["embalagem"],
    },
    {
        "chave": "fabricacao embalagem",
        "nome": "FABRICAÇÃO DE EMBALAGEM",
        "meta": 2350.0,
        "unidade": "uni",
        "aliases": ["fabricacao embalagem", "fabricacao de embalagem"],
    },
    {
        "chave": "tapecaria",
        "nome": "TAPEÇARIA",
        "meta": 1920.0,
        "unidade": "uni",
        "aliases": ["tapecaria"],
    },
    {
        "chave": "montagem de cadeiras",
        "nome": "MONTAGEM DE CADEIRAS",
        "meta": 1030.0,
        "unidade": "uni",
        "aliases": ["montagem cadeira", "montagem de cadeiras"],
    },
    {
        "chave": "cola de espuma",
        "nome": "COLA DE ESPUMA",
        "meta": 1920.0,
        "unidade": "uni",
        "aliases": ["cola de espuma"],
    },
    {
        "chave": "pintura uv",
        "nome": "PINTURA U.V.",
        "meta": 800.0,
        "unidade": "M²",
        "aliases": ["pintura uv", "uv"],
    },
]


def mapa_alias_metas_setor() -> dict[str, str]:
    mapa: dict[str, str] = {}
    for cadastro in METAS_DIARIAS_SETOR:
        chave = texto_limpo(cadastro["chave"])
        for alias in [chave, *cadastro.get("aliases", [])]:
            alias_norm = normalizar(alias)
            if alias_norm:
                mapa[alias_norm] = chave
    return mapa


_ALIAS_METAS_SETOR = mapa_alias_metas_setor()
_CADASTRO_META_POR_CHAVE = {
    texto_limpo(item["chave"]): item for item in METAS_DIARIAS_SETOR
}


def chave_meta_setor(valor: Any) -> str:
    """Converte nomes como '50 SECCIONADORA 2' para a chave da meta."""
    setor_sem_codigo = remover_codigo_inicial(valor)
    return _ALIAS_METAS_SETOR.get(normalizar(setor_sem_codigo), "")


def unidade_meta_legivel(unidade: str) -> str:
    if unidade == "M²":
        return "M²"
    if unidade == "m linear":
        return "METROS LINEARES"
    return "PEÇAS"


def formatar_valor_setor(valor: Any, unidade: str) -> str:
    try:
        numero = float(valor)
    except Exception:
        numero = 0.0
    casas = 0 if unidade == "uni" else 1
    return fmt_num(numero, casas)


def preparar_base_acumulado_setor(df: pd.DataFrame) -> pd.DataFrame:
    """Prepara os apontamentos manuais para o painel de metas por setor."""
    base = preparar_base_programado_equipamento(df)
    base["Chave Meta Setor"] = base["Setor"].map(chave_meta_setor)
    base["Data Acumulado"] = pd.to_datetime(base["Data Realizada DT"], errors="coerce")

    base["Unidade Meta"] = base["Chave Meta Setor"].map(
        lambda chave: texto_limpo(_CADASTRO_META_POR_CHAVE.get(chave, {}).get("unidade", ""))
    )
    base["Valor Apontado Setor"] = pd.to_numeric(
        base["Qtde Apontada Num"], errors="coerce"
    ).fillna(0).clip(lower=0)

    usa_metrica = base["Unidade Meta"].isin(["M²", "m linear"])
    base.loc[usa_metrica, "Valor Apontado Setor"] = pd.to_numeric(
        base.loc[usa_metrica, "Métrica Apontada Num"], errors="coerce"
    ).fillna(0).clip(lower=0)

    base["Medida necessária ausente?"] = (
        usa_metrica
        & (pd.to_numeric(base["Qtde Apontada Num"], errors="coerce").fillna(0) > 0)
        & ~base["Medida Válida?"].fillna(False)
    )
    return base


def limites_acumulado_setor(base: pd.DataFrame) -> tuple[date, date] | None:
    datas = pd.to_datetime(base.get("Data Acumulado"), errors="coerce").dropna()
    if datas.empty:
        return None
    return datas.min().date(), datas.max().date()


def escolher_periodo_acumulado_setor(
    base: pd.DataFrame,
) -> tuple[date | None, date | None, bool, str]:
    limites = limites_acumulado_setor(base)
    if limites is None:
        st.warning("Não existem apontamentos manuais com Data Realizada válida.")
        return None, None, False, ""

    data_min, data_max = limites
    dia_anterior = date.today() - timedelta(days=1)
    referencia_padrao = dia_anterior if data_min <= dia_anterior <= data_max else data_max

    modo = st.radio(
        "Período analisado",
        ["Dia específico", "Semana de trabalho", "Mês", "Período personalizado"],
        index=0,
        horizontal=True,
        key="acumulado_setor_modo_periodo",
    )

    if modo == "Dia específico":
        inicio = fim = st.date_input(
            "Data analisada",
            value=referencia_padrao,
            min_value=data_min,
            max_value=data_max,
            format="DD/MM/YYYY",
            key="acumulado_setor_dia",
        )
    elif modo == "Semana de trabalho":
        referencia = st.date_input(
            "Escolha uma data da semana",
            value=referencia_padrao,
            min_value=data_min,
            max_value=data_max,
            format="DD/MM/YYYY",
            key="acumulado_setor_semana",
        )
        inicio = referencia - timedelta(days=referencia.weekday())
        fim = inicio + timedelta(days=4)
    elif modo == "Mês":
        meses = list(
            pd.period_range(
                start=pd.Timestamp(data_min).to_period("M"),
                end=pd.Timestamp(data_max).to_period("M"),
                freq="M",
            )
        )
        periodo_padrao = pd.Timestamp(referencia_padrao).to_period("M")
        indice_padrao = meses.index(periodo_padrao) if periodo_padrao in meses else len(meses) - 1
        periodo_escolhido = st.selectbox(
            "Mês analisado",
            meses,
            index=indice_padrao,
            format_func=lambda periodo: periodo.strftime("%m/%Y"),
            key="acumulado_setor_mes",
        )
        inicio = periodo_escolhido.start_time.date()
        fim = periodo_escolhido.end_time.date()
    else:
        periodo = st.date_input(
            "Datas analisadas",
            value=(referencia_padrao, referencia_padrao),
            min_value=data_min,
            max_value=data_max,
            format="DD/MM/YYYY",
            key="acumulado_setor_periodo",
        )
        if isinstance(periodo, (tuple, list)) and len(periodo) == 2:
            inicio, fim = periodo
        elif isinstance(periodo, (tuple, list)) and len(periodo) == 1:
            inicio = fim = periodo[0]
        else:
            inicio = fim = periodo

    if inicio > fim:
        inicio, fim = fim, inicio

    incluir_fim_semana = st.checkbox(
        "Incluir sábados e domingos na tabela e na meta acumulada",
        value=False,
        key="acumulado_setor_fim_semana",
    )
    return inicio, fim, incluir_fim_semana, modo


def datas_painel_acumulado(
    inicio: date,
    fim: date,
    incluir_fim_semana: bool,
) -> list[date]:
    datas = [valor.date() for valor in pd.date_range(inicio, fim, freq="D")]
    if incluir_fim_semana:
        return datas
    uteis = [data_item for data_item in datas if data_item.weekday() < 5]
    # Se o usuário escolheu especificamente um sábado ou domingo, mantém a data.
    return uteis or datas


def montar_acumulado_diario_setor(
    base: pd.DataFrame,
    inicio: date,
    fim: date,
    incluir_fim_semana: bool,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Monta produção diária, acumulada, metas, lotes e resumo por setor."""
    datas = datas_painel_acumulado(inicio, fim, incluir_fim_semana)
    chaves = [texto_limpo(item["chave"]) for item in METAS_DIARIAS_SETOR]

    filtro = base[
        base["Data Acumulado"].notna()
        & base["Chave Meta Setor"].ne("")
        & (base["Data Acumulado"].dt.date >= inicio)
        & (base["Data Acumulado"].dt.date <= fim)
        & (pd.to_numeric(base["Qtde Apontada Num"], errors="coerce").fillna(0) > 0)
    ].copy()
    filtro["Data"] = filtro["Data Acumulado"].dt.date

    if not incluir_fim_semana:
        filtro = filtro[filtro["Data"].map(lambda valor: valor.weekday() < 5)]

    if filtro.empty:
        diario = pd.DataFrame(0.0, index=pd.Index(datas, name="Data"), columns=chaves)
        lotes = pd.DataFrame("", index=pd.Index(datas, name="Data"), columns=chaves)
    else:
        agrupado = (
            filtro.groupby(["Data", "Chave Meta Setor"], as_index=False)["Valor Apontado Setor"]
            .sum()
        )
        diario = agrupado.pivot(
            index="Data", columns="Chave Meta Setor", values="Valor Apontado Setor"
        )
        diario = diario.reindex(index=datas, columns=chaves).fillna(0.0)

        lotes_agrupados = (
            filtro.groupby(["Data", "Chave Meta Setor"])["OP/Lote"]
            .agg(juntar_lotes)
            .reset_index()
        )
        lotes = lotes_agrupados.pivot(
            index="Data", columns="Chave Meta Setor", values="OP/Lote"
        )
        lotes = lotes.reindex(index=datas, columns=chaves).fillna("")

    acumulado = diario.cumsum()
    dias_meta = len(datas)
    resumo_linhas: list[dict[str, Any]] = []
    for cadastro in METAS_DIARIAS_SETOR:
        chave = texto_limpo(cadastro["chave"])
        meta_diaria = float(cadastro["meta"])
        apontado = float(diario[chave].sum()) if chave in diario.columns else 0.0
        meta_periodo = meta_diaria * dias_meta
        percentual = (apontado / meta_periodo * 100) if meta_periodo > 0 else 0.0
        resumo_linhas.append(
            {
                "Setor": cadastro["nome"],
                "Unidade": unidade_meta_legivel(cadastro["unidade"]),
                "Meta Diária": meta_diaria,
                "Dias da Meta": dias_meta,
                "Meta Acumulada": meta_periodo,
                "Apontado Acumulado": apontado,
                "Saldo": apontado - meta_periodo,
                "% da Meta": percentual,
                "Lotes Apontados": juntar_lotes(
                    filtro.loc[filtro["Chave Meta Setor"].eq(chave), "OP/Lote"]
                ) if not filtro.empty else "",
            }
        )
    resumo = pd.DataFrame(resumo_linhas)
    return filtro, diario, acumulado, lotes, resumo


def classe_meta_painel(valor: float, meta: float) -> str:
    if valor <= 0:
        return "ac-sem"
    percentual = valor / meta * 100 if meta > 0 else 0.0
    if percentual >= 100:
        return "ac-ok"
    if percentual >= 80:
        return "ac-atencao"
    return "ac-abaixo"


def html_painel_acumulado_setor(
    valores: pd.DataFrame,
    modo_acumulado: bool,
) -> str:
    """Gera o quadro horizontal com metas, setores e valores por dia."""
    cabecalho_meta = [
        '<th class="ac-primeira ac-meta">META DIÁRIA</th>'
    ]
    cabecalho_setor = [
        '<th class="ac-primeira ac-setor">DATA</th>'
    ]

    for cadastro in METAS_DIARIAS_SETOR:
        meta_texto = (
            f'{formatar_valor_setor(cadastro["meta"], cadastro["unidade"])} '
            f'{unidade_meta_legivel(cadastro["unidade"])}'
        )
        cabecalho_meta.append(f'<th class="ac-meta">{escape(meta_texto)}</th>')
        cabecalho_setor.append(f'<th class="ac-setor">{escape(cadastro["nome"])}</th>')

    linhas: list[str] = []
    for posicao, (data_item, linha) in enumerate(valores.iterrows(), start=1):
        numero_dias = posicao if modo_acumulado else 1
        celulas = [
            f'<td class="ac-primeira ac-data"><strong>{data_item:%d/%m/%Y}</strong>'
            f'<span>{"Acumulado" if modo_acumulado else "Produção do dia"}</span></td>'
        ]
        for cadastro in METAS_DIARIAS_SETOR:
            chave = texto_limpo(cadastro["chave"])
            valor = float(linha.get(chave, 0) or 0)
            meta_referencia = float(cadastro["meta"]) * numero_dias
            percentual = valor / meta_referencia * 100 if meta_referencia > 0 else 0.0
            classe = classe_meta_painel(valor, meta_referencia)
            valor_texto = formatar_valor_setor(valor, cadastro["unidade"])
            celulas.append(
                f'<td class="ac-valor {classe}"><strong>{escape(valor_texto)}</strong>'
                f'<span>{fmt_num(percentual, 1)}%</span></td>'
            )
        linhas.append(f'<tr>{"".join(celulas)}</tr>')

    tabela = f"""
    <style>
        .ac-wrapper {{ width: 100%; overflow-x: auto; border: 1px solid #94a3b8;
            border-radius: 8px; background: #ffffff; margin-top: 8px; }}
        .ac-table {{ border-collapse: separate; border-spacing: 0; min-width: 2450px;
            width: 100%; table-layout: fixed; font-family: Arial, sans-serif; }}
        .ac-table th, .ac-table td {{ border-right: 1px solid #27364a;
            border-bottom: 1px solid #27364a; text-align: center; vertical-align: middle;
            min-width: 155px; padding: 8px 6px; }}
        .ac-table tr:last-child td {{ border-bottom: 0; }}
        .ac-table th:last-child, .ac-table td:last-child {{ border-right: 0; }}
        .ac-meta {{ background: #dc7300 !important; color: white !important;
            font-weight: 900; font-size: .94rem; height: 48px; }}
        .ac-setor {{ background: #0d8279 !important; color: white !important;
            font-weight: 900; font-size: .92rem; line-height: 1.18; height: 54px; }}
        .ac-primeira {{ position: sticky; left: 0; z-index: 3; min-width: 128px !important;
            width: 128px; }}
        .ac-meta.ac-primeira {{ z-index: 5; }}
        .ac-setor.ac-primeira {{ z-index: 5; }}
        .ac-data {{ background: #e2e8f0 !important; color: #0f172a !important; }}
        .ac-data strong {{ display: block; font-size: .90rem; }}
        .ac-data span {{ display: block; margin-top: 3px; font-size: .69rem;
            font-weight: 700; color: #475569 !important; }}
        .ac-valor {{ height: 58px; background: #cfcfcf !important; }}
        .ac-valor strong {{ display: block; font-size: 1rem; color: #0f172a !important; }}
        .ac-valor span {{ display: block; margin-top: 4px; font-size: .72rem;
            font-weight: 800; color: #334155 !important; }}
        .ac-sem {{ background: #d5d7da !important; }}
        .ac-ok {{ background: #c8f0d3 !important; }}
        .ac-atencao {{ background: #fff0b8 !important; }}
        .ac-abaixo {{ background: #ffd7d7 !important; }}
    </style>
    <div class="ac-wrapper">
      <table class="ac-table">
        <thead>
          <tr>{''.join(cabecalho_meta)}</tr>
          <tr>{''.join(cabecalho_setor)}</tr>
        </thead>
        <tbody>{''.join(linhas)}</tbody>
      </table>
    </div>
    """
    return tabela


def excel_acumulado_setor(
    diario: pd.DataFrame,
    acumulado: pd.DataFrame,
    lotes: pd.DataFrame,
    resumo: pd.DataFrame,
) -> bytes:
    saida = BytesIO()
    nomes = {
        texto_limpo(item["chave"]): item["nome"] for item in METAS_DIARIAS_SETOR
    }

    def preparar_matriz(df_origem: pd.DataFrame) -> pd.DataFrame:
        df_saida = df_origem.rename(columns=nomes).reset_index()
        df_saida["Data"] = pd.to_datetime(df_saida["Data"])
        return df_saida

    with pd.ExcelWriter(saida, engine="openpyxl") as writer:
        preparar_matriz(diario).to_excel(writer, index=False, sheet_name="Produção diária")
        preparar_matriz(acumulado).to_excel(writer, index=False, sheet_name="Acumulado")
        preparar_matriz(lotes).to_excel(writer, index=False, sheet_name="Lotes por dia")
        resumo.to_excel(writer, index=False, sheet_name="Resumo por setor")

        for ws in writer.book.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            ws.sheet_view.showGridLines = False
            for celula in ws[1]:
                celula.fill = PatternFill("solid", fgColor="0D8279")
                celula.font = Font(color="FFFFFF", bold=True)
                celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[1].height = 34
            for indice, coluna in enumerate(ws.columns, start=1):
                maior = max((len(texto_limpo(celula.value)) for celula in coluna), default=10)
                ws.column_dimensions[get_column_letter(indice)].width = min(max(maior + 2, 12), 42)
                for celula in coluna[1:]:
                    if isinstance(celula.value, (datetime, date)):
                        celula.number_format = "dd/mm/yyyy"
                    elif isinstance(celula.value, (int, float)):
                        celula.number_format = "#,##0.00"
                    celula.alignment = Alignment(vertical="center", wrap_text=False)

    saida.seek(0)
    return saida.getvalue()


def renderizar_acumulado_diario_setor(df: pd.DataFrame, caminho: Path) -> None:
    st.title("Acumulado Diário por Setor")
    st.caption(
        "Agrupa os apontamentos manuais pela Data Realizada e compara cada setor com "
        "a meta diária informada pelo PPCP. Para Seccionadoras e Pintura U.V. o valor "
        "é calculado em M²; para Coladeiras, em metros lineares; os demais setores usam peças."
    )

    base = preparar_base_acumulado_setor(df)
    inicio, fim, incluir_fim_semana, _modo_periodo = escolher_periodo_acumulado_setor(base)
    if inicio is None or fim is None:
        return

    setores_meta = {texto_limpo(item["chave"]) for item in METAS_DIARIAS_SETOR}
    setores_sem_meta = sorted(
        {
            texto_limpo(setor)
            for setor, chave in zip(base["Setor"], base["Chave Meta Setor"])
            if texto_limpo(setor) and not texto_limpo(chave)
        },
        key=chave_ordenacao_texto,
    )

    filtro, diario, acumulado, lotes, resumo = montar_acumulado_diario_setor(
        base, inicio, fim, incluir_fim_semana
    )
    del setores_meta

    modo_visualizacao = st.radio(
        "Valores exibidos nas linhas",
        ["Produção de cada dia", "Acumulado até cada dia"],
        index=1,
        horizontal=True,
        key="acumulado_setor_visualizacao",
    )
    modo_acumulado = modo_visualizacao == "Acumulado até cada dia"
    valores_painel = acumulado if modo_acumulado else diario

    total_apontado = float(resumo["Apontado Acumulado"].sum()) if not resumo.empty else 0.0
    setores_atingidos = int((resumo["% da Meta"] >= 100).sum()) if not resumo.empty else 0
    setores_com_producao = int((resumo["Apontado Acumulado"] > 0).sum()) if not resumo.empty else 0
    dias_meta = int(resumo["Dias da Meta"].max()) if not resumo.empty else 0

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Dias considerados", dias_meta)
    c2.metric("Setores com produção", f"{setores_com_producao}/{len(METAS_DIARIAS_SETOR)}")
    c3.metric("Setores que atingiram a meta", f"{setores_atingidos}/{len(METAS_DIARIAS_SETOR)}")
    c4.metric("Linhas apontadas consideradas", len(filtro))

    st.caption(
        f"Período: {inicio:%d/%m/%Y} a {fim:%d/%m/%Y}. "
        "Verde = meta atingida; amarelo = pelo menos 80%; vermelho = abaixo de 80%; cinza = sem produção."
    )
    st.markdown(
        html_painel_acumulado_setor(valores_painel, modo_acumulado),
        unsafe_allow_html=True,
    )

    medidas_ausentes = base[
        base["Medida necessária ausente?"].fillna(False)
        & base["Data Acumulado"].notna()
        & (base["Data Acumulado"].dt.date >= inicio)
        & (base["Data Acumulado"].dt.date <= fim)
    ].copy()
    if not medidas_ausentes.empty:
        st.warning(
            f"{len(medidas_ausentes)} linha(s) apontada(s) de setores medidos em M² ou metros "
            "lineares não possuem medida identificável. Essas linhas ficaram com valor métrico zero."
        )
        with st.expander("Ver linhas sem medida"):
            colunas = [
                "Data Realizada", "Setor", "OP/Lote", "Código Peça", "Descrição Peça",
                "Qtde Realizada", "Unidade Meta", "Origem Medida", "Medida Catálogo",
            ]
            st.dataframe(
                medidas_ausentes[[c for c in colunas if c in medidas_ausentes.columns]],
                use_container_width=True,
                hide_index=True,
            )

    aba_resumo, aba_diario, aba_lotes = st.tabs(
        ["Resumo por setor", "Valores diários", "Lotes apontados"]
    )

    with aba_resumo:
        st.dataframe(
            resumo,
            use_container_width=True,
            hide_index=True,
            height=540,
            column_config={
                "Meta Diária": st.column_config.NumberColumn("Meta diária", format="%.2f"),
                "Meta Acumulada": st.column_config.NumberColumn("Meta acumulada", format="%.2f"),
                "Apontado Acumulado": st.column_config.NumberColumn("Apontado acumulado", format="%.2f"),
                "Saldo": st.column_config.NumberColumn("Saldo", format="%.2f"),
                "% da Meta": st.column_config.NumberColumn("% da meta", format="%.1f%%"),
                "Lotes Apontados": st.column_config.TextColumn("Lotes apontados", width="large"),
            },
        )

    with aba_diario:
        nomes = {texto_limpo(item["chave"]): item["nome"] for item in METAS_DIARIAS_SETOR}
        tabela_diaria = diario.rename(columns=nomes).reset_index()
        tabela_diaria["Data"] = pd.to_datetime(tabela_diaria["Data"]).dt.strftime("%d/%m/%Y")
        st.dataframe(tabela_diaria, use_container_width=True, hide_index=True, height=520)

    with aba_lotes:
        nomes = {texto_limpo(item["chave"]): item["nome"] for item in METAS_DIARIAS_SETOR}
        tabela_lotes = lotes.rename(columns=nomes).reset_index()
        tabela_lotes["Data"] = pd.to_datetime(tabela_lotes["Data"]).dt.strftime("%d/%m/%Y")
        st.dataframe(
            tabela_lotes,
            use_container_width=True,
            hide_index=True,
            height=520,
            column_config={
                coluna: st.column_config.TextColumn(coluna, width="large")
                for coluna in tabela_lotes.columns
                if coluna != "Data"
            },
        )

    if setores_sem_meta:
        with st.expander("Setores encontrados sem meta cadastrada"):
            st.caption(
                "Esses setores existem na planilha, mas não aparecem no modelo de metas enviado. "
                "Por isso não entram no painel até que uma meta seja informada."
            )
            st.write(", ".join(setores_sem_meta))

    arquivo_excel = excel_acumulado_setor(diario, acumulado, lotes, resumo)
    st.download_button(
        "Baixar acumulado por setor em Excel",
        data=arquivo_excel,
        file_name=f"acumulado_setores_{inicio:%Y%m%d}_{fim:%Y%m%d}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_acumulado_setor_xlsx",
        use_container_width=True,
    )
    st.caption(f"Fonte dos dados: {caminho.name} — somente apontamentos manuais da planilha.")


def preparar_base_calendario_semanal(df: pd.DataFrame) -> pd.DataFrame:
    """Classifica o programado em realizado no prazo, após o prazo e pendente."""
    base = preparar_base_producao_setor(df)
    programada = base["Qtde Programada Num"].clip(lower=0)
    realizada = base["Qtde Realizada Considerada Num"].clip(lower=0)
    base["Qtde Finalizada Classificada Num"] = pd.concat(
        [programada, realizada], axis=1
    ).min(axis=1)

    tem_data_realizada = base["Data Realizada DT"].notna()
    tem_data_programada = base["Data Programada DT"].notna()
    finalizada_com_data = base["Qtde Finalizada Classificada Num"].where(
        tem_data_realizada & tem_data_programada,
        0.0,
    )
    no_prazo = (
        tem_data_realizada
        & tem_data_programada
        & (base["Data Realizada DT"] <= base["Data Programada DT"])
    )
    apos_prazo = (
        tem_data_realizada
        & tem_data_programada
        & (base["Data Realizada DT"] > base["Data Programada DT"])
    )
    base["Realizada no Prazo Num"] = finalizada_com_data.where(no_prazo, 0.0)
    base["Realizada Após Num"] = finalizada_com_data.where(apos_prazo, 0.0)
    base["Pendente Num"] = (
        base["Qtde Programada Num"]
        - base["Realizada no Prazo Num"]
        - base["Realizada Após Num"]
    ).clip(lower=0)
    return base


def limites_datas_calendario(
    base: pd.DataFrame,
    paradas: pd.DataFrame,
) -> tuple[date, date] | None:
    series: list[pd.Series] = []
    if "Data Programada DT" in base.columns:
        series.append(base["Data Programada DT"])
    if not paradas.empty and "Data Parada Valor" in paradas.columns:
        series.append(pd.to_datetime(paradas["Data Parada Valor"], errors="coerce"))
    if not series:
        return None
    datas = pd.concat(series, ignore_index=True).dropna()
    if datas.empty:
        return None
    return datas.min().date(), datas.max().date()


def inicio_da_semana(data_referencia: date) -> date:
    return data_referencia - timedelta(days=data_referencia.weekday())


def montar_paradas_calendario(
    paradas: pd.DataFrame,
    setores: list[str],
    inicio: date,
    fim: date,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    col_diario = ["Setor", "Data Parada Dia", "Tempo Parada (min)", "Quantidade de Paradas", "Itens"]
    col_semanal = ["Setor", "Tempo Parada Semana (min)", "Quantidade de Paradas"]
    if paradas.empty:
        return pd.DataFrame(columns=col_diario), pd.DataFrame(columns=col_semanal)

    dados = paradas.copy()
    dados["Data Parada DT"] = pd.to_datetime(dados["Data Parada Valor"], errors="coerce")
    dados["Tempo Total Parada (min)"] = to_num(dados["Tempo Total Parada (min)"].copy()).clip(lower=0)
    dados = dados[
        dados["Setor"].isin(setores)
        & entre_datas(dados["Data Parada DT"], inicio, fim)
    ].copy()
    if dados.empty:
        return pd.DataFrame(columns=col_diario), pd.DataFrame(columns=col_semanal)

    dados["Data Parada Dia"] = dados["Data Parada DT"].dt.date
    linhas_diarias: list[dict[str, Any]] = []
    for (setor, dia), grupo in dados.groupby(["Setor", "Data Parada Dia"], sort=False):
        itens: list[dict[str, Any]] = []
        consolidados = (
            grupo.groupby(["Máquina/Posto", "Motivo da Parada"], dropna=False)["Tempo Total Parada (min)"]
            .sum()
            .reset_index()
        )
        for _, item in consolidados.iterrows():
            itens.append(
                {
                    "maquina": texto_limpo(item["Máquina/Posto"]) or "Sem máquina informada",
                    "motivo": texto_limpo(item["Motivo da Parada"]) or "Sem motivo informado",
                    "tempo_min": float(item["Tempo Total Parada (min)"]),
                }
            )
        linhas_diarias.append(
            {
                "Setor": setor,
                "Data Parada Dia": dia,
                "Tempo Parada (min)": float(grupo["Tempo Total Parada (min)"].sum()),
                "Quantidade de Paradas": int(len(grupo)),
                "Itens": itens,
            }
        )

    diario = pd.DataFrame(linhas_diarias, columns=col_diario)
    semanal = (
        dados.groupby("Setor", as_index=False)
        .agg(
            **{
                "Tempo Parada Semana (min)": ("Tempo Total Parada (min)", "sum"),
                "Quantidade de Paradas": ("Tempo Total Parada (min)", "size"),
            }
        )
    )
    return diario, semanal


def montar_dados_calendario_semanal(
    base: pd.DataFrame,
    setores_sel: list[str],
    inicio: date,
    fim: date,
    somente_movimento: bool,
    setores_com_paradas: set[str] | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    todos_setores = opcoes_unicas(base, "Setor")
    setores_base = setores_sel if setores_sel else todos_setores
    setores_com_paradas = setores_com_paradas or set()

    periodo = base[
        base["Setor"].isin(setores_base)
        & entre_datas(base["Data Programada DT"], inicio, fim)
    ].copy()
    periodo["Data Programada Dia"] = periodo["Data Programada DT"].dt.date
    colunas_soma = [
        "Qtde Programada Num",
        "Realizada no Prazo Num",
        "Realizada Após Num",
        "Pendente Num",
    ]

    if periodo.empty:
        diario = pd.DataFrame(
            columns=["Setor", "Data Programada Dia", "Lotes", *colunas_soma, "% Atraso"]
        )
        resumo = pd.DataFrame(
            {
                "Setor": setores_base,
                "Lotes da Semana": "",
                "Programada": 0.0,
                "Realizada no Prazo": 0.0,
                "Realizada Após": 0.0,
                "Pendente": 0.0,
                "% Atraso Semanal": 0.0,
            }
        )
        if somente_movimento:
            resumo = resumo[resumo["Setor"].isin(setores_com_paradas)].copy()
        return diario, resumo, resumo["Setor"].tolist()

    diario = periodo.groupby(["Setor", "Data Programada Dia"], as_index=False)[colunas_soma].sum()
    lotes_diarios = (
        periodo.groupby(["Setor", "Data Programada Dia"], as_index=False)["OP/Lote"]
        .agg(juntar_lotes)
        .rename(columns={"OP/Lote": "Lotes"})
    )
    diario = diario.merge(lotes_diarios, on=["Setor", "Data Programada Dia"], how="left")
    diario["Lotes"] = diario["Lotes"].fillna("")
    diario["% Atraso"] = 0.0
    mask_diario = diario["Qtde Programada Num"] > 0
    diario.loc[mask_diario, "% Atraso"] = (
        diario.loc[mask_diario, "Realizada Após Num"]
        / diario.loc[mask_diario, "Qtde Programada Num"]
        * 100
    )

    resumo_agrupado = periodo.groupby("Setor")[colunas_soma].sum()
    lotes_semanais = periodo.groupby("Setor")["OP/Lote"].agg(juntar_lotes).rename("Lotes da Semana")
    resumo = pd.DataFrame(index=setores_base).join(resumo_agrupado).join(lotes_semanais)
    resumo[colunas_soma] = resumo[colunas_soma].fillna(0)
    resumo["Lotes da Semana"] = resumo["Lotes da Semana"].fillna("")
    resumo.index.name = "Setor"
    resumo = resumo.reset_index().rename(
        columns={
            "Qtde Programada Num": "Programada",
            "Realizada no Prazo Num": "Realizada no Prazo",
            "Realizada Após Num": "Realizada Após",
            "Pendente Num": "Pendente",
        }
    )
    resumo["% Atraso Semanal"] = 0.0
    mask_semana = resumo["Programada"] > 0
    resumo.loc[mask_semana, "% Atraso Semanal"] = (
        resumo.loc[mask_semana, "Realizada Após"] / resumo.loc[mask_semana, "Programada"] * 100
    )

    if somente_movimento:
        resumo = resumo[
            (resumo["Programada"] > 0) | resumo["Setor"].isin(setores_com_paradas)
        ].copy()
    return diario, resumo, resumo["Setor"].tolist()


def classe_celula_calendario(
    programada: float,
    realizada_apos: float,
    pendente: float,
    tempo_parada_min: float = 0.0,
) -> str:
    if programada <= 0:
        return "cal-com-parada" if tempo_parada_min > 0 else "cal-sem-programacao"
    if realizada_apos > 0:
        return "cal-atraso"
    if pendente > 0:
        return "cal-pendente"
    return "cal-ok"


def html_paradas_calendario(
    itens: list[dict[str, Any]] | None,
    tempo_total_min: float,
    quantidade: int,
    semanal: bool = False,
) -> str:
    if tempo_total_min <= 0 and not itens:
        return ""
    titulo = "Paradas na semana" if semanal else "Paradas"
    cabecalho = (
        f'<div class="cal-paradas-titulo"><span>⏸ {titulo}</span>'
        f'<strong>{formatar_tempo_minutos(tempo_total_min)}</strong></div>'
    )
    if semanal:
        sufixo = "registro" if quantidade == 1 else "registros"
        return (
            '<div class="cal-paradas">'
            + cabecalho
            + f'<div class="cal-parada-resumo">{quantidade} {sufixo}</div>'
            + '</div>'
        )

    linhas: list[str] = []
    for item in itens or []:
        maquina = escape(texto_limpo(item.get("maquina")) or "Sem máquina informada")
        motivo = escape(texto_limpo(item.get("motivo")) or "Sem motivo informado")
        tempo_item = formatar_tempo_minutos(item.get("tempo_min", 0))
        linhas.append(
            '<div class="cal-parada-item">'
            f'<div><strong>{maquina}</strong><span>{motivo}</span></div>'
            f'<b>{escape(tempo_item)}</b>'
            '</div>'
        )
    return '<div class="cal-paradas">' + cabecalho + ''.join(linhas) + '</div>'


def html_metricas_calendario(
    programada: float,
    no_prazo: float,
    apos: float,
    pendente: float,
    percentual_atraso: float,
    lotes: str = "",
    semanal: bool = False,
    paradas: list[dict[str, Any]] | None = None,
    tempo_parada_min: float = 0.0,
    quantidade_paradas: int = 0,
) -> str:
    classe = classe_celula_calendario(programada, apos, pendente, tempo_parada_min)
    if programada <= 0 and tempo_parada_min <= 0:
        return f'<div class="cal-celula {classe}"><span class="cal-vazio">—</span></div>'

    partes: list[str] = [f'<div class="cal-celula {classe}">']
    if programada > 0:
        titulo_percentual = "% atraso semanal" if semanal else "% atraso"
        titulo_lotes = "Lotes da semana" if semanal else "Lotes"
        lotes_html = escape(texto_limpo(lotes)) or "—"
        partes.extend(
            [
                f'<div class="cal-lotes"><span>{titulo_lotes}</span><strong>{lotes_html}</strong></div>',
                f'<div class="cal-linha"><span>Programada</span><strong>{fmt_qtd(programada)}</strong></div>',
                f'<div class="cal-linha"><span>No prazo</span><strong>{fmt_qtd(no_prazo)}</strong></div>',
                f'<div class="cal-linha"><span>Após</span><strong>{fmt_qtd(apos)}</strong></div>',
                f'<div class="cal-linha"><span>Pendente</span><strong>{fmt_qtd(pendente)}</strong></div>',
                f'<div class="cal-linha cal-percentual"><span>{titulo_percentual}</span>'
                f'<strong>{fmt_num(percentual_atraso, 1)}%</strong></div>',
            ]
        )
    else:
        partes.append('<div class="cal-sem-producao-texto">Sem programação de produção</div>')

    partes.append(
        html_paradas_calendario(
            paradas,
            tempo_parada_min,
            quantidade_paradas,
            semanal=semanal,
        )
    )
    partes.append('</div>')
    return ''.join(partes)

# Nomes dos dias exibidos no Calendário Semanal
DIAS_SEMANA_PT = [
    "Segunda",
    "Terça",
    "Quarta",
    "Quinta",
    "Sexta",
]


def _carregar_fonte_pil(tamanho: int, negrito: bool = False) -> ImageFont.ImageFont:
    candidatos = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if negrito else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf" if negrito else "/usr/share/fonts/dejavu/DejaVuSans.ttf",
        "DejaVuSans-Bold.ttf" if negrito else "DejaVuSans.ttf",
    ]
    for fonte in candidatos:
        try:
            return ImageFont.truetype(fonte, tamanho)
        except Exception:
            continue
    return ImageFont.load_default()


def _quebrar_texto_pil(draw: ImageDraw.ImageDraw, texto: str, fonte: ImageFont.ImageFont, largura_max: int) -> list[str]:
    texto = texto_limpo(texto)
    if not texto:
        return [""]
    linhas: list[str] = []
    for paragrafo in str(texto).splitlines() or [str(texto)]:
        palavras = paragrafo.split()
        if not palavras:
            linhas.append("")
            continue
        atual = palavras[0]
        for palavra in palavras[1:]:
            tentativa = f"{atual} {palavra}"
            caixa = draw.textbbox((0, 0), tentativa, font=fonte)
            if (caixa[2] - caixa[0]) <= largura_max:
                atual = tentativa
            else:
                linhas.append(atual)
                atual = palavra
        linhas.append(atual)
    return linhas or [""]


def _altura_linha_pil(draw: ImageDraw.ImageDraw, fonte: ImageFont.ImageFont) -> int:
    caixa = draw.textbbox((0, 0), "Ag", font=fonte)
    return caixa[3] - caixa[1]


def _montar_lookups_calendario(
    diario: pd.DataFrame,
    resumo: pd.DataFrame,
    paradas_diarias: pd.DataFrame,
    paradas_semanais: pd.DataFrame,
) -> tuple[dict[tuple[str, date], dict[str, Any]], dict[tuple[str, date], dict[str, Any]], dict[str, dict[str, Any]], dict[str, dict[str, Any]]]:
    lookup_diario: dict[tuple[str, date], dict[str, Any]] = {}
    for _, linha in diario.iterrows():
        lookup_diario[(str(linha["Setor"]), linha["Data Programada Dia"])] = {
            "lotes": texto_limpo(linha.get("Lotes", "")),
            "programada": float(linha["Qtde Programada Num"]),
            "no_prazo": float(linha["Realizada no Prazo Num"]),
            "apos": float(linha["Realizada Após Num"]),
            "pendente": float(linha["Pendente Num"]),
            "percentual": float(linha["% Atraso"]),
        }
    lookup_paradas: dict[tuple[str, date], dict[str, Any]] = {}
    for _, linha in paradas_diarias.iterrows():
        lookup_paradas[(str(linha["Setor"]), linha["Data Parada Dia"])] = {
            "tempo": float(linha["Tempo Parada (min)"]),
            "quantidade": int(linha["Quantidade de Paradas"]),
            "itens": linha["Itens"],
        }
    lookup_resumo = resumo.set_index("Setor").to_dict(orient="index") if not resumo.empty else {}
    lookup_paradas_semana = (
        paradas_semanais.set_index("Setor").to_dict(orient="index")
        if not paradas_semanais.empty
        else {}
    )
    return lookup_diario, lookup_paradas, lookup_resumo, lookup_paradas_semana


def _gerar_linhas_celula_calendario_png(
    programada: float,
    no_prazo: float,
    apos: float,
    pendente: float,
    percentual_atraso: float,
    lotes: str = "",
    semanal: bool = False,
    paradas: list[dict[str, Any]] | None = None,
    tempo_parada_min: float = 0.0,
    quantidade_paradas: int = 0,
) -> list[tuple[str, str]]:
    linhas: list[tuple[str, str]] = []
    if programada <= 0 and tempo_parada_min <= 0:
        return [("normal", "—")]
    if programada > 0:
        linhas.extend(
            [
                ("label", "Lotes da semana:" if semanal else "Lotes:"),
                ("value", texto_limpo(lotes) or "—"),
                ("normal", f"Programada: {fmt_qtd(programada)}"),
                ("normal", f"No prazo: {fmt_qtd(no_prazo)}"),
                ("normal", f"Após: {fmt_qtd(apos)}"),
                ("normal", f"Pendente: {fmt_qtd(pendente)}"),
                ("strong", f"% atraso semanal: {fmt_num(percentual_atraso, 1)}%" if semanal else f"% atraso: {fmt_num(percentual_atraso, 1)}%"),
            ]
        )
    else:
        linhas.append(("normal", "Sem programação de produção"))
    if tempo_parada_min > 0 or paradas:
        linhas.append(("label", "Paradas na semana:" if semanal else "Paradas:"))
        linhas.append(("strong", formatar_tempo_minutos(tempo_parada_min)))
        if semanal:
            sufixo = "registro" if quantidade_paradas == 1 else "registros"
            linhas.append(("normal", f"{quantidade_paradas} {sufixo}"))
        else:
            for item in paradas or []:
                maquina = texto_limpo(item.get("maquina")) or "Sem máquina"
                motivo = texto_limpo(item.get("motivo")) or "Sem motivo"
                tempo_item = formatar_tempo_minutos(item.get("tempo_min", 0))
                linhas.append(("normal", f"{maquina} — {motivo} ({tempo_item})"))
    return linhas


def gerar_png_calendario_semanal(
    diario: pd.DataFrame,
    resumo: pd.DataFrame,
    paradas_diarias: pd.DataFrame,
    paradas_semanais: pd.DataFrame,
    setores: list[str],
    inicio: date,
    fim: date,
) -> bytes:
    datas_semana = [inicio + timedelta(days=i) for i in range(5)]
    lookup_diario, lookup_paradas, lookup_resumo, lookup_paradas_semana = _montar_lookups_calendario(
        diario, resumo, paradas_diarias, paradas_semanais
    )

    total_programado = float(resumo["Programada"].sum()) if not resumo.empty else 0.0
    total_no_prazo = float(resumo["Realizada no Prazo"].sum()) if not resumo.empty else 0.0
    total_apos = float(resumo["Realizada Após"].sum()) if not resumo.empty else 0.0
    total_pendente = float(resumo["Pendente"].sum()) if not resumo.empty else 0.0
    total_parada = float(paradas_semanais["Tempo Parada Semana (min)"].sum()) if not paradas_semanais.empty else 0.0
    percentual_atraso = total_apos / total_programado * 100 if total_programado > 0 else 0.0

    cores = {
        "bg": "#ffffff",
        "grid": "#d7dee8",
        "header": "#e8edf4",
        "setor": "#f1f5f9",
        "resumo": "#eef4ff",
        "texto": "#172033",
        "subtexto": "#475569",
        "ok": "#eaf8ef",
        "pendente": "#fff3e4",
        "atraso": "#fdebec",
        "parada": "#f5f1ff",
        "sem": "#f1f4f8",
        "parada_box": "#ede9fe",
    }

    largura_setor = 300
    largura_dia = 295
    largura_resumo = 330
    margem = 28
    gap_metricas = 12
    altura_titulo = 66
    altura_subtitulo = 34
    altura_metricas = 90
    altura_header = 56
    padding_celula = 12

    largura_total = margem * 2 + largura_setor + len(datas_semana) * largura_dia + largura_resumo

    img_aux = Image.new("RGB", (largura_total, 2000), cores["bg"])
    draw_aux = ImageDraw.Draw(img_aux)
    fonte_titulo = _carregar_fonte_pil(28, negrito=True)
    fonte_subtitulo = _carregar_fonte_pil(18, negrito=False)
    fonte_cab = _carregar_fonte_pil(18, negrito=True)
    fonte_setor = _carregar_fonte_pil(18, negrito=True)
    fonte_label = _carregar_fonte_pil(15, negrito=True)
    fonte_normal = _carregar_fonte_pil(15, negrito=False)
    fonte_strong = _carregar_fonte_pil(15, negrito=True)
    fonte_metrica_rot = _carregar_fonte_pil(15, negrito=True)
    fonte_metrica_val = _carregar_fonte_pil(18, negrito=True)
    h_label = _altura_linha_pil(draw_aux, fonte_label)
    h_normal = _altura_linha_pil(draw_aux, fonte_normal)
    h_strong = _altura_linha_pil(draw_aux, fonte_strong)

    metricas = [
        ("Programada na semana", fmt_qtd(total_programado)),
        ("Realizada no prazo", fmt_qtd(total_no_prazo)),
        ("Realizada após", fmt_qtd(total_apos)),
        ("Pendente", fmt_qtd(total_pendente)),
        ("% atraso semanal", f"{fmt_num(percentual_atraso, 1)}%"),
        ("Tempo parado", formatar_tempo_minutos(total_parada)),
    ]

    estrutura_linhas: list[dict[str, Any]] = []
    altura_corpo = 0
    largura_conteudo_normal = largura_dia - padding_celula * 2 - 4
    largura_conteudo_resumo = largura_resumo - padding_celula * 2 - 4
    largura_setor_texto = largura_setor - padding_celula * 2

    for setor in setores:
        celulas_row = []
        alturas_candidatas = []

        setor_linhas = _quebrar_texto_pil(draw_aux, setor, fonte_setor, largura_setor_texto)
        altura_setor = padding_celula * 2 + len(setor_linhas) * (h_strong + 2)
        alturas_candidatas.append(max(altura_setor, 70))

        for dia in datas_semana:
            valores = lookup_diario.get(
                (setor, dia),
                {"lotes": "", "programada": 0.0, "no_prazo": 0.0, "apos": 0.0, "pendente": 0.0, "percentual": 0.0},
            )
            parada = lookup_paradas.get((setor, dia), {"tempo": 0.0, "quantidade": 0, "itens": []})
            classe = classe_celula_calendario(
                valores["programada"], valores["apos"], valores["pendente"], parada["tempo"]
            )
            linhas = _gerar_linhas_celula_calendario_png(
                valores["programada"], valores["no_prazo"], valores["apos"], valores["pendente"], valores["percentual"],
                lotes=valores["lotes"], paradas=parada["itens"], tempo_parada_min=parada["tempo"], quantidade_paradas=parada["quantidade"],
            )
            linhas_quebradas: list[tuple[str, str]] = []
            altura = padding_celula * 2
            for estilo, texto in linhas:
                fonte = fonte_label if estilo == "label" else fonte_strong if estilo == "strong" else fonte_normal
                altura_fonte = h_label if estilo == "label" else h_strong if estilo == "strong" else h_normal
                for parte in _quebrar_texto_pil(draw_aux, texto, fonte, largura_conteudo_normal):
                    linhas_quebradas.append((estilo, parte))
                    altura += altura_fonte + 3
                altura += 1
            altura += 4
            alturas_candidatas.append(max(altura, 170))
            celulas_row.append({"tipo": "dia", "classe": classe, "linhas": linhas_quebradas, "dia": dia})

        semana = lookup_resumo.get(
            setor,
            {"Lotes da Semana": "", "Programada": 0.0, "Realizada no Prazo": 0.0, "Realizada Após": 0.0, "Pendente": 0.0, "% Atraso Semanal": 0.0},
        )
        parada_semana = lookup_paradas_semana.get(setor, {"Tempo Parada Semana (min)": 0.0, "Quantidade de Paradas": 0})
        classe_resumo = classe_celula_calendario(
            float(semana["Programada"]), float(semana["Realizada Após"]), float(semana["Pendente"]), float(parada_semana["Tempo Parada Semana (min)"])
        )
        linhas_resumo = _gerar_linhas_celula_calendario_png(
            float(semana["Programada"]), float(semana["Realizada no Prazo"]), float(semana["Realizada Após"]),
            float(semana["Pendente"]), float(semana["% Atraso Semanal"]), lotes=texto_limpo(semana.get("Lotes da Semana", "")),
            semanal=True, tempo_parada_min=float(parada_semana["Tempo Parada Semana (min)"]), quantidade_paradas=int(parada_semana["Quantidade de Paradas"]),
        )
        linhas_resumo_quebradas: list[tuple[str, str]] = []
        altura_resumo = padding_celula * 2
        for estilo, texto in linhas_resumo:
            fonte = fonte_label if estilo == "label" else fonte_strong if estilo == "strong" else fonte_normal
            altura_fonte = h_label if estilo == "label" else h_strong if estilo == "strong" else h_normal
            for parte in _quebrar_texto_pil(draw_aux, texto, fonte, largura_conteudo_resumo):
                linhas_resumo_quebradas.append((estilo, parte))
                altura_resumo += altura_fonte + 3
            altura_resumo += 1
        altura_resumo += 4
        alturas_candidatas.append(max(altura_resumo, 170))
        celulas_row.append({"tipo": "resumo", "classe": classe_resumo, "linhas": linhas_resumo_quebradas})

        altura_linha = max(170, max(alturas_candidatas))
        estrutura_linhas.append({"setor": setor, "setor_linhas": setor_linhas, "altura": altura_linha, "celulas": celulas_row})
        altura_corpo += altura_linha

    altura_total = margem * 2 + altura_titulo + altura_subtitulo + altura_metricas + altura_header + altura_corpo + 20
    imagem = Image.new("RGB", (largura_total, altura_total), cores["bg"])
    draw = ImageDraw.Draw(imagem)

    y = margem
    draw.text((margem, y), "Calendário Semanal dos Setores", fill=cores["texto"], font=fonte_titulo)
    y += altura_titulo - 16
    draw.text(
        (margem, y),
        f"Período exibido: {inicio:%d/%m/%Y} a {fim:%d/%m/%Y} | Setores: {len(setores)}",
        fill=cores["subtexto"],
        font=fonte_subtitulo,
    )
    y += altura_subtitulo

    largura_cartao = (largura_total - margem * 2 - gap_metricas * 5) // 6
    for i, (rotulo, valor) in enumerate(metricas):
        x0 = margem + i * (largura_cartao + gap_metricas)
        y0 = y
        x1 = x0 + largura_cartao
        y1 = y0 + altura_metricas - 10
        draw.rounded_rectangle((x0, y0, x1, y1), radius=10, fill="#f8fafc", outline=cores["grid"], width=1)
        draw.text((x0 + 12, y0 + 12), rotulo, fill=cores["subtexto"], font=fonte_metrica_rot)
        draw.text((x0 + 12, y0 + 42), valor, fill=cores["texto"], font=fonte_metrica_val)
    y += altura_metricas

    x = margem
    headers = [("Setor", largura_setor)]
    for idx, dia in enumerate(datas_semana):
        headers.append((f"{DIAS_SEMANA_PT[idx]}\n{dia:%d/%m/%Y}", largura_dia))
    headers.append(("Resumo semanal", largura_resumo))

    for titulo, largura in headers:
        draw.rectangle((x, y, x + largura, y + altura_header), fill=cores["header"], outline=cores["grid"], width=1)
        partes = titulo.split("\n")
        if len(partes) == 2:
            w1 = draw.textbbox((0, 0), partes[0], font=fonte_cab)[2]
            draw.text((x + (largura - w1) / 2, y + 7), partes[0], fill=cores["texto"], font=fonte_cab)
            w2 = draw.textbbox((0, 0), partes[1], font=fonte_subtitulo)[2]
            draw.text((x + (largura - w2) / 2, y + 30), partes[1], fill=cores["subtexto"], font=fonte_subtitulo)
        else:
            w = draw.textbbox((0, 0), titulo, font=fonte_cab)[2]
            draw.text((x + (largura - w) / 2, y + 16), titulo, fill=cores["texto"], font=fonte_cab)
        x += largura
    y += altura_header

    fill_map = {
        "cal-ok": cores["ok"],
        "cal-pendente": cores["pendente"],
        "cal-atraso": cores["atraso"],
        "cal-com-parada": cores["parada"],
        "cal-sem-programacao": cores["sem"],
    }

    for linha in estrutura_linhas:
        altura_linha = linha["altura"]
        x = margem
        draw.rectangle((x, y, x + largura_setor, y + altura_linha), fill=cores["setor"], outline=cores["grid"], width=1)
        texto_y = y + padding_celula
        for parte in linha["setor_linhas"]:
            draw.text((x + padding_celula, texto_y), parte, fill=cores["texto"], font=fonte_setor)
            texto_y += h_strong + 3
        x += largura_setor

        for idx, celula in enumerate(linha["celulas"]):
            largura = largura_resumo if celula.get("tipo") == "resumo" else largura_dia
            if celula.get("tipo") == "resumo":
                fundo = cores["resumo"]
            else:
                fundo = fill_map.get(celula["classe"], cores["bg"])
            draw.rectangle((x, y, x + largura, y + altura_linha), fill="#ffffff", outline=cores["grid"], width=1)
            draw.rounded_rectangle((x + 6, y + 6, x + largura - 6, y + altura_linha - 6), radius=10, fill=fundo, outline="#e2e8f0", width=1)
            if celula["classe"] in fill_map:
                draw.rounded_rectangle((x + 6, y + 6, x + 14, y + altura_linha - 6), radius=6, fill={
                    "cal-ok": "#15803d",
                    "cal-pendente": "#c96a00",
                    "cal-atraso": "#d11f2f",
                    "cal-com-parada": "#6d28d9",
                    "cal-sem-programacao": "#8293aa",
                }.get(celula["classe"], "#8293aa"))
            texto_x = x + padding_celula + 8
            texto_y = y + padding_celula
            for estilo, texto in celula["linhas"]:
                fonte = fonte_label if estilo == "label" else fonte_strong if estilo == "strong" else fonte_normal
                cor = "#4c1d95" if estilo == "label" and texto.startswith("Paradas") else cores["texto"]
                if estilo == "label" and texto.endswith(":"):
                    cor = cores["subtexto"]
                draw.text((texto_x, texto_y), texto, fill=cor, font=fonte)
                altura_fonte = h_label if estilo == "label" else h_strong if estilo == "strong" else h_normal
                texto_y += altura_fonte + 4
            x += largura
        y += altura_linha

    rodape = (
        "Legenda: concluído no prazo, pendente, realizado após a Data Programada e células com paradas registradas. "
        "Arquivo gerado em alta resolução para impressão e compartilhamento."
    )
    draw.text((margem, altura_total - margem), rodape, fill=cores["subtexto"], font=fonte_subtitulo)

    buffer = BytesIO()
    imagem.save(buffer, format="PNG", optimize=True)
    return buffer.getvalue()


def gerar_pdf_calendario_semanal(
    diario: pd.DataFrame,
    resumo: pd.DataFrame,
    paradas_diarias: pd.DataFrame,
    paradas_semanais: pd.DataFrame,
    setores: list[str],
    inicio: date,
    fim: date,
) -> bytes:
    """Gera o calendário em PDF vetorial, sem converter uma imagem PNG."""
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A3, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        LongTable,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    lookup_diario, lookup_paradas, lookup_resumo, lookup_paradas_semana = _montar_lookups_calendario(
        diario, resumo, paradas_diarias, paradas_semanais
    )
    datas_semana = [inicio + timedelta(days=i) for i in range(5)]

    total_programado = float(resumo["Programada"].sum()) if not resumo.empty else 0.0
    total_no_prazo = float(resumo["Realizada no Prazo"].sum()) if not resumo.empty else 0.0
    total_apos = float(resumo["Realizada Após"].sum()) if not resumo.empty else 0.0
    total_pendente = float(resumo["Pendente"].sum()) if not resumo.empty else 0.0
    total_parada = (
        float(paradas_semanais["Tempo Parada Semana (min)"].sum())
        if not paradas_semanais.empty
        else 0.0
    )
    percentual_atraso = total_apos / total_programado * 100 if total_programado > 0 else 0.0

    buffer = BytesIO()
    pagina = landscape(A3)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=pagina,
        leftMargin=6 * mm,
        rightMargin=6 * mm,
        topMargin=7 * mm,
        bottomMargin=8 * mm,
        title=f"Calendário Semanal dos Setores - {inicio:%d/%m/%Y} a {fim:%d/%m/%Y}",
        author="PPCP",
    )

    estilos_base = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(
        "CalendarioTitulo",
        parent=estilos_base["Title"],
        fontName="Helvetica-Bold",
        fontSize=16,
        leading=18,
        textColor=colors.HexColor("#0f172a"),
        alignment=TA_LEFT,
        spaceAfter=2 * mm,
    )
    estilo_subtitulo = ParagraphStyle(
        "CalendarioSubtitulo",
        parent=estilos_base["Normal"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=10,
        textColor=colors.HexColor("#475569"),
        spaceAfter=2 * mm,
    )
    estilo_cabecalho = ParagraphStyle(
        "CalendarioCabecalho",
        parent=estilos_base["Normal"],
        fontName="Helvetica-Bold",
        fontSize=7.8,
        leading=9,
        textColor=colors.HexColor("#0f172a"),
        alignment=TA_CENTER,
    )
    estilo_setor = ParagraphStyle(
        "CalendarioSetor",
        parent=estilos_base["Normal"],
        fontName="Helvetica-Bold",
        fontSize=7.6,
        leading=9,
        textColor=colors.HexColor("#0f172a"),
    )
    estilo_celula = ParagraphStyle(
        "CalendarioCelula",
        parent=estilos_base["Normal"],
        fontName="Helvetica",
        fontSize=6.6,
        leading=8.1,
        textColor=colors.HexColor("#172033"),
        wordWrap="CJK",
    )
    estilo_metrica_rotulo = ParagraphStyle(
        "MetricaRotulo",
        parent=estilos_base["Normal"],
        fontName="Helvetica-Bold",
        fontSize=6.8,
        leading=8,
        textColor=colors.HexColor("#475569"),
        alignment=TA_CENTER,
    )
    estilo_metrica_valor = ParagraphStyle(
        "MetricaValor",
        parent=estilos_base["Normal"],
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=11,
        textColor=colors.HexColor("#0f172a"),
        alignment=TA_CENTER,
    )
    estilo_legenda = ParagraphStyle(
        "CalendarioLegenda",
        parent=estilos_base["Normal"],
        fontName="Helvetica",
        fontSize=7.2,
        leading=8.5,
        textColor=colors.HexColor("#475569"),
        spaceAfter=2 * mm,
    )

    story: list[Any] = []
    story.append(Paragraph("Calendário Semanal dos Setores", estilo_titulo))
    story.append(
        Paragraph(
            f"Período filtrado: <b>{inicio:%d/%m/%Y}</b> a <b>{fim:%d/%m/%Y}</b> | "
            f"Setores exibidos: <b>{len(setores)}</b>",
            estilo_subtitulo,
        )
    )

    metricas = [
        ("Programada na semana", fmt_qtd(total_programado)),
        ("Realizada no prazo", fmt_qtd(total_no_prazo)),
        ("Realizada após", fmt_qtd(total_apos)),
        ("Pendente", fmt_qtd(total_pendente)),
        ("% atraso semanal", f"{fmt_num(percentual_atraso, 1)}%"),
        ("Tempo parado", formatar_tempo_minutos(total_parada)),
    ]
    metricas_tabela = [
        [Paragraph(rotulo, estilo_metrica_rotulo) for rotulo, _ in metricas],
        [Paragraph(valor, estilo_metrica_valor) for _, valor in metricas],
    ]
    largura_util = pagina[0] - doc.leftMargin - doc.rightMargin
    tabela_metricas = Table(metricas_tabela, colWidths=[largura_util / 6.0] * 6)
    tabela_metricas.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#e2e8f0")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    story.append(tabela_metricas)
    story.append(Spacer(1, 2.5 * mm))
    story.append(
        Paragraph(
            "Legenda: verde = concluído sem atraso; laranja = possui pendência; "
            "vermelho = possui realização após a Data Programada; roxo = possui parada registrada.",
            estilo_legenda,
        )
    )

    def cor_status(classe: str) -> tuple[colors.Color, colors.Color]:
        mapa = {
            "cal-ok": (colors.HexColor("#eaf8ef"), colors.HexColor("#15803d")),
            "cal-pendente": (colors.HexColor("#fff3e4"), colors.HexColor("#c96a00")),
            "cal-atraso": (colors.HexColor("#fdebec"), colors.HexColor("#d11f2f")),
            "cal-com-parada": (colors.HexColor("#f5f1ff"), colors.HexColor("#6d28d9")),
            "cal-sem-programacao": (colors.HexColor("#f1f4f8"), colors.HexColor("#8293aa")),
        }
        return mapa.get(classe, (colors.white, colors.HexColor("#8293aa")))

    def texto_celula(
        programada: float,
        no_prazo: float,
        apos: float,
        pendente: float,
        percentual: float,
        lotes: str,
        paradas: list[dict[str, Any]] | None,
        tempo_parada: float,
        quantidade_paradas: int,
        semanal: bool = False,
    ) -> str:
        partes: list[str] = []
        if programada <= 0 and tempo_parada <= 0 and not paradas:
            return "-"
        if programada > 0:
            rotulo_lotes = "Lotes da semana" if semanal else "Lotes"
            rotulo_atraso = "% atraso semanal" if semanal else "% atraso"
            partes.extend(
                [
                    f"<b>{rotulo_lotes}:</b> {escape(texto_limpo(lotes) or '-')}",
                    f"<b>Programada:</b> {escape(fmt_qtd(programada))}",
                    f"<b>No prazo:</b> {escape(fmt_qtd(no_prazo))}",
                    f"<b>Após:</b> {escape(fmt_qtd(apos))}",
                    f"<b>Pendente:</b> {escape(fmt_qtd(pendente))}",
                    f"<b>{rotulo_atraso}:</b> {escape(fmt_num(percentual, 1))}%",
                ]
            )
        else:
            partes.append("<i>Sem programação de produção</i>")

        if tempo_parada > 0 or paradas:
            titulo_parada = "Paradas na semana" if semanal else "Paradas"
            partes.append(
                f'<font color="#4c1d95"><b>{titulo_parada}:</b> '
                f"{escape(formatar_tempo_minutos(tempo_parada))}</font>"
            )
            if semanal:
                sufixo = "registro" if quantidade_paradas == 1 else "registros"
                partes.append(f"{quantidade_paradas} {sufixo}")
            else:
                for item in paradas or []:
                    maquina = escape(texto_limpo(item.get("maquina")) or "Sem máquina")
                    motivo = escape(texto_limpo(item.get("motivo")) or "Sem motivo")
                    tempo_item = escape(formatar_tempo_minutos(item.get("tempo_min", 0)))
                    partes.append(f"<b>{maquina}</b> - {motivo} ({tempo_item})")
        return "<br/>".join(partes)

    cabecalho = [Paragraph("Setor", estilo_cabecalho)]
    for indice, dia in enumerate(datas_semana):
        cabecalho.append(
            Paragraph(
                f"{escape(DIAS_SEMANA_PT[indice])}<br/>{dia:%d/%m/%Y}",
                estilo_cabecalho,
            )
        )
    cabecalho.append(Paragraph("Resumo semanal", estilo_cabecalho))

    dados_tabela: list[list[Any]] = [cabecalho]
    estilos_tabela: list[tuple[Any, ...]] = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e8edf4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
        ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]

    for indice_linha, setor in enumerate(setores, start=1):
        linha_pdf: list[Any] = [Paragraph(escape(setor), estilo_setor)]
        estilos_tabela.append(
            ("BACKGROUND", (0, indice_linha), (0, indice_linha), colors.HexColor("#f1f5f9"))
        )

        for indice_dia, dia in enumerate(datas_semana, start=1):
            valores = lookup_diario.get(
                (setor, dia),
                {
                    "lotes": "",
                    "programada": 0.0,
                    "no_prazo": 0.0,
                    "apos": 0.0,
                    "pendente": 0.0,
                    "percentual": 0.0,
                },
            )
            parada = lookup_paradas.get(
                (setor, dia), {"tempo": 0.0, "quantidade": 0, "itens": []}
            )
            classe = classe_celula_calendario(
                valores["programada"], valores["apos"], valores["pendente"], parada["tempo"]
            )
            fundo, borda = cor_status(classe)
            linha_pdf.append(
                Paragraph(
                    texto_celula(
                        valores["programada"],
                        valores["no_prazo"],
                        valores["apos"],
                        valores["pendente"],
                        valores["percentual"],
                        valores["lotes"],
                        parada["itens"],
                        parada["tempo"],
                        parada["quantidade"],
                    ),
                    estilo_celula,
                )
            )
            estilos_tabela.extend(
                [
                    ("BACKGROUND", (indice_dia, indice_linha), (indice_dia, indice_linha), fundo),
                    ("LINEBEFORE", (indice_dia, indice_linha), (indice_dia, indice_linha), 2.2, borda),
                ]
            )

        semana = lookup_resumo.get(
            setor,
            {
                "Lotes da Semana": "",
                "Programada": 0.0,
                "Realizada no Prazo": 0.0,
                "Realizada Após": 0.0,
                "Pendente": 0.0,
                "% Atraso Semanal": 0.0,
            },
        )
        parada_semana = lookup_paradas_semana.get(
            setor, {"Tempo Parada Semana (min)": 0.0, "Quantidade de Paradas": 0}
        )
        classe_semana = classe_celula_calendario(
            float(semana["Programada"]),
            float(semana["Realizada Após"]),
            float(semana["Pendente"]),
            float(parada_semana["Tempo Parada Semana (min)"]),
        )
        fundo_semana, borda_semana = cor_status(classe_semana)
        linha_pdf.append(
            Paragraph(
                texto_celula(
                    float(semana["Programada"]),
                    float(semana["Realizada no Prazo"]),
                    float(semana["Realizada Após"]),
                    float(semana["Pendente"]),
                    float(semana["% Atraso Semanal"]),
                    texto_limpo(semana.get("Lotes da Semana", "")),
                    None,
                    float(parada_semana["Tempo Parada Semana (min)"]),
                    int(parada_semana["Quantidade de Paradas"]),
                    semanal=True,
                ),
                estilo_celula,
            )
        )
        coluna_resumo = 6
        estilos_tabela.extend(
            [
                ("BACKGROUND", (coluna_resumo, indice_linha), (coluna_resumo, indice_linha), fundo_semana),
                ("LINEBEFORE", (coluna_resumo, indice_linha), (coluna_resumo, indice_linha), 2.2, borda_semana),
            ]
        )
        dados_tabela.append(linha_pdf)

    larguras = [55 * mm, 53 * mm, 53 * mm, 53 * mm, 53 * mm, 53 * mm, 65 * mm]
    tabela_calendario = LongTable(
        dados_tabela,
        colWidths=larguras,
        repeatRows=1,
        splitByRow=1,
        hAlign="LEFT",
    )
    tabela_calendario.setStyle(TableStyle(estilos_tabela))
    story.append(tabela_calendario)

    def rodape(canvas, documento) -> None:
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#64748b"))
        canvas.drawString(
            doc.leftMargin,
            4 * mm,
            f"Calendário {inicio:%d/%m/%Y} a {fim:%d/%m/%Y}",
        )
        canvas.drawRightString(
            pagina[0] - doc.rightMargin,
            4 * mm,
            f"Página {documento.page}",
        )
        canvas.restoreState()

    doc.build(story, onFirstPage=rodape, onLaterPages=rodape)
    return buffer.getvalue()


def renderizar_tabela_calendario_semanal(
    diario: pd.DataFrame,
    resumo: pd.DataFrame,
    paradas_diarias: pd.DataFrame,
    paradas_semanais: pd.DataFrame,
    setores: list[str],
    inicio: date,
) -> None:
    datas_semana = [inicio + timedelta(days=i) for i in range(5)]
    lookup_diario: dict[tuple[str, date], dict[str, Any]] = {}
    for _, linha in diario.iterrows():
        lookup_diario[(str(linha["Setor"]), linha["Data Programada Dia"])] = {
            "lotes": texto_limpo(linha.get("Lotes", "")),
            "programada": float(linha["Qtde Programada Num"]),
            "no_prazo": float(linha["Realizada no Prazo Num"]),
            "apos": float(linha["Realizada Após Num"]),
            "pendente": float(linha["Pendente Num"]),
            "percentual": float(linha["% Atraso"]),
        }
    lookup_paradas: dict[tuple[str, date], dict[str, Any]] = {}
    for _, linha in paradas_diarias.iterrows():
        lookup_paradas[(str(linha["Setor"]), linha["Data Parada Dia"])] = {
            "tempo": float(linha["Tempo Parada (min)"]),
            "quantidade": int(linha["Quantidade de Paradas"]),
            "itens": linha["Itens"],
        }
    lookup_resumo = resumo.set_index("Setor").to_dict(orient="index") if not resumo.empty else {}
    lookup_paradas_semana = (
        paradas_semanais.set_index("Setor").to_dict(orient="index")
        if not paradas_semanais.empty
        else {}
    )

    cabecalhos = ['<th class="cal-setor-col">Setor</th>']
    for indice, dia in enumerate(datas_semana):
        cabecalhos.append(
            '<th>'
            f'<div class="cal-dia">{DIAS_SEMANA_PT[indice]}</div>'
            f'<div class="cal-data">{dia:%d/%m/%Y}</div>'
            '</th>'
        )
    cabecalhos.append('<th class="cal-resumo-col">Resumo semanal</th>')

    linhas_html: list[str] = []
    for setor in setores:
        celulas = [f'<td class="cal-setor-col"><b>{escape(setor)}</b></td>']
        for dia in datas_semana:
            valores = lookup_diario.get(
                (setor, dia),
                {"lotes": "", "programada": 0.0, "no_prazo": 0.0, "apos": 0.0, "pendente": 0.0, "percentual": 0.0},
            )
            parada = lookup_paradas.get(
                (setor, dia),
                {"tempo": 0.0, "quantidade": 0, "itens": []},
            )
            celulas.append(
                '<td>'
                + html_metricas_calendario(
                    valores["programada"],
                    valores["no_prazo"],
                    valores["apos"],
                    valores["pendente"],
                    valores["percentual"],
                    lotes=valores["lotes"],
                    paradas=parada["itens"],
                    tempo_parada_min=parada["tempo"],
                    quantidade_paradas=parada["quantidade"],
                )
                + '</td>'
            )

        semana = lookup_resumo.get(
            setor,
            {"Lotes da Semana": "", "Programada": 0.0, "Realizada no Prazo": 0.0, "Realizada Após": 0.0, "Pendente": 0.0, "% Atraso Semanal": 0.0},
        )
        parada_semana = lookup_paradas_semana.get(
            setor,
            {"Tempo Parada Semana (min)": 0.0, "Quantidade de Paradas": 0},
        )
        celulas.append(
            '<td class="cal-resumo-col">'
            + html_metricas_calendario(
                float(semana["Programada"]),
                float(semana["Realizada no Prazo"]),
                float(semana["Realizada Após"]),
                float(semana["Pendente"]),
                float(semana["% Atraso Semanal"]),
                lotes=texto_limpo(semana.get("Lotes da Semana", "")),
                semanal=True,
                tempo_parada_min=float(parada_semana["Tempo Parada Semana (min)"]),
                quantidade_paradas=int(parada_semana["Quantidade de Paradas"]),
            )
            + '</td>'
        )
        linhas_html.append('<tr>' + ''.join(celulas) + '</tr>')

    tabela_html = f"""
    <style>
        .cal-wrapper {{ overflow-x: auto; width: 100%; border: 1px solid #cbd5e1; border-radius: 10px; margin-top: .5rem; background: #fff; color-scheme: light; }}
        .cal-table {{ width: max-content; min-width: 100%; border-collapse: separate; border-spacing: 0; font-size: .88rem; color: #111827 !important; background: #fff !important; }}
        .cal-table th, .cal-table td {{ min-width: 170px; padding: 9px; border-right: 1px solid #d7dee8; border-bottom: 1px solid #d7dee8; vertical-align: top; background: #fff !important; color: #111827 !important; }}
        .cal-table th {{ position: sticky; top: 0; z-index: 3; text-align: center; background: #e8edf4 !important; color: #0f172a !important; font-weight: 750; }}
        .cal-table .cal-setor-col {{ min-width: 190px; max-width: 260px; position: sticky; left: 0; z-index: 2; background: #f1f5f9 !important; color: #0f172a !important; font-weight: 750; }}
        .cal-table th.cal-setor-col {{ z-index: 4; }}
        .cal-table .cal-resumo-col {{ min-width: 225px; background: #eef4ff !important; color: #0f172a !important; }}
        .cal-dia {{ font-weight: 800; color: #0f172a !important; }}
        .cal-data {{ font-size: .79rem; color: #475569 !important; margin-top: 2px; }}
        .cal-celula {{ border-left: 6px solid transparent; border-radius: 8px; padding: 8px 9px; line-height: 1.35; min-height: 164px; color: #172033 !important; box-shadow: inset 0 0 0 1px rgba(15,23,42,.04); }}
        .cal-celula *, .cal-celula span, .cal-celula strong, .cal-celula b {{ color: #172033 !important; opacity: 1 !important; text-shadow: none !important; }}
        .cal-lotes {{ display: block; margin-bottom: 5px; padding-bottom: 6px; border-bottom: 1px solid rgba(15,23,42,.18); }}
        .cal-lotes span {{ display: block; margin-bottom: 2px; font-size: .73rem; font-weight: 800; letter-spacing: .02em; text-transform: uppercase; color: #475569 !important; }}
        .cal-lotes strong {{ display: block; white-space: normal; overflow-wrap: anywhere; word-break: break-word; line-height: 1.25; font-size: .84rem; font-weight: 850; }}
        .cal-linha {{ display: flex; align-items: baseline; justify-content: space-between; gap: 10px; padding: 2px 0; white-space: nowrap; }}
        .cal-linha span {{ font-weight: 650; }} .cal-linha strong {{ font-weight: 800; text-align: right; }}
        .cal-ok {{ border-left-color: #15803d; background: #eaf8ef !important; }}
        .cal-pendente {{ border-left-color: #c96a00; background: #fff3e4 !important; }}
        .cal-atraso {{ border-left-color: #d11f2f; background: #fdebec !important; }}
        .cal-com-parada {{ border-left-color: #6d28d9; background: #f5f1ff !important; }}
        .cal-sem-programacao {{ border-left-color: #8293aa; background: #f1f4f8 !important; display: flex; align-items: center; justify-content: center; }}
        .cal-vazio {{ font-size: 1.15rem; color: #64748b !important; }}
        .cal-percentual {{ margin-top: 6px; padding-top: 6px; border-top: 1px solid rgba(15,23,42,.18); }}
        .cal-percentual span, .cal-percentual strong {{ font-weight: 850; }}
        .cal-sem-producao-texto {{ font-size: .78rem; font-weight: 750; color: #64748b !important; margin-bottom: 7px; }}
        .cal-paradas {{ margin-top: 8px; padding: 7px; border: 1px solid #c4b5fd; border-radius: 7px; background: #ede9fe !important; }}
        .cal-paradas-titulo {{ display: flex; justify-content: space-between; gap: 8px; padding-bottom: 5px; margin-bottom: 4px; border-bottom: 1px solid #c4b5fd; }}
        .cal-paradas-titulo span, .cal-paradas-titulo strong {{ color: #4c1d95 !important; font-weight: 850; }}
        .cal-parada-item {{ display: flex; justify-content: space-between; gap: 8px; padding: 4px 0; border-bottom: 1px dashed #c4b5fd; }}
        .cal-parada-item:last-child {{ border-bottom: 0; }}
        .cal-parada-item div {{ min-width: 0; }}
        .cal-parada-item strong, .cal-parada-item span {{ display: block; white-space: normal; overflow-wrap: anywhere; }}
        .cal-parada-item strong {{ font-size: .77rem; }} .cal-parada-item span {{ font-size: .75rem; color: #475569 !important; }}
        .cal-parada-item b {{ white-space: nowrap; font-size: .78rem; color: #4c1d95 !important; }}
        .cal-parada-resumo {{ font-size: .78rem; font-weight: 700; color: #4c1d95 !important; }}
        @media (max-width: 900px) {{ .cal-table {{ font-size: .84rem; }} .cal-table th, .cal-table td {{ min-width: 165px; }} .cal-table .cal-setor-col {{ min-width: 170px; }} }}
    </style>
    <div class="cal-wrapper"><table class="cal-table"><thead><tr>{''.join(cabecalhos)}</tr></thead><tbody>{''.join(linhas_html)}</tbody></table></div>
    """
    st.markdown(tabela_html, unsafe_allow_html=True)


def renderizar_calendario_semanal(df: pd.DataFrame, caminho: Path) -> None:
    st.title("Calendário Semanal dos Setores")
    st.caption(
        "Tela pública por Data Programada. Além da produção, cada dia mostra as paradas "
        "registradas na aba 6_Paradas, com máquina, motivo e duração."
    )

    base = preparar_base_calendario_semanal(df)
    try:
        paradas = obter_paradas_cacheadas(caminho)
    except Exception as exc:
        paradas = dataframe_paradas_vazio()
        st.warning(f"Não foi possível ler a aba de paradas: {exc}")

    limites = limites_datas_calendario(base, paradas)
    if limites is None:
        st.warning("Não existem Datas Programadas ou Datas de Parada válidas para montar o calendário.")
        return
    data_min, data_max = limites
    dia_anterior = date.today() - timedelta(days=1)
    data_padrao = (
        dia_anterior
        if data_min <= dia_anterior <= data_max
        else data_max
    )

    setores_producao = opcoes_unicas(base, "Setor")
    setores_paradas = opcoes_unicas(paradas, "Setor") if not paradas.empty else []
    setores_disponiveis = sorted(set(setores_producao) | set(setores_paradas), key=chave_ordenacao_texto)

    col_data, col_setores, col_movimento = st.columns([1, 2.2, 1.35])
    referencia = col_data.date_input(
        "Escolha uma data da semana",
        value=data_padrao,
        format="DD/MM/YYYY",
        key="calendario_semana_referencia",
        help="O sistema exibirá de segunda a sexta-feira da semana escolhida.",
    )
    setores_sel = col_setores.multiselect(
        "Setores",
        setores_disponiveis,
        placeholder="Vazio = todos os setores",
        key="calendario_setores",
    )
    somente_movimento = col_movimento.checkbox(
        "Somente setores com programação ou parada",
        value=True,
        key="calendario_somente_movimento",
    )

    inicio = inicio_da_semana(referencia)
    fim = inicio + timedelta(days=4)
    st.info(f"Semana útil exibida: **{inicio:%d/%m/%Y} a {fim:%d/%m/%Y}** (segunda a sexta-feira).")
    setores_calculo = setores_sel if setores_sel else setores_disponiveis
    paradas_diarias, paradas_semanais = montar_paradas_calendario(
        paradas, setores_calculo, inicio, fim
    )
    setores_com_paradas = set(paradas_diarias["Setor"].tolist()) if not paradas_diarias.empty else set()
    diario, resumo, setores_exibicao = montar_dados_calendario_semanal(
        base=base,
        setores_sel=setores_calculo,
        inicio=inicio,
        fim=fim,
        somente_movimento=somente_movimento,
        setores_com_paradas=setores_com_paradas,
    )

    total_programado = float(resumo["Programada"].sum()) if not resumo.empty else 0.0
    total_no_prazo = float(resumo["Realizada no Prazo"].sum()) if not resumo.empty else 0.0
    total_apos = float(resumo["Realizada Após"].sum()) if not resumo.empty else 0.0
    total_pendente = float(resumo["Pendente"].sum()) if not resumo.empty else 0.0
    total_parada = float(paradas_semanais["Tempo Parada Semana (min)"].sum()) if not paradas_semanais.empty else 0.0
    percentual_atraso = total_apos / total_programado * 100 if total_programado > 0 else 0.0

    k1, k2, k3, k4, k5, k6 = st.columns(6)
    k1.metric("Programada na semana", fmt_qtd(total_programado))
    k2.metric("Realizada no prazo", fmt_qtd(total_no_prazo))
    k3.metric("Realizada após", fmt_qtd(total_apos))
    k4.metric("Pendente", fmt_qtd(total_pendente))
    k5.metric("% atraso semanal", f"{fmt_num(percentual_atraso, 1)}%")
    k6.metric("Tempo parado", formatar_tempo_minutos(total_parada))

    st.caption(
        "Atraso semanal = quantidade finalizada após a Data Programada ÷ quantidade programada "
        "de segunda a sexta-feira. O tempo parado é a soma dos lançamentos da aba 6_Paradas."
    )
    if not setores_exibicao:
        st.warning("Nenhum setor possui programação ou parada na semana e nos filtros selecionados.")
        return

    aba_calendario, aba_resumo, aba_detalhes = st.tabs(
        ["Calendário semanal", "Resumo por setor", "Detalhamento"]
    )
    with aba_calendario:
        st.markdown(
            "**Legenda:** 🟢 concluído sem atraso &nbsp;&nbsp; "
            "🟠 possui pendência &nbsp;&nbsp; 🔴 possui realização após o prazo &nbsp;&nbsp; "
            "🟣 possui parada registrada",
            unsafe_allow_html=True,
        )
        pdf_calendario = gerar_pdf_calendario_semanal(
            diario, resumo, paradas_diarias, paradas_semanais, setores_exibicao, inicio, fim
        )
        st.download_button(
            "Baixar calendário em PDF",
            data=pdf_calendario,
            file_name=f"calendario_semanal_setores_{inicio:%Y%m%d}_{fim:%Y%m%d}.pdf",
            mime="application/pdf",
            use_container_width=False,
        )
        renderizar_tabela_calendario_semanal(
            diario, resumo, paradas_diarias, paradas_semanais, setores_exibicao, inicio
        )

    with aba_resumo:
        resumo_exibicao = resumo.copy()
        resumo_exibicao = resumo_exibicao.merge(paradas_semanais, on="Setor", how="left")
        resumo_exibicao["Tempo Parada Semana (min)"] = resumo_exibicao["Tempo Parada Semana (min)"].fillna(0)
        resumo_exibicao["Quantidade de Paradas"] = resumo_exibicao["Quantidade de Paradas"].fillna(0).astype(int)
        resumo_exibicao["Tempo de Parada"] = resumo_exibicao["Tempo Parada Semana (min)"].apply(formatar_tempo_minutos)
        resumo_exibicao["Situação"] = "Concluído no prazo"
        resumo_exibicao.loc[resumo_exibicao["Pendente"] > 0, "Situação"] = "Com pendência"
        resumo_exibicao.loc[resumo_exibicao["Realizada Após"] > 0, "Situação"] = "Com atraso"
        resumo_exibicao.loc[resumo_exibicao["Programada"] <= 0, "Situação"] = "Sem programação"
        st.dataframe(
            resumo_exibicao,
            use_container_width=True,
            hide_index=True,
            height=520,
            column_config={
                "Programada": st.column_config.NumberColumn("Programada", format="%.2f"),
                "Realizada no Prazo": st.column_config.NumberColumn("Realizada no Prazo", format="%.2f"),
                "Realizada Após": st.column_config.NumberColumn("Realizada Após", format="%.2f"),
                "Pendente": st.column_config.NumberColumn("Pendente", format="%.2f"),
                "% Atraso Semanal": st.column_config.ProgressColumn("% Atraso Semanal", min_value=0.0, max_value=100.0, format="%.1f%%"),
                "Tempo Parada Semana (min)": st.column_config.NumberColumn("Parada (min)", format="%.0f"),
            },
        )
        csv_resumo = resumo_dataframe_para_csv_excel(exibicao)
        st.download_button(
            "Baixar resumo semanal em CSV",
            data=csv_resumo,
            file_name=f"calendario_semanal_setores_{inicio:%Y%m%d}_{fim:%Y%m%d}.csv",
            mime="text/csv",
        )

    with aba_detalhes:
        sub_producao, sub_paradas = st.tabs(["Produção", "Paradas"])
        with sub_producao:
            detalhes = diario.copy().rename(
                columns={
                    "Data Programada Dia": "Data Programada",
                    "Qtde Programada Num": "Programada",
                    "Realizada no Prazo Num": "Realizada no Prazo",
                    "Realizada Após Num": "Realizada Após",
                    "Pendente Num": "Pendente",
                }
            )
            if detalhes.empty:
                st.info("Não existem linhas de produção para detalhar nesta semana.")
            else:
                detalhes["Data Programada"] = pd.to_datetime(detalhes["Data Programada"]).dt.strftime("%d/%m/%Y")
                detalhes = detalhes[["Data Programada", "Setor", "Lotes", "Programada", "Realizada no Prazo", "Realizada Após", "Pendente", "% Atraso"]]
                st.dataframe(detalhes, use_container_width=True, hide_index=True, height=520)

        with sub_paradas:
            if paradas.empty:
                st.info("Ainda não existem paradas registradas na aba 6_Paradas.")
            else:
                detalhes_paradas = paradas.copy()
                detalhes_paradas["Data Parada DT"] = pd.to_datetime(detalhes_paradas["Data Parada Valor"], errors="coerce")
                detalhes_paradas = detalhes_paradas[
                    detalhes_paradas["Setor"].isin(setores_calculo)
                    & entre_datas(detalhes_paradas["Data Parada DT"], inicio, fim)
                ].copy()
                if detalhes_paradas.empty:
                    st.info("Não existem paradas registradas para esta semana e filtros.")
                else:
                    detalhes_paradas["Tempo de Parada"] = detalhes_paradas["Tempo Total Parada (min)"].apply(formatar_tempo_minutos)
                    detalhes_paradas = detalhes_paradas.sort_values(["Data Parada DT", "Setor", "Máquina/Posto"])
                    st.dataframe(
                        detalhes_paradas[["Data Parada", "Setor", "Máquina/Posto", "Motivo da Parada", "Tempo Total Parada (min)", "Tempo de Parada", "Observação"]],
                        use_container_width=True,
                        hide_index=True,
                        height=520,
                    )

    st.divider()
    st.caption(f"Fonte dos dados: {caminho.name} — abas {NOME_ABA_PADRAO} e {NOME_ABA_PARADAS}")



# -----------------------------------------------------------------------------
# Tela pública: Consulta de Paradas
# -----------------------------------------------------------------------------
def renderizar_consulta_paradas(caminho: Path) -> None:
    """Exibe os apontamentos da aba de paradas sem permitir alterações."""
    st.title("Paradas por Setor, Equipamento e Data")
    st.caption(
        f"Consulta pública e somente leitura dos registros da aba {NOME_ABA_PARADAS}. "
        "Use os filtros para analisar o histórico de paradas sem alterar a planilha."
    )

    try:
        with st.spinner("Carregando os registros de parada..."):
            paradas = obter_paradas_cacheadas(caminho)
    except Exception as exc:
        st.error(f"Não foi possível ler a aba de paradas: {exc}")
        return

    if paradas.empty:
        st.info(f"Ainda não existem registros válidos na aba {NOME_ABA_PARADAS}.")
        return

    base = paradas.copy()
    base["Data Parada DT"] = pd.to_datetime(base["Data Parada Valor"], errors="coerce")
    base["Tempo Total Parada (min)"] = to_num(base["Tempo Total Parada (min)"].copy()).clip(lower=0)

    datas_validas = base["Data Parada DT"].dropna()
    if datas_validas.empty:
        st.warning("Existem registros na aba de paradas, mas nenhuma Data Parada pôde ser reconhecida.")
        return

    data_min = datas_validas.min().date()
    data_max = datas_validas.max().date()

    st.subheader("Filtros")
    col_periodo, col_setor, col_maquina = st.columns([1.05, 1.25, 1.45])

    modo_data = col_periodo.radio(
        "Período",
        ["Dia específico", "Período personalizado", "Período completo"],
        index=1,
        key="consulta_paradas_modo_data",
    )

    if modo_data == "Dia específico":
        hoje = date.today()
        data_padrao = hoje if data_min <= hoje <= data_max else data_max
        inicio = fim = col_periodo.date_input(
            "Data analisada",
            value=data_padrao,
            min_value=data_min,
            max_value=data_max,
            format="DD/MM/YYYY",
            key="consulta_paradas_dia",
        )
    elif modo_data == "Período personalizado":
        periodo = col_periodo.date_input(
            "Datas analisadas",
            value=(data_min, data_max),
            min_value=data_min,
            max_value=data_max,
            format="DD/MM/YYYY",
            key="consulta_paradas_periodo",
        )
        if isinstance(periodo, (tuple, list)) and len(periodo) == 2:
            inicio, fim = periodo
        elif isinstance(periodo, (tuple, list)) and len(periodo) == 1:
            inicio = fim = periodo[0]
        else:
            inicio = fim = periodo
    else:
        inicio, fim = data_min, data_max
        col_periodo.caption(f"Base completa: {inicio:%d/%m/%Y} a {fim:%d/%m/%Y}")

    setores = opcoes_unicas(base, "Setor")
    setores_sel = col_setor.multiselect(
        "Setor",
        setores,
        placeholder="Vazio = todos os setores",
        key="consulta_paradas_setores",
    )

    base_opcoes_maquina = base.copy()
    if setores_sel:
        base_opcoes_maquina = base_opcoes_maquina[base_opcoes_maquina["Setor"].isin(setores_sel)]
    maquinas = opcoes_unicas(base_opcoes_maquina, "Máquina/Posto")
    maquinas_sel = col_maquina.multiselect(
        "Equipamento/Máquina",
        maquinas,
        placeholder="Vazio = todos os equipamentos",
        key="consulta_paradas_maquinas",
    )

    col_motivo, col_obs = st.columns(2)
    busca_motivo = col_motivo.text_input(
        "Buscar no motivo",
        placeholder="Ex.: manutenção, material, energia...",
        key="consulta_paradas_busca_motivo",
    )
    busca_observacao = col_obs.text_input(
        "Buscar na observação",
        placeholder="Digite uma palavra ou parte do texto",
        key="consulta_paradas_busca_observacao",
    )

    filtrado = base[entre_datas(base["Data Parada DT"], inicio, fim)].copy()
    if setores_sel:
        filtrado = filtrado[filtrado["Setor"].isin(setores_sel)]
    if maquinas_sel:
        filtrado = filtrado[filtrado["Máquina/Posto"].isin(maquinas_sel)]
    if busca_motivo:
        filtrado = filtrado[
            filtrado["Motivo da Parada"].astype(str).str.contains(busca_motivo, case=False, na=False)
        ]
    if busca_observacao:
        filtrado = filtrado[
            filtrado["Observação"].astype(str).str.contains(busca_observacao, case=False, na=False)
        ]

    if filtrado.empty:
        st.warning("Nenhuma parada foi encontrada para os filtros selecionados.")
        return

    total_min = float(filtrado["Tempo Total Parada (min)"].sum())
    quantidade = int(len(filtrado))
    setores_movimento = int(filtrado["Setor"].nunique())
    maquinas_movimento = int(filtrado["Máquina/Posto"].nunique())
    media_min = total_min / quantidade if quantidade else 0.0

    st.divider()
    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("Paradas", fmt_num(quantidade))
    k2.metric("Tempo total parado", formatar_tempo_minutos(total_min))
    k3.metric("Tempo médio/parada", formatar_tempo_minutos(media_min))
    k4.metric("Setores", fmt_num(setores_movimento))
    k5.metric("Equipamentos", fmt_num(maquinas_movimento))
    st.caption(f"Resultado de {inicio:%d/%m/%Y} a {fim:%d/%m/%Y}.")

    aba_registros, aba_setor, aba_equipamento, aba_data = st.tabs(
        ["Registros", "Resumo por setor", "Resumo por equipamento", "Resumo por data"]
    )

    with aba_registros:
        st.subheader("Registros encontrados")
        exibicao = filtrado.copy()
        exibicao["Tempo de Parada"] = exibicao["Tempo Total Parada (min)"].apply(formatar_tempo_minutos)
        exibicao["Registrado Em Exibição"] = pd.to_datetime(
            exibicao["Registrado Em"], errors="coerce"
        ).dt.strftime("%d/%m/%Y %H:%M")
        exibicao["Registrado Em Exibição"] = exibicao["Registrado Em Exibição"].fillna(
            exibicao["Registrado Em"].fillna("").astype(str)
        )
        exibicao = exibicao.sort_values(
            ["Data Parada DT", "Setor", "Máquina/Posto", "Linha Excel"],
            ascending=[False, True, True, False],
        )
        tabela = exibicao[
            [
                "Data Parada",
                "Setor",
                "Máquina/Posto",
                "Motivo da Parada",
                "Tempo Total Parada (min)",
                "Tempo de Parada",
                "Observação",
                "Registrado Em Exibição",
            ]
        ].rename(columns={"Registrado Em Exibição": "Registrado Em"})
        st.dataframe(
            tabela,
            use_container_width=True,
            hide_index=True,
            height=540,
            column_config={
                "Data Parada": st.column_config.TextColumn("Data", width="small"),
                "Setor": st.column_config.TextColumn("Setor", width="medium"),
                "Máquina/Posto": st.column_config.TextColumn("Equipamento/Máquina", width="large"),
                "Motivo da Parada": st.column_config.TextColumn("Motivo", width="large"),
                "Tempo Total Parada (min)": st.column_config.NumberColumn("Tempo (min)", format="%.0f"),
                "Tempo de Parada": st.column_config.TextColumn("Tempo", width="small"),
                "Observação": st.column_config.TextColumn("Observação", width="large"),
                "Registrado Em": st.column_config.TextColumn("Registrado em", width="medium"),
            },
        )
        csv = dataframe_para_csv_excel(tabela)
        st.download_button(
            "Baixar registros filtrados em CSV",
            data=csv,
            file_name=f"paradas_{inicio:%Y%m%d}_{fim:%Y%m%d}.csv",
            mime="text/csv",
            key="consulta_paradas_csv_registros",
        )

    with aba_setor:
        resumo_setor = (
            filtrado.groupby("Setor", as_index=False)
            .agg(
                **{
                    "Quantidade de Paradas": ("Linha Excel", "count"),
                    "Tempo Total (min)": ("Tempo Total Parada (min)", "sum"),
                    "Equipamentos": ("Máquina/Posto", "nunique"),
                }
            )
            .sort_values(["Tempo Total (min)", "Quantidade de Paradas"], ascending=[False, False])
        )
        resumo_setor["Tempo Total"] = resumo_setor["Tempo Total (min)"].apply(formatar_tempo_minutos)
        st.bar_chart(resumo_setor.set_index("Setor")["Tempo Total (min)"])
        st.dataframe(
            resumo_setor,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Quantidade de Paradas": st.column_config.NumberColumn("Paradas", format="%d"),
                "Tempo Total (min)": st.column_config.NumberColumn("Tempo (min)", format="%.0f"),
                "Equipamentos": st.column_config.NumberColumn("Equipamentos", format="%d"),
            },
        )

    with aba_equipamento:
        resumo_maquina = (
            filtrado.groupby(["Setor", "Máquina/Posto"], as_index=False)
            .agg(
                **{
                    "Quantidade de Paradas": ("Linha Excel", "count"),
                    "Tempo Total (min)": ("Tempo Total Parada (min)", "sum"),
                }
            )
            .sort_values(["Tempo Total (min)", "Quantidade de Paradas"], ascending=[False, False])
        )
        resumo_maquina["Tempo Total"] = resumo_maquina["Tempo Total (min)"].apply(formatar_tempo_minutos)
        grafico_maquina = resumo_maquina.copy()
        grafico_maquina["Equipamento"] = (
            grafico_maquina["Setor"].astype(str) + " — " + grafico_maquina["Máquina/Posto"].astype(str)
        )
        st.bar_chart(grafico_maquina.set_index("Equipamento")["Tempo Total (min)"])
        st.dataframe(
            resumo_maquina,
            use_container_width=True,
            hide_index=True,
            height=520,
            column_config={
                "Máquina/Posto": st.column_config.TextColumn("Equipamento/Máquina", width="large"),
                "Quantidade de Paradas": st.column_config.NumberColumn("Paradas", format="%d"),
                "Tempo Total (min)": st.column_config.NumberColumn("Tempo (min)", format="%.0f"),
            },
        )

    with aba_data:
        resumo_data = filtrado.copy()
        resumo_data["Data"] = resumo_data["Data Parada DT"].dt.date
        resumo_data = (
            resumo_data.groupby("Data", as_index=False)
            .agg(
                **{
                    "Quantidade de Paradas": ("Linha Excel", "count"),
                    "Tempo Total (min)": ("Tempo Total Parada (min)", "sum"),
                    "Setores": ("Setor", "nunique"),
                    "Equipamentos": ("Máquina/Posto", "nunique"),
                }
            )
            .sort_values("Data")
        )
        resumo_data["Tempo Total"] = resumo_data["Tempo Total (min)"].apply(formatar_tempo_minutos)
        grafico_data = resumo_data.copy()
        grafico_data["Data"] = pd.to_datetime(grafico_data["Data"])
        st.line_chart(grafico_data.set_index("Data")["Tempo Total (min)"])
        tabela_data = resumo_data.copy()
        tabela_data["Data"] = pd.to_datetime(tabela_data["Data"]).dt.strftime("%d/%m/%Y")
        st.dataframe(
            tabela_data,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Quantidade de Paradas": st.column_config.NumberColumn("Paradas", format="%d"),
                "Tempo Total (min)": st.column_config.NumberColumn("Tempo (min)", format="%.0f"),
                "Setores": st.column_config.NumberColumn("Setores", format="%d"),
                "Equipamentos": st.column_config.NumberColumn("Equipamentos", format="%d"),
            },
        )

    st.divider()
    st.caption(f"Fonte dos dados: {caminho.name} — aba {NOME_ABA_PARADAS}. Tela somente leitura.")


# -----------------------------------------------------------------------------
# Tela protegida: Apontamento de Paradas
# -----------------------------------------------------------------------------
def renderizar_apontamento_paradas(
    df: pd.DataFrame,
    caminho: Path,
    nome_aba: str,
) -> None:
    st.title("Apontamento de Paradas")
    st.caption(
        f"Registre as paradas por setor e máquina. Os dados são gravados na aba {NOME_ABA_PARADAS} "
        f"do arquivo {caminho.name}."
    )

    mensagem_edicao = st.session_state.pop("mensagem_edicao_parada", None)
    if mensagem_edicao:
        st.success(mensagem_edicao)

    setores = opcoes_unicas(df, "Setor")
    if not setores:
        st.warning("Não encontrei setores na aba de acompanhamento.")
        return

    setor = st.selectbox(
        "Setor",
        ["Selecione o setor", *setores],
        key="parada_setor",
    )
    maquinas_setor: list[str] = []
    if setor != "Selecione o setor":
        maquinas_setor = opcoes_unicas(df[df["Setor"] == setor], "Máquina/Posto")

    # Primeiro aparecem as máquinas já relacionadas ao setor na planilha.
    # Em seguida, entram as máquinas do cadastro fixo solicitado pelo PPCP.
    maquinas_disponiveis = combinar_opcoes_maquinas(
        maquinas_setor,
        MAQUINAS_PARADAS_CADASTRADAS,
    )
    opcoes_maquina = [
        "Selecione a máquina",
        *maquinas_disponiveis,
        "SETOR GERAL (sem máquina específica)",
        "Outra máquina/posto",
    ]
    maquina_escolhida = st.selectbox(
        "Máquina/Posto",
        opcoes_maquina,
        key="parada_maquina",
        help=(
            "A lista combina as máquinas encontradas na planilha para o setor "
            "com o cadastro fixo de máquinas do apontamento de paradas."
        ),
    )
    st.caption(
        f"{len(maquinas_disponiveis)} máquina(s)/posto(s) disponível(is). "
        "É possível digitar no campo aberto para localizar uma opção."
    )
    maquina_digitada = ""
    if maquina_escolhida == "Outra máquina/posto":
        maquina_digitada = st.text_input(
            "Informe a máquina/posto",
            placeholder="Ex.: Coladeira 02",
            key="parada_maquina_outra",
        )
    maquina_final = maquina_digitada.strip() if maquina_escolhida == "Outra máquina/posto" else maquina_escolhida

    with st.form("form_apontamento_parada", clear_on_submit=True):
        col_data, col_horas, col_minutos = st.columns([1.3, 1, 1])
        data_parada = col_data.date_input(
            "Data da parada",
            value=date.today(),
            format="DD/MM/YYYY",
        )
        horas = col_horas.number_input(
            "Horas paradas",
            min_value=0,
            max_value=999,
            value=0,
            step=1,
        )
        minutos = col_minutos.number_input(
            "Minutos adicionais",
            min_value=0,
            max_value=59,
            value=0,
            step=1,
        )
        motivo = st.text_input(
            "Motivo da parada",
            placeholder="Ex.: manutenção corretiva, falta de material, falta de energia...",
        )
        observacao = st.text_area(
            "Observação (opcional)",
            placeholder="Informe detalhes adicionais que ajudem na análise da parada.",
            height=100,
        )
        criar_backup = False
        salvar = st.form_submit_button(
            "Salvar parada",
            type="primary",
            use_container_width=True,
        )

    if salvar:
        tempo_total_min = int(horas) * 60 + int(minutos)
        erros: list[str] = []
        if setor == "Selecione o setor":
            erros.append("selecione o setor")
        if maquina_final in {"", "Selecione a máquina", "Outra máquina/posto"}:
            erros.append("informe a máquina/posto")
        if not motivo.strip():
            erros.append("informe o motivo da parada")
        if tempo_total_min <= 0:
            erros.append("informe um tempo de parada maior que zero")

        if erros:
            st.error("Não foi possível salvar: " + "; ".join(erros) + ".")
        else:
            try:
                with st.spinner("Salvando parada na planilha..."):
                    backup = salvar_parada(
                        caminho=caminho,
                        data_parada=data_parada,
                        setor=setor,
                        maquina=maquina_final,
                        motivo=motivo,
                        tempo_total_min=tempo_total_min,
                        observacao=observacao,
                        criar_backup=criar_backup,
                    )
                carregar_paradas_cacheado.clear()
                carregar_dados_cacheado.clear()
                st.success(
                    f"Parada salva: {setor} — {maquina_final} — {formatar_tempo_minutos(tempo_total_min)}."
                )
                if backup:
                    st.info(f"Backup criado: {backup.name}")
                st.rerun()
            except PermissionError:
                st.error("Não consegui salvar. Feche a planilha no Excel e tente novamente.")
            except Exception as exc:
                st.error(f"Erro ao salvar a parada: {exc}")

    st.divider()
    try:
        paradas = obter_paradas_cacheadas(caminho)
    except Exception as exc:
        st.error(f"Não foi possível carregar os apontamentos de parada: {exc}")
        paradas = dataframe_paradas_vazio()

    paradas = garantir_colunas_dataframe_paradas(paradas)

    if paradas.empty:
        st.info(
            f"Ainda não há paradas registradas. No primeiro salvamento, os cabeçalhos serão criados na aba {NOME_ABA_PARADAS}."
        )
    else:
        total_min = float(paradas["Tempo Total Parada (min)"].sum())
        ultima_data = pd.to_datetime(paradas["Data Parada Valor"], errors="coerce").max()
        m1, m2, m3 = st.columns(3)
        m1.metric("Paradas registradas", fmt_num(len(paradas)))
        m2.metric("Tempo total registrado", formatar_tempo_minutos(total_min))
        m3.metric("Última data", ultima_data.strftime("%d/%m/%Y") if pd.notna(ultima_data) else "—")

        st.subheader("Últimos apontamentos")
        exibicao = paradas.copy()
        exibicao["Data Ordenação"] = pd.to_datetime(exibicao["Data Parada Valor"], errors="coerce")
        exibicao["Tempo de Parada"] = exibicao["Tempo Total Parada (min)"].apply(formatar_tempo_minutos)
        exibicao = exibicao.sort_values(["Data Ordenação", "Linha Excel"], ascending=[False, False]).head(300)
        colunas_exibicao_paradas = [
            "Data Parada",
            "Setor",
            "Máquina/Posto",
            "Motivo da Parada",
            "Tempo Total Parada (min)",
            "Tempo de Parada",
            "Observação",
            "Registrado Em",
            "Editado Em",
            "Editado Por",
        ]
        st.dataframe(
            exibicao[colunas_exibicao_paradas],
            use_container_width=True,
            hide_index=True,
            height=480,
        )

        st.subheader("Editar apontamento salvo")
        st.caption(
            "Selecione um registro, altere os campos necessários e salve. "
            "A edição atualiza a mesma linha da aba de paradas e registra a data da alteração."
        )

        registros_edicao = paradas.copy()
        registros_edicao["Data Ordenação"] = pd.to_datetime(
            registros_edicao["Data Parada Valor"], errors="coerce"
        )
        registros_edicao = registros_edicao.sort_values(
            ["Data Ordenação", "Linha Excel"], ascending=[False, False]
        )
        linhas_edicao = registros_edicao["Linha Excel"].astype(int).tolist()
        registros_por_linha = {
            int(linha["Linha Excel"]): linha
            for _, linha in registros_edicao.iterrows()
        }

        def rotulo_registro_parada(linha_excel: int) -> str:
            registro = registros_por_linha[int(linha_excel)]
            motivo_resumido = texto_limpo(registro.get("Motivo da Parada", ""))
            if len(motivo_resumido) > 55:
                motivo_resumido = motivo_resumido[:52] + "..."
            return (
                f"{texto_limpo(registro.get('Data Parada'))} | "
                f"{texto_limpo(registro.get('Setor'))} | "
                f"{texto_limpo(registro.get('Máquina/Posto'))} | "
                f"{motivo_resumido} | "
                f"{formatar_tempo_minutos(registro.get('Tempo Total Parada (min)', 0))}"
            )

        linha_edicao = st.selectbox(
            "Registro que será editado",
            options=linhas_edicao,
            format_func=rotulo_registro_parada,
            key="parada_registro_edicao",
        )
        registro_edicao = registros_por_linha[int(linha_edicao)]

        data_atual_edicao = pd.to_datetime(
            registro_edicao.get("Data Parada Valor"), errors="coerce"
        )
        data_padrao_edicao = (
            data_atual_edicao.date() if pd.notna(data_atual_edicao) else date.today()
        )
        tempo_atual_edicao = int(
            round(minutos_seguro(registro_edicao.get("Tempo Total Parada (min)", 0)))
        )
        horas_padrao_edicao, minutos_padrao_edicao = divmod(tempo_atual_edicao, 60)

        setor_atual_edicao = texto_limpo(registro_edicao.get("Setor"))
        setores_edicao = sorted(
            set([*setores, setor_atual_edicao]),
            key=chave_ordenacao_texto,
        )
        indice_setor_edicao = (
            setores_edicao.index(setor_atual_edicao)
            if setor_atual_edicao in setores_edicao
            else 0
        )

        with st.form(f"form_editar_parada_{int(linha_edicao)}"):
            col_data_edit, col_setor_edit = st.columns([1, 2])
            nova_data_parada = col_data_edit.date_input(
                "Data da parada",
                value=data_padrao_edicao,
                format="DD/MM/YYYY",
                key=f"parada_editar_data_{int(linha_edicao)}",
            )
            novo_setor = col_setor_edit.selectbox(
                "Setor",
                options=setores_edicao,
                index=indice_setor_edicao,
                key=f"parada_editar_setor_{int(linha_edicao)}",
            )
            nova_maquina = st.text_input(
                "Máquina/Posto",
                value=texto_limpo(registro_edicao.get("Máquina/Posto")),
                key=f"parada_editar_maquina_{int(linha_edicao)}",
            )
            col_horas_edit, col_minutos_edit = st.columns(2)
            novas_horas = col_horas_edit.number_input(
                "Horas paradas",
                min_value=0,
                max_value=999,
                value=int(horas_padrao_edicao),
                step=1,
                key=f"parada_editar_horas_{int(linha_edicao)}",
            )
            novos_minutos = col_minutos_edit.number_input(
                "Minutos adicionais",
                min_value=0,
                max_value=59,
                value=int(minutos_padrao_edicao),
                step=1,
                key=f"parada_editar_minutos_{int(linha_edicao)}",
            )
            novo_motivo = st.text_input(
                "Motivo da parada",
                value=texto_limpo(registro_edicao.get("Motivo da Parada")),
                key=f"parada_editar_motivo_{int(linha_edicao)}",
            )
            nova_observacao = st.text_area(
                "Observação (opcional)",
                value=texto_limpo(registro_edicao.get("Observação")),
                height=100,
                key=f"parada_editar_observacao_{int(linha_edicao)}",
            )
            confirmar_edicao = st.checkbox(
                "Confirmo que desejo substituir os dados desse registro",
                value=False,
                key=f"parada_confirmar_edicao_{int(linha_edicao)}",
            )
            salvar_edicao = st.form_submit_button(
                "Salvar alterações",
                type="primary",
                use_container_width=True,
            )

        if salvar_edicao:
            novo_tempo_total_min = int(novas_horas) * 60 + int(novos_minutos)
            erros_edicao: list[str] = []
            if not confirmar_edicao:
                erros_edicao.append("confirme a substituição do registro")
            if not texto_limpo(novo_setor):
                erros_edicao.append("informe o setor")
            if not nova_maquina.strip():
                erros_edicao.append("informe a máquina/posto")
            if not novo_motivo.strip():
                erros_edicao.append("informe o motivo da parada")
            if novo_tempo_total_min <= 0:
                erros_edicao.append("informe um tempo de parada maior que zero")

            if erros_edicao:
                st.error("Não foi possível editar: " + "; ".join(erros_edicao) + ".")
            else:
                try:
                    with st.spinner("Atualizando o registro na planilha..."):
                        editar_parada(
                            caminho=caminho,
                            linha_excel=int(linha_edicao),
                            data_parada=nova_data_parada,
                            setor=novo_setor,
                            maquina=nova_maquina,
                            motivo=novo_motivo,
                            tempo_total_min=novo_tempo_total_min,
                            observacao=nova_observacao,
                            criar_backup=False,
                        )
                    carregar_paradas_cacheado.clear()
                    carregar_dados_cacheado.clear()
                    st.session_state["mensagem_edicao_parada"] = (
                        f"Registro atualizado: {novo_setor} — {nova_maquina} — "
                        f"{formatar_tempo_minutos(novo_tempo_total_min)}."
                    )
                    st.rerun()
                except PermissionError:
                    st.error("Não consegui editar. Feche a planilha no Excel e tente novamente.")
                except Exception as exc:
                    st.error(f"Erro ao editar a parada: {exc}")

    with open(caminho, "rb") as arquivo:
        st.download_button(
            "Baixar planilha atualizada",
            data=arquivo.read(),
            file_name=caminho.name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="baixar_planilha_paradas",
        )
    st.caption(f"Base de produção: {nome_aba} | Base de paradas: {NOME_ABA_PARADAS}")



# -----------------------------------------------------------------------------
# Apontamento operacional por equipamento - SQLite + catálogo validado de fichas
# -----------------------------------------------------------------------------
def _candidatos_banco_operacional() -> list[Path]:
    """Lista locais compatíveis com versões anteriores sem criar arquivos."""
    candidatos: list[Path] = []
    configurado = texto_limpo(os.environ.get("PPCP_BANCO_OPERACIONAL", ""))
    if configurado:
        candidatos.append(Path(configurado).expanduser())

    pasta_script = Path(__file__).resolve().parent
    pasta_execucao = Path.cwd().resolve()
    candidatos.extend(
        [
            pasta_script / ARQUIVO_BANCO_OPERACIONAL,
            pasta_execucao / ARQUIVO_BANCO_OPERACIONAL,
            pasta_script / "dados" / ARQUIVO_BANCO_OPERACIONAL,
            pasta_execucao / "dados" / ARQUIVO_BANCO_OPERACIONAL,
        ]
    )

    unicos: list[Path] = []
    vistos: set[str] = set()
    for candidato in candidatos:
        chave = str(candidato.resolve(strict=False)).lower()
        if chave not in vistos:
            vistos.add(chave)
            unicos.append(candidato)
    return unicos


def _pontuacao_banco_operacional(caminho: Path) -> tuple[int, int, int, int]:
    """Prioriza o banco que realmente contém mais apontamentos finalizados."""
    if not caminho.exists() or not caminho.is_file():
        return (-1, -1, -1, -1)

    total_ordens = 0
    total_finalizadas = 0
    try:
        uri = caminho.resolve().as_uri() + "?mode=ro"
        conexao = sqlite3.connect(uri, uri=True, timeout=3)
        try:
            existe = conexao.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name='ordens_operacionais'"
            ).fetchone()
            if existe:
                total_ordens = int(
                    conexao.execute("SELECT COUNT(*) FROM ordens_operacionais").fetchone()[0]
                )
                total_finalizadas = int(
                    conexao.execute(
                        "SELECT COUNT(*) FROM ordens_operacionais WHERE fechada_em IS NOT NULL"
                    ).fetchone()[0]
                )
        finally:
            conexao.close()
    except sqlite3.Error:
        pass

    estatistica = caminho.stat()
    return (
        total_finalizadas,
        total_ordens,
        int(estatistica.st_size),
        int(estatistica.st_mtime_ns),
    )


def caminho_banco_operacional() -> Path:
    """
    Retorna o banco operacional existente mais completo.

    Isso evita que uma versão nova do arquivo Python, executada em outra pasta,
    consulte um SQLite vazio enquanto o banco com os apontamentos permanece no
    diretório de execução ou na pasta dados.
    """
    candidatos = _candidatos_banco_operacional()
    configurado = texto_limpo(os.environ.get("PPCP_BANCO_OPERACIONAL", ""))
    if configurado:
        return candidatos[0]

    existentes = [caminho for caminho in candidatos if caminho.exists()]
    if existentes:
        return max(existentes, key=_pontuacao_banco_operacional)

    destino = candidatos[0]
    destino.parent.mkdir(parents=True, exist_ok=True)
    return destino


def caminho_catalogo_fichas() -> Path:
    return Path(__file__).resolve().parent / ARQUIVO_CATALOGO_FICHAS


def agora_iso() -> str:
    return datetime.now().replace(microsecond=0).isoformat(sep=" ")


def hash_senha_operador(usuario: str, senha: str) -> str:
    sal = f"PPCP-OPERADOR|{usuario.strip().lower()}".encode("utf-8")
    derivada = hashlib.pbkdf2_hmac("sha256", senha.encode("utf-8"), sal, 150_000)
    return derivada.hex()


def conectar_banco_operacional() -> sqlite3.Connection:
    conexao = sqlite3.connect(caminho_banco_operacional(), timeout=15)
    conexao.row_factory = sqlite3.Row
    conexao.execute("PRAGMA foreign_keys = ON")
    conexao.execute("PRAGMA busy_timeout = 15000")
    return conexao


def conectar_catalogo_fichas() -> sqlite3.Connection:
    conexao = sqlite3.connect(caminho_catalogo_fichas(), timeout=15)
    conexao.row_factory = sqlite3.Row
    conexao.execute("PRAGMA foreign_keys = ON")
    conexao.execute("PRAGMA busy_timeout = 15000")
    return conexao


def adicionar_coluna_sqlite(conexao: sqlite3.Connection, tabela: str, coluna: str, definicao: str) -> None:
    existentes = {linha["name"] for linha in conexao.execute(f"PRAGMA table_info({tabela})").fetchall()}
    if coluna not in existentes:
        conexao.execute(f"ALTER TABLE {tabela} ADD COLUMN {coluna} {definicao}")


@st.cache_resource(show_spinner=False)
def inicializar_banco_operacional() -> bool:
    """Inicializa uma única vez por processo; evita recalcular 17 hashes a cada clique."""
    with conectar_banco_operacional() as conexao:
        conexao.execute("PRAGMA journal_mode = WAL")
        conexao.execute("PRAGMA synchronous = NORMAL")
        conexao.executescript(
            """
            CREATE TABLE IF NOT EXISTS equipamentos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                usuario TEXT NOT NULL UNIQUE,
                senha_hash TEXT NOT NULL,
                codigo TEXT NOT NULL,
                setor TEXT NOT NULL,
                maquina TEXT NOT NULL,
                ativo INTEGER NOT NULL DEFAULT 1,
                criado_em TEXT NOT NULL,
                atualizado_em TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS ordens_operacionais (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                equipamento_id INTEGER NOT NULL,
                usuario TEXT NOT NULL,
                setor TEXT NOT NULL,
                maquina TEXT NOT NULL,
                codigo_barras TEXT NOT NULL,
                codigo_normalizado TEXT NOT NULL,
                lote_identificado TEXT,
                codigo_peca_identificado TEXT,
                descricao_peca TEXT,
                quantidade_programada REAL,
                aberta_em TEXT NOT NULL,
                fechada_em TEXT,
                duracao_segundos INTEGER,
                quantidade_boa REAL NOT NULL DEFAULT 0,
                refugo REAL NOT NULL DEFAULT 0,
                retrabalho REAL NOT NULL DEFAULT 0,
                observacao_fechamento TEXT,
                status TEXT NOT NULL DEFAULT 'ABERTA',
                FOREIGN KEY (equipamento_id) REFERENCES equipamentos(id)
            );

            CREATE TABLE IF NOT EXISTS paradas_operacionais (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                equipamento_id INTEGER NOT NULL,
                usuario TEXT NOT NULL,
                setor TEXT NOT NULL,
                maquina TEXT NOT NULL,
                motivo TEXT NOT NULL,
                iniciada_em TEXT NOT NULL,
                encerrada_em TEXT,
                duracao_segundos INTEGER,
                status TEXT NOT NULL DEFAULT 'ABERTA',
                FOREIGN KEY (equipamento_id) REFERENCES equipamentos(id)
            );
            """
        )

        novas_colunas = {
            "ficha_op": "TEXT",
            "rota_ordem": "INTEGER",
            "rota_codigo_equipamento": "TEXT",
            "rota_equipamento": "TEXT",
            "rota_operacao": "TEXT",
            "tipo_material": "TEXT",
            "revestimento_ficha": "TEXT",
            "medida_ficha": "TEXT",
            "obs_lote_ficha": "TEXT",
            "ficha_snapshot_json": "TEXT",
        }
        for coluna, definicao in novas_colunas.items():
            adicionar_coluna_sqlite(conexao, "ordens_operacionais", coluna, definicao)

        conexao.executescript(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS ux_ordem_aberta_por_equipamento
                ON ordens_operacionais(equipamento_id) WHERE status = 'ABERTA';
            CREATE UNIQUE INDEX IF NOT EXISTS ux_parada_aberta_por_equipamento
                ON paradas_operacionais(equipamento_id) WHERE status = 'ABERTA';
            CREATE UNIQUE INDEX IF NOT EXISTS ux_ordem_rota_processada
                ON ordens_operacionais(equipamento_id, ficha_op, rota_ordem)
                WHERE ficha_op IS NOT NULL AND rota_ordem IS NOT NULL;
            CREATE INDEX IF NOT EXISTS ix_ordens_codigo ON ordens_operacionais(codigo_normalizado);
            CREATE INDEX IF NOT EXISTS ix_ordens_ficha_op ON ordens_operacionais(ficha_op);
            CREATE INDEX IF NOT EXISTS ix_ordens_abertura ON ordens_operacionais(aberta_em);
            CREATE INDEX IF NOT EXISTS ix_paradas_inicio ON paradas_operacionais(iniciada_em);
            """
        )

        momento = agora_iso()
        existentes = {
            linha["usuario"]: dict(linha)
            for linha in conexao.execute("SELECT usuario, senha_hash FROM equipamentos").fetchall()
        }
        for cadastro in EQUIPAMENTOS_OPERADORES:
            atual = existentes.get(cadastro["usuario"])
            if atual is None:
                conexao.execute(
                    """
                    INSERT INTO equipamentos
                        (usuario, senha_hash, codigo, setor, maquina, ativo, criado_em, atualizado_em)
                    VALUES (?, ?, ?, ?, ?, 1, ?, ?)
                    """,
                    (
                        cadastro["usuario"],
                        hash_senha_operador(cadastro["usuario"], SENHA_PADRAO_OPERADORES),
                        cadastro["codigo"],
                        cadastro["setor"],
                        cadastro["maquina"],
                        momento,
                        momento,
                    ),
                )
            else:
                conexao.execute(
                    """
                    UPDATE equipamentos
                    SET codigo = ?, setor = ?, maquina = ?, ativo = 1, atualizado_em = ?
                    WHERE usuario = ?
                    """,
                    (
                        cadastro["codigo"],
                        cadastro["setor"],
                        cadastro["maquina"],
                        momento,
                        cadastro["usuario"],
                    ),
                )
        conexao.commit()
    return True


@st.cache_resource(show_spinner=False)
def inicializar_catalogo_fichas() -> bool:
    with conectar_catalogo_fichas() as conexao:
        conexao.execute("PRAGMA journal_mode = WAL")
        conexao.execute("PRAGMA synchronous = NORMAL")
        conexao.executescript(
            """
            CREATE TABLE IF NOT EXISTS fichas_catalogo (
                op TEXT PRIMARY KEY,
                codigo_barras TEXT NOT NULL UNIQUE,
                lote TEXT,
                codigo_peca TEXT,
                descricao_peca TEXT,
                codigo_revestimento TEXT,
                revestimento TEXT,
                medida TEXT,
                tipo_material TEXT,
                quantidade_material REAL NOT NULL DEFAULT 0,
                quantidade_produzir REAL NOT NULL DEFAULT 0,
                obs_lote TEXT,
                emissao_data TEXT,
                emissao_hora TEXT,
                pagina_pdf INTEGER,
                produtos_json TEXT NOT NULL DEFAULT '[]',
                rotas_json TEXT NOT NULL DEFAULT '[]',
                fonte_arquivo TEXT,
                importado_em TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS rotas_catalogo (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                op TEXT NOT NULL,
                ordem_rota INTEGER NOT NULL,
                codigo_setor TEXT,
                setor TEXT,
                codigo_equipamento TEXT NOT NULL,
                equipamento TEXT,
                operacao TEXT,
                pecas_por_vez INTEGER NOT NULL DEFAULT 0,
                FOREIGN KEY (op) REFERENCES fichas_catalogo(op) ON DELETE CASCADE,
                UNIQUE(op, ordem_rota)
            );

            CREATE TABLE IF NOT EXISTS importacoes_catalogo (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                arquivo TEXT NOT NULL,
                importado_em TEXT NOT NULL,
                fichas_importadas INTEGER NOT NULL,
                lote_principal TEXT
            );

            CREATE INDEX IF NOT EXISTS ix_catalogo_lote ON fichas_catalogo(lote);
            CREATE INDEX IF NOT EXISTS ix_catalogo_peca ON fichas_catalogo(codigo_peca);
            CREATE INDEX IF NOT EXISTS ix_rotas_op_equipamento
                ON rotas_catalogo(op, codigo_equipamento, ordem_rota);
            """
        )
        conexao.commit()
    return True


def numero_br_para_float(valor: str) -> float:
    try:
        return float(valor.replace(".", "").replace(",", "."))
    except Exception:
        return 0.0


def separar_fichas_texto_pdf(texto: str) -> list[str]:
    partes = re.split(r"(?=1-DOBUE MOVELARIA)", texto or "")
    return [parte.strip() for parte in partes if re.search(r"O\.P\.\s+\d{8}", parte)]


def interpretar_ficha_texto(bloco: str, pagina_pdf: int) -> dict[str, Any] | None:
    """Interpreta o layout das fichas Dobue; preserva todos os campos úteis e o roteiro."""
    linhas = [linha.rstrip() for linha in bloco.splitlines()]
    texto = "\n".join(linhas)
    numero = r"[\d\.]+,\d+"

    op_match = re.search(r"O\.P\.\s+(\d{8})\s+([^\n]+)", texto)
    if not op_match:
        return None

    emissao = re.search(
        r"Emiss[aã]o:\s*(\d{2}/\d{2}/\d{2})\s+(\d{2}:\d{2}:\d{2})",
        texto,
        re.IGNORECASE,
    )
    lote = re.search(r"Obs\. Do Lote:\s*(.*?)\s+Lote:\s*(\d+)", texto, re.IGNORECASE)
    peca = re.search(r"Pe[çc]a:\s*(\d+)\s*-\s*([^\n]+)", texto, re.IGNORECASE)
    revestimento = re.search(
        r"Revestimento:\s*(\d+)\s+(.*?)\s*Medida:\s*(\S+)\s+Qtde:\s*("
        + numero
        + r")\s+("
        + numero
        + r")",
        texto,
        re.IGNORECASE,
    )

    ficha: dict[str, Any] = {
        "op": op_match.group(1),
        "codigo_barras": op_match.group(1),
        "tipo_material": op_match.group(2).strip(),
        "pagina_pdf": pagina_pdf,
        "emissao_data": emissao.group(1) if emissao else "",
        "emissao_hora": emissao.group(2) if emissao else "",
        "obs_lote": lote.group(1).strip() if lote else "",
        "lote": lote.group(2) if lote else "",
        "codigo_peca": peca.group(1) if peca else "",
        "descricao_peca": peca.group(2).strip() if peca else "",
        "codigo_revestimento": revestimento.group(1) if revestimento else "",
        "revestimento": revestimento.group(2).strip() if revestimento else "",
        "medida": revestimento.group(3) if revestimento else "",
        "quantidade_material": numero_br_para_float(revestimento.group(4)) if revestimento else 0.0,
        "quantidade_produzir": numero_br_para_float(revestimento.group(5)) if revestimento else 0.0,
        "produtos": [],
        "rotas": [],
    }

    try:
        indice_produtos = next(i for i, linha in enumerate(linhas) if "Produtos Relacionados" in linha)
        indice_setor = next(
            i
            for i, linha in enumerate(linhas[indice_produtos + 1 :], indice_produtos + 1)
            if re.match(r"\s*Setor\s+", linha)
        )
    except StopIteration:
        return ficha

    for linha in linhas[indice_produtos + 2 : indice_setor]:
        conteudo = linha.strip()
        if not conteudo or conteudo.startswith("Produto:"):
            continue
        partes = [p.strip() for p in re.split(r"\s{2,}", conteudo) if p.strip()]
        if len(partes) < 5:
            continue
        if not re.fullmatch(numero, partes[-1]) or not re.fullmatch(numero, partes[-2]):
            continue
        if not partes[-3].isdigit():
            continue

        primeiro = partes[0]
        primeiro_match = re.match(r"^(\d+)\s+(.+)$", primeiro)
        if primeiro_match:
            codigo_produto = primeiro_match.group(1)
            descricao_produto = primeiro_match.group(2)
            cor = " ".join(partes[1:-3])
        elif primeiro.isdigit() and len(partes) >= 6:
            codigo_produto = primeiro
            descricao_produto = partes[1]
            cor = " ".join(partes[2:-3])
        else:
            continue

        ficha["produtos"].append(
            {
                "codigo": codigo_produto,
                "descricao": descricao_produto,
                "revestimento_cor": cor,
                "sequencia": int(partes[-3]),
                "qtde_por_produto": numero_br_para_float(partes[-2]),
                "qtde_produzir": numero_br_para_float(partes[-1]),
            }
        )

    indice = indice_setor + 1
    while indice < len(linhas):
        linha = linhas[indice]
        if linha.startswith("1-DOBUE"):
            break
        rota_match = re.match(
            r"^\s*(\d+)\s+(.+?)\s{2,}(\d+)\s+(.+?)\s{2,}(.+?)\s{2,}(\d+)\s*$",
            linha,
        )
        if rota_match:
            rota = {
                "ordem": len(ficha["rotas"]) + 1,
                "codigo_setor": rota_match.group(1),
                "setor": rota_match.group(2).strip(),
                "codigo_equipamento": rota_match.group(3),
                "equipamento": rota_match.group(4).strip(),
                "operacao": rota_match.group(5).strip(),
                "pecas_por_vez": int(rota_match.group(6)),
            }
            continuacao: list[str] = []
            proxima = indice + 1
            while proxima < len(linhas):
                seguinte = linhas[proxima].strip()
                if not seguinte:
                    proxima += 1
                    continue
                if seguinte.startswith("Responsável:") or re.match(r"^\d+\s+", seguinte):
                    break
                continuacao.append(seguinte)
                proxima += 1
            if continuacao:
                rota["operacao"] = f"{rota['operacao']} {' '.join(continuacao)}".strip()
            ficha["rotas"].append(rota)
        indice += 1

    return ficha


def extrair_fichas_pdf(conteudo_pdf: bytes) -> list[dict[str, Any]]:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise RuntimeError(
            "Para importar fichas PDF, instale o pacote pypdf: python -m pip install pypdf"
        ) from exc

    import io

    leitor = PdfReader(io.BytesIO(conteudo_pdf))
    fichas: list[dict[str, Any]] = []
    for numero_pagina, pagina in enumerate(leitor.pages, start=1):
        try:
            texto = pagina.extract_text(extraction_mode="layout")
        except TypeError:
            texto = pagina.extract_text()
        for bloco in separar_fichas_texto_pdf(texto or ""):
            ficha = interpretar_ficha_texto(bloco, numero_pagina)
            if ficha:
                fichas.append(ficha)

    ops = [ficha["op"] for ficha in fichas]
    if not fichas:
        raise ValueError("Nenhuma ficha com O.P. de oito dígitos foi encontrada no PDF.")
    if len(set(ops)) != len(ops):
        repetidas = sorted({op for op in ops if ops.count(op) > 1})
        raise ValueError(f"O PDF contém O.P.s repetidas: {', '.join(repetidas[:10])}.")
    if any(not ficha["rotas"] for ficha in fichas):
        sem_rota = [ficha["op"] for ficha in fichas if not ficha["rotas"]]
        raise ValueError(f"Não foi possível ler o roteiro destas O.P.s: {', '.join(sem_rota[:10])}.")
    return fichas


def gravar_fichas_catalogo(fichas: list[dict[str, Any]], nome_arquivo: str) -> int:
    inicializar_catalogo_fichas()
    momento = agora_iso()
    lotes = [texto_limpo(ficha.get("lote")) for ficha in fichas if texto_limpo(ficha.get("lote"))]
    lote_principal = max(set(lotes), key=lotes.count) if lotes else ""

    conexao = conectar_catalogo_fichas()
    try:
        conexao.execute("BEGIN IMMEDIATE")
        for ficha in fichas:
            op = ficha["op"]
            conexao.execute("DELETE FROM fichas_catalogo WHERE op = ?", (op,))
            conexao.execute(
                """
                INSERT INTO fichas_catalogo (
                    op, codigo_barras, lote, codigo_peca, descricao_peca,
                    codigo_revestimento, revestimento, medida, tipo_material,
                    quantidade_material, quantidade_produzir, obs_lote,
                    emissao_data, emissao_hora, pagina_pdf, produtos_json,
                    rotas_json, fonte_arquivo, importado_em
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    op,
                    ficha.get("codigo_barras", op),
                    ficha.get("lote", ""),
                    ficha.get("codigo_peca", ""),
                    ficha.get("descricao_peca", ""),
                    ficha.get("codigo_revestimento", ""),
                    ficha.get("revestimento", ""),
                    ficha.get("medida", ""),
                    ficha.get("tipo_material", ""),
                    float(ficha.get("quantidade_material", 0) or 0),
                    float(ficha.get("quantidade_produzir", 0) or 0),
                    ficha.get("obs_lote", ""),
                    ficha.get("emissao_data", ""),
                    ficha.get("emissao_hora", ""),
                    int(ficha.get("pagina_pdf", 0) or 0),
                    json.dumps(ficha.get("produtos", []), ensure_ascii=False),
                    json.dumps(ficha.get("rotas", []), ensure_ascii=False),
                    nome_arquivo,
                    momento,
                ),
            )
            for rota in ficha.get("rotas", []):
                conexao.execute(
                    """
                    INSERT INTO rotas_catalogo (
                        op, ordem_rota, codigo_setor, setor, codigo_equipamento,
                        equipamento, operacao, pecas_por_vez
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        op,
                        int(rota.get("ordem", 0)),
                        texto_limpo(rota.get("codigo_setor")),
                        texto_limpo(rota.get("setor")),
                        texto_limpo(rota.get("codigo_equipamento")),
                        texto_limpo(rota.get("equipamento")),
                        texto_limpo(rota.get("operacao")),
                        int(rota.get("pecas_por_vez", 0) or 0),
                    ),
                )
        conexao.execute(
            """
            INSERT INTO importacoes_catalogo (arquivo, importado_em, fichas_importadas, lote_principal)
            VALUES (?, ?, ?, ?)
            """,
            (nome_arquivo, momento, len(fichas), lote_principal),
        )
        conexao.commit()
        return len(fichas)
    except Exception:
        conexao.rollback()
        raise
    finally:
        conexao.close()


def resumo_catalogo_fichas() -> dict[str, Any]:
    inicializar_catalogo_fichas()
    with conectar_catalogo_fichas() as conexao:
        linha = conexao.execute(
            """
            SELECT COUNT(*) AS fichas, COUNT(DISTINCT lote) AS lotes,
                   MIN(op) AS primeira_op, MAX(op) AS ultima_op,
                   MAX(importado_em) AS ultima_importacao
            FROM fichas_catalogo
            """
        ).fetchone()
    return dict(linha) if linha else {"fichas": 0, "lotes": 0}


def normalizar_codigo_operacional(valor: Any) -> str:
    return re.sub(r"\s+", "", texto_limpo(valor))


def validar_formato_codigo_barras(valor: Any) -> str:
    codigo = normalizar_codigo_operacional(valor)
    if not re.fullmatch(r"\d{8}", codigo):
        raise ValueError("Código inválido. Leia o código de barras da ficha: ele deve resultar em uma O.P. de 8 dígitos.")
    return codigo


def ficha_linha_para_dict(linha: sqlite3.Row) -> dict[str, Any]:
    ficha = dict(linha)
    for campo in ["produtos_json", "rotas_json"]:
        try:
            ficha[campo.removesuffix("_json")] = json.loads(ficha.get(campo) or "[]")
        except Exception:
            ficha[campo.removesuffix("_json")] = []
    return ficha


def buscar_ficha_para_equipamento(codigo_lido: Any, equipamento: dict[str, Any]) -> tuple[dict[str, Any], dict[str, Any]]:
    """Aceita somente O.P. existente no catálogo e prevista no roteiro da máquina logada."""
    codigo = validar_formato_codigo_barras(codigo_lido)
    inicializar_catalogo_fichas()
    with conectar_catalogo_fichas() as conexao:
        ficha_linha = conexao.execute(
            "SELECT * FROM fichas_catalogo WHERE codigo_barras = ? OR op = ?",
            (codigo, codigo),
        ).fetchone()
        if ficha_linha is None:
            raise ValueError(
                f"O código {codigo} não existe no catálogo importado. A ordem não foi aberta."
            )
        rotas = conexao.execute(
            """
            SELECT * FROM rotas_catalogo
            WHERE op = ? AND ltrim(codigo_equipamento, '0') = ltrim(?, '0')
            ORDER BY ordem_rota
            """,
            (codigo, texto_limpo(equipamento["codigo"])),
        ).fetchall()

    ficha = ficha_linha_para_dict(ficha_linha)
    if not rotas:
        codigos_rota = ", ".join(
            texto_limpo(rota.get("codigo_equipamento"))
            for rota in ficha.get("rotas", [])
            if texto_limpo(rota.get("codigo_equipamento"))
        )
        raise ValueError(
            f"A O.P. {codigo} é válida, mas não está programada para o equipamento "
            f"{equipamento['codigo']}. Equipamentos do roteiro: {codigos_rota or 'não identificados'}."
        )

    with conectar_banco_operacional() as conexao:
        utilizadas = {
            int(linha["rota_ordem"])
            for linha in conexao.execute(
                """
                SELECT rota_ordem FROM ordens_operacionais
                WHERE equipamento_id = ? AND ficha_op = ?
                  AND rota_ordem IS NOT NULL AND status IN ('ABERTA', 'FINALIZADA')
                """,
                (int(equipamento["id"]), codigo),
            ).fetchall()
        }

    rota_disponivel = next((dict(rota) for rota in rotas if int(rota["ordem_rota"]) not in utilizadas), None)
    if rota_disponivel is None:
        raise ValueError(
            f"A O.P. {codigo} já foi apontada em todas as passagens previstas para este equipamento."
        )
    return ficha, rota_disponivel


def autenticar_operador(usuario: str, senha: str) -> dict[str, Any] | None:
    inicializar_banco_operacional()
    with conectar_banco_operacional() as conexao:
        linha = conexao.execute(
            "SELECT * FROM equipamentos WHERE lower(usuario) = lower(?) AND ativo = 1",
            (usuario.strip(),),
        ).fetchone()
    if linha is None:
        return None
    informado = hash_senha_operador(linha["usuario"], senha)
    return dict(linha) if hmac.compare_digest(linha["senha_hash"], informado) else None


def operador_autenticado() -> bool:
    return bool(st.session_state.get("operador_equipamento_id"))


def obter_equipamento_sessao() -> dict[str, Any] | None:
    equipamento_id = st.session_state.get("operador_equipamento_id")
    if not equipamento_id:
        return None
    with conectar_banco_operacional() as conexao:
        linha = conexao.execute(
            "SELECT * FROM equipamentos WHERE id = ? AND ativo = 1",
            (int(equipamento_id),),
        ).fetchone()
    return dict(linha) if linha else None


def estado_operacional_equipamento(equipamento_id: int) -> tuple[str, dict[str, Any] | None]:
    with conectar_banco_operacional() as conexao:
        linhas = conexao.execute(
            """
            SELECT 'ORDEM' AS tipo, id, aberta_em AS inicio FROM ordens_operacionais
            WHERE equipamento_id = ? AND status = 'ABERTA'
            UNION ALL
            SELECT 'PARADA' AS tipo, id, iniciada_em AS inicio FROM paradas_operacionais
            WHERE equipamento_id = ? AND status = 'ABERTA'
            ORDER BY inicio DESC
            """,
            (equipamento_id, equipamento_id),
        ).fetchall()
        if len(linhas) > 1:
            return "INCONSISTENTE", None
        if not linhas:
            return "LIVRE", None
        tipo = linhas[0]["tipo"]
        tabela = "ordens_operacionais" if tipo == "ORDEM" else "paradas_operacionais"
        registro = conexao.execute(
            f"SELECT * FROM {tabela} WHERE id = ?",
            (int(linhas[0]["id"]),),
        ).fetchone()
    return tipo, dict(registro) if registro else None


def abrir_ordem_operacional(
    equipamento: dict[str, Any],
    codigo_lido: Any,
    ficha: dict[str, Any],
    rota: dict[str, Any],
) -> int:
    codigo = validar_formato_codigo_barras(codigo_lido)
    conexao = conectar_banco_operacional()
    try:
        conexao.execute("BEGIN IMMEDIATE")
        ocupada = conexao.execute(
            """
            SELECT 1 FROM ordens_operacionais WHERE equipamento_id = ? AND status = 'ABERTA'
            UNION ALL
            SELECT 1 FROM paradas_operacionais WHERE equipamento_id = ? AND status = 'ABERTA'
            LIMIT 1
            """,
            (equipamento["id"], equipamento["id"]),
        ).fetchone()
        if ocupada:
            raise ValueError("Este equipamento já possui uma ordem ou parada em andamento.")

        cursor = conexao.execute(
            """
            INSERT INTO ordens_operacionais (
                equipamento_id, usuario, setor, maquina, codigo_barras,
                codigo_normalizado, lote_identificado, codigo_peca_identificado,
                descricao_peca, quantidade_programada, aberta_em, status,
                ficha_op, rota_ordem, rota_codigo_equipamento, rota_equipamento,
                rota_operacao, tipo_material, revestimento_ficha, medida_ficha,
                obs_lote_ficha, ficha_snapshot_json
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'ABERTA', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                int(equipamento["id"]),
                equipamento["usuario"],
                equipamento["setor"],
                equipamento["maquina"],
                codigo,
                codigo,
                texto_limpo(ficha.get("lote")),
                texto_limpo(ficha.get("codigo_peca")),
                texto_limpo(ficha.get("descricao_peca")),
                float(ficha.get("quantidade_produzir", 0) or 0),
                agora_iso(),
                codigo,
                int(rota["ordem_rota"]),
                texto_limpo(rota.get("codigo_equipamento")),
                texto_limpo(rota.get("equipamento")),
                texto_limpo(rota.get("operacao")),
                texto_limpo(ficha.get("tipo_material")),
                texto_limpo(ficha.get("revestimento")),
                texto_limpo(ficha.get("medida")),
                texto_limpo(ficha.get("obs_lote")),
                json.dumps({"ficha": ficha, "rota": rota}, ensure_ascii=False),
            ),
        )
        conexao.commit()
        return int(cursor.lastrowid)
    except sqlite3.IntegrityError as exc:
        conexao.rollback()
        raise ValueError("Esta passagem da O.P. já foi apontada neste equipamento.") from exc
    except Exception:
        conexao.rollback()
        raise
    finally:
        conexao.close()


def finalizar_ordem_operacional(
    equipamento_id: int,
    ordem_id: int,
    quantidade_boa: float,
    refugo: float,
    retrabalho: float,
    observacao: str,
) -> int:
    if min(quantidade_boa, refugo, retrabalho) < 0:
        raise ValueError("As quantidades não podem ser negativas.")
    conexao = conectar_banco_operacional()
    try:
        conexao.execute("BEGIN IMMEDIATE")
        ordem = conexao.execute(
            """
            SELECT * FROM ordens_operacionais
            WHERE id = ? AND equipamento_id = ? AND status = 'ABERTA'
            """,
            (ordem_id, equipamento_id),
        ).fetchone()
        if ordem is None:
            raise ValueError("A ordem não está mais aberta neste equipamento.")
        encerrada = datetime.now().replace(microsecond=0)
        aberta = datetime.fromisoformat(ordem["aberta_em"])
        duracao = max(0, int((encerrada - aberta).total_seconds()))
        conexao.execute(
            """
            UPDATE ordens_operacionais
            SET fechada_em = ?, duracao_segundos = ?, quantidade_boa = ?,
                refugo = ?, retrabalho = ?, observacao_fechamento = ?, status = 'FINALIZADA'
            WHERE id = ?
            """,
            (
                encerrada.isoformat(sep=" "),
                duracao,
                float(quantidade_boa),
                float(refugo),
                float(retrabalho),
                texto_limpo(observacao),
                ordem_id,
            ),
        )
        conexao.commit()
        return duracao
    except Exception:
        conexao.rollback()
        raise
    finally:
        conexao.close()


def iniciar_parada_operacional(equipamento: dict[str, Any], motivo: str) -> int:
    if motivo not in MOTIVOS_PARADA_OPERACIONAL:
        raise ValueError("Selecione um motivo de parada válido.")
    conexao = conectar_banco_operacional()
    try:
        conexao.execute("BEGIN IMMEDIATE")
        ocupada = conexao.execute(
            """
            SELECT 1 FROM ordens_operacionais WHERE equipamento_id = ? AND status = 'ABERTA'
            UNION ALL
            SELECT 1 FROM paradas_operacionais WHERE equipamento_id = ? AND status = 'ABERTA'
            LIMIT 1
            """,
            (equipamento["id"], equipamento["id"]),
        ).fetchone()
        if ocupada:
            raise ValueError("Este equipamento já possui uma ordem ou parada em andamento.")
        cursor = conexao.execute(
            """
            INSERT INTO paradas_operacionais
                (equipamento_id, usuario, setor, maquina, motivo, iniciada_em, status)
            VALUES (?, ?, ?, ?, ?, ?, 'ABERTA')
            """,
            (
                int(equipamento["id"]),
                equipamento["usuario"],
                equipamento["setor"],
                equipamento["maquina"],
                motivo,
                agora_iso(),
            ),
        )
        conexao.commit()
        return int(cursor.lastrowid)
    except Exception:
        conexao.rollback()
        raise
    finally:
        conexao.close()


def encerrar_parada_operacional(equipamento_id: int, parada_id: int) -> int:
    conexao = conectar_banco_operacional()
    try:
        conexao.execute("BEGIN IMMEDIATE")
        parada = conexao.execute(
            """
            SELECT * FROM paradas_operacionais
            WHERE id = ? AND equipamento_id = ? AND status = 'ABERTA'
            """,
            (parada_id, equipamento_id),
        ).fetchone()
        if parada is None:
            raise ValueError("A parada não está mais aberta neste equipamento.")
        encerrada = datetime.now().replace(microsecond=0)
        iniciada = datetime.fromisoformat(parada["iniciada_em"])
        duracao = max(0, int((encerrada - iniciada).total_seconds()))
        conexao.execute(
            """
            UPDATE paradas_operacionais
            SET encerrada_em = ?, duracao_segundos = ?, status = 'FINALIZADA'
            WHERE id = ?
            """,
            (encerrada.isoformat(sep=" "), duracao, parada_id),
        )
        conexao.commit()
        return duracao
    except Exception:
        conexao.rollback()
        raise
    finally:
        conexao.close()


def duracao_desde(inicio_iso: str) -> int:
    try:
        return max(0, int((datetime.now() - datetime.fromisoformat(inicio_iso)).total_seconds()))
    except Exception:
        return 0


def formatar_duracao_segundos(segundos: Any) -> str:
    try:
        total = max(0, int(float(segundos or 0)))
    except Exception:
        total = 0
    horas, resto = divmod(total, 3600)
    minutos, segundos = divmod(resto, 60)
    return f"{horas:02d}:{minutos:02d}:{segundos:02d}"


# Telas Monitor Operacional e Apontamento Operacional por Equipamento removidas na versão 5.19.


def apontamento_autenticado() -> bool:
    """Retorna True somente quando o usuário efetuou login nesta sessão."""
    return bool(st.session_state.get("apontamento_autenticado", False))


def autenticar_apontamento(usuario: str, senha: str) -> bool:
    """Compara as credenciais sem expor o resultado de comparações parciais."""
    usuario_ok = hmac.compare_digest(usuario.strip(), LOGIN_APONTAMENTO)
    senha_ok = hmac.compare_digest(senha, SENHA_APONTAMENTO)
    return usuario_ok and senha_ok


def renderizar_login_apontamento() -> None:
    """Exibe o formulário de login antes de permitir qualquer gravação."""
    st.title("Acesso aos Apontamentos PPCP")
    st.caption(
        "As consultas públicas permanecem abertas. Os apontamentos manuais "
        "exigem o usuário e a senha do PPCP."
    )

    col_esquerda, col_login, col_direita = st.columns([1, 1.15, 1])
    with col_login:
        with st.form("form_login_apontamento", clear_on_submit=False):
            usuario = st.text_input(
                "Login",
                placeholder="Digite o login",
                autocomplete="username",
            )
            senha = st.text_input(
                "Senha",
                type="password",
                placeholder="Digite a senha",
                autocomplete="current-password",
            )
            entrar = st.form_submit_button(
                "Entrar no modo apontamento",
                type="primary",
                use_container_width=True,
            )

        if entrar:
            if autenticar_apontamento(usuario, senha):
                st.session_state["apontamento_autenticado"] = True
                st.session_state.pop("erro_login_apontamento", None)
                st.rerun()
            else:
                st.session_state["erro_login_apontamento"] = True

        if st.session_state.get("erro_login_apontamento", False):
            st.error("Login ou senha incorretos.")

        st.info("O acesso permanece ativo somente nesta sessão do navegador.")


def renderizar_status_acesso() -> None:
    """Mostra o estado do acesso e oferece encerramento da sessão protegida."""
    with st.sidebar:
        st.divider()
        st.markdown("### Acesso")
        if apontamento_autenticado():
            st.success("Modo apontamento liberado")
            if st.button("Sair do apontamento", use_container_width=True):
                st.session_state["apontamento_autenticado"] = False
                st.session_state.pop("erro_login_apontamento", None)
                st.rerun()
        else:
            st.info("OPEN SOURCE")
            st.caption("Consultas públicas abertas.")


def localizar_arquivo_padrao() -> Path:
    """Localiza a planilha padrão no diretório atual ou ao lado do arquivo Python."""
    candidatos = [
        Path(ARQUIVO_PADRAO).expanduser(),
        Path(__file__).resolve().parent / ARQUIVO_PADRAO,
    ]
    for candidato in candidatos:
        caminho = candidato.resolve()
        if caminho.exists():
            return caminho
    return candidatos[0].resolve()


def carregar_arquivo_modo_leitura() -> tuple[Path, str] | None:
    """No acesso público, usa somente a planilha padrão e não permite trocar arquivos."""
    caminho = localizar_arquivo_padrao()
    with st.sidebar:
        st.header("Fonte dos dados")
        st.caption(f"Arquivo: {caminho.name}")
        st.caption(f"Aba: {NOME_ABA_PADRAO}")

    if not caminho.exists():
        st.error(
            f'Não encontrei o arquivo "{ARQUIVO_PADRAO}". '
            "Coloque a planilha na mesma pasta do aplicativo para liberar a consulta pública."
        )
        return None
    return caminho, NOME_ABA_PADRAO


# -----------------------------------------------------------------------------
# Inicialização do aplicativo
# -----------------------------------------------------------------------------
def rodando_via_streamlit() -> bool:
    """Retorna True quando o arquivo foi iniciado pelo comando streamlit run."""
    try:
        from streamlit.runtime.scriptrunner import get_script_run_ctx

        return get_script_run_ctx() is not None
    except Exception:
        return False


def abrir_com_streamlit_se_necessario() -> bool:
    """Permite abrir o arquivo pelo PyCharm/duplo clique sem gerar erro de ScriptRunContext."""
    if rodando_via_streamlit():
        return False

    script = Path(__file__).resolve()
    print("\nEste é um aplicativo Streamlit.")
    print("Abrindo corretamente no navegador com o comando:")
    print(f'  "{sys.executable}" -m streamlit run "{script}"\n')

    try:
        subprocess.run([sys.executable, "-m", "streamlit", "run", str(script)], check=False)
    except FileNotFoundError:
        print("Python não encontrado para iniciar o Streamlit.")
    return True


def carregar_arquivo_interface() -> tuple[Path, str] | None:
    with st.sidebar:
        st.header("Arquivo")
        caminho_txt = st.text_input("Caminho do Excel", value=ARQUIVO_PADRAO)
        nome_aba = st.text_input("Aba", value=NOME_ABA_PADRAO)
        st.info("Deixe a planilha Excel fechada antes de salvar, senão o Windows pode bloquear o arquivo.")

    caminho = Path(caminho_txt).expanduser().resolve()

    if not caminho.exists():
        st.warning("Arquivo não encontrado pelo caminho informado.")
        arquivo = st.file_uploader("Ou envie a planilha aqui", type=["xlsx"])
        if arquivo is None:
            st.info("Informe um caminho válido para a planilha ou envie o arquivo Excel acima.")
            st.stop()
            return None
        pasta_tmp = Path("_arquivo_trabalho")
        pasta_tmp.mkdir(exist_ok=True)
        caminho = pasta_tmp / arquivo.name
        conteudo = arquivo.getvalue()
        if not caminho.exists() or caminho.stat().st_size != len(conteudo):
            caminho.write_bytes(conteudo)
            carregar_dados_cacheado.clear()
        st.success(f"Arquivo carregado temporariamente: {caminho}")

    return caminho, nome_aba


def main() -> None:
    st.set_page_config(page_title="Acompanhamento PPCP", layout="wide")

    st.sidebar.success(f"VERSÃO {VERSAO_APP}")
    st.sidebar.markdown("## SISTEMA MES")

    pagina_param = str(st.query_params.get("pagina", "")).lower()
    mapa_parametros = {
        "programado-equipamento": 0,
        "programado-dia-equipamento": 0,
        "equipamentos": 0,
        "acumulado-setor": 1,
        "acumulado-diario-setor": 1,
        "metas-setor": 1,
        "calendario": 2,
        "semanal": 2,
        "consulta-paradas": 3,
        "paradas-leitura": 3,
        "apontamento": 4,
        "producao-apontamento": 4,
        "paradas": 5,
        "apontamento-paradas": 5,
    }
    pagina_padrao = mapa_parametros.get(pagina_param, 0)

    paginas = [
        "🏭 Programado por Dia/Equipamento — leitura",
        "📈 Acumulado Diário por Setor — leitura",
        "🗓️ Calendário Semanal dos Setores — leitura",
        "📋 Paradas por Setor/Equipamento — leitura",
        "🔐 Apontamento Manual de Produção",
        "⏸️ Apontamento Manual de Paradas",
    ]
    pagina_padrao = min(max(pagina_padrao, 0), len(paginas) - 1)
    pagina = st.sidebar.radio("MENU", paginas, index=pagina_padrao)
    st.sidebar.caption(
        "SISTEMA DE APONTAMENTO DE PRODUÇÃO"
    )
    renderizar_status_acesso()

    paginas_pcp = {
        "🔐 Apontamento Manual de Produção",
        "⏸️ Apontamento Manual de Paradas",
    }
    if pagina in paginas_pcp:
        if not apontamento_autenticado():
            renderizar_login_apontamento()
            return
        arquivo_config = carregar_arquivo_interface()
    else:
        arquivo_config = carregar_arquivo_modo_leitura()

    if arquivo_config is None:
        return
    caminho, nome_aba = arquivo_config

    if "Paradas por Setor/Equipamento" in pagina:
        renderizar_consulta_paradas(caminho)
        return

    try:
        with st.spinner("Carregando dados da planilha..."):
            df, _ = obter_dados_cacheados(caminho, nome_aba)
    except Exception as exc:
        st.error(f"Erro ao carregar a planilha: {exc}")
        st.stop()
        return

    if df.empty:
        st.warning("Não encontrei linhas de acompanhamento para exibir.")
        st.stop()
        return

    if pagina == "⏸️ Apontamento Manual de Paradas":
        renderizar_apontamento_paradas(df, caminho, nome_aba)
    elif pagina == "🔐 Apontamento Manual de Produção":
        renderizar_apontamento(df, caminho, nome_aba)
    elif "Acumulado Diário por Setor" in pagina:
        renderizar_acumulado_diario_setor(df, caminho)
    elif "Calendário Semanal" in pagina:
        renderizar_calendario_semanal(df, caminho)
    else:
        renderizar_programado_dia_equipamento(df, caminho)


if __name__ == "__main__":
    if abrir_com_streamlit_se_necessario():
        raise SystemExit(0)
    main()


## python -m streamlit run app_acompanhamento_ppcp.py

## python -m streamlit run j2i_mes.py --server.address 127.0.0.1 --server.port 8501

## cd "C:\Users\alanf\Videos\Project\Sistema de apontamento"

