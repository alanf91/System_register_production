import os
import re
import shutil
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook


# ============================================================
# CONFIGURAÇÃO
# ============================================================

load_dotenv()

ARQUIVO = Path(
    os.getenv(
        "EXCEL_FILE",
        "controle_transferencias_configurado.xlsx",
    )
)

# Deixe vazio para o programa localizar automaticamente a aba
# que contém os cabeçalhos obrigatórios.
ABA_PREFERENCIAL = os.getenv("EXCEL_SHEET", "").strip()

TOKEN = os.getenv("WHATSAPP_TOKEN", "").strip()
PHONE_NUMBER_ID = os.getenv(
    "WHATSAPP_PHONE_NUMBER_ID",
    "",
).strip()

API_VERSION = os.getenv(
    "GRAPH_API_VERSION",
    "v26.0",
).strip()

# "texto": útil para testes ou conversas dentro da janela permitida.
# "template": recomendado para notificações proativas.
MODO = os.getenv(
    "WHATSAPP_MODE",
    "template",
).strip().lower()

TEMPLATE = os.getenv(
    "WHATSAPP_TEMPLATE_NAME",
    "atualizacao_movimentacao_fabrica",
).strip()

IDIOMA_TEMPLATE = os.getenv(
    "WHATSAPP_TEMPLATE_LANGUAGE",
    "pt_BR",
).strip()

# Limite conservador para dividir atualizações muito grandes.
LIMITE_DETALHES = int(
    os.getenv(
        "WHATSAPP_DETAILS_LIMIT",
        "2800",
    )
)


CABECALHOS_OBRIGATORIOS = {
    "ID",
    "FABRICA",
    "TELEFONE",
    "ITEM",
    "QUANTIDADE",
    "DATA_ENVIO",
    "DATA_RECEBIMENTO",
}


# ============================================================
# FUNÇÕES AUXILIARES
# ============================================================

def texto_limpo(valor):
    return str(valor or "").strip()


def normalizar_cabecalho(valor):
    return texto_limpo(valor).upper()


def limpar_telefone(valor):
    """
    Converte telefone brasileiro para:
    55 + DDD + número.
    """
    telefone = re.sub(r"\D", "", texto_limpo(valor))

    if len(telefone) in (10, 11):
        telefone = "55" + telefone

    if not telefone.startswith("55"):
        raise ValueError(
            f"Telefone sem código do Brasil: {valor}"
        )

    if len(telefone) not in (12, 13):
        raise ValueError(
            f"Telefone inválido: {valor}. "
            "Use DDD + número ou 55 + DDD + número."
        )

    return telefone


def converter_data(valor):
    """
    Retorna datetime.date para permitir agrupamento correto.
    """
    if valor is None or texto_limpo(valor) == "":
        return None

    if isinstance(valor, datetime):
        return valor.date()

    if isinstance(valor, date):
        return valor

    texto = texto_limpo(valor)

    formatos = (
        "%d/%m/%Y",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%Y/%m/%d",
    )

    for formato in formatos:
        try:
            return datetime.strptime(
                texto,
                formato,
            ).date()
        except ValueError:
            continue

    raise ValueError(
        f"Data inválida: {valor}. "
        "Use uma data real do Excel ou dd/mm/aaaa."
    )


def formatar_data(valor):
    return valor.strftime("%d/%m/%Y")


def formatar_quantidade(valor):
    if valor is None:
        return "-"

    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))

    return str(valor)


def localizar_aba(wb):
    """
    Usa a aba configurada ou localiza automaticamente uma aba
    com todos os cabeçalhos necessários.
    """
    if ABA_PREFERENCIAL:
        if ABA_PREFERENCIAL not in wb.sheetnames:
            raise RuntimeError(
                f"A aba '{ABA_PREFERENCIAL}' não existe. "
                f"Abas disponíveis: {wb.sheetnames}"
            )

        return wb[ABA_PREFERENCIAL]

    for ws in wb.worksheets:
        cabecalhos = {
            normalizar_cabecalho(celula.value)
            for celula in ws[1]
            if celula.value is not None
        }

        if CABECALHOS_OBRIGATORIOS.issubset(cabecalhos):
            return ws

    raise RuntimeError(
        "Nenhuma aba contém todos os cabeçalhos obrigatórios: "
        + ", ".join(sorted(CABECALHOS_OBRIGATORIOS))
    )


def mapear_colunas(ws):
    return {
        normalizar_cabecalho(celula.value): celula.column
        for celula in ws[1]
        if celula.value is not None
    }


def criar_coluna_se_nao_existir(ws, colunas, nome):
    if nome in colunas:
        return colunas[nome]

    numero = ws.max_column + 1
    ws.cell(
        row=1,
        column=numero,
        value=nome,
    )
    colunas[nome] = numero
    return numero


def extrair_message_id(retorno):
    mensagens = retorno.get("messages") or []

    if mensagens:
        return mensagens[0].get("id", "SEM_ID")

    return "SEM_ID"


def status_ja_enviado(valor):
    return texto_limpo(valor).upper() == "SIM"


# ============================================================
# MONTAGEM DAS MENSAGENS
# ============================================================

def titulo_evento(evento):
    if evento["tipo"] == "ENVIO":
        return "📤 ENVIOS"

    return "📥 RECEBIMENTOS"


def linha_evento(evento):
    return (
        f"• ID: {evento['id']} | "
        f"Item: {evento['item']} | "
        f"Qtd.: {formatar_quantidade(evento['quantidade'])}"
    )


def montar_detalhes(eventos):
    """
    Separa os dados em seções de envios e recebimentos.
    """
    envios = [
        evento
        for evento in eventos
        if evento["tipo"] == "ENVIO"
    ]

    recebimentos = [
        evento
        for evento in eventos
        if evento["tipo"] == "RECEBIMENTO"
    ]

    blocos = []

    if envios:
        blocos.append(
            "📤 ENVIOS\n"
            + "\n".join(linha_evento(e) for e in envios)
        )

    if recebimentos:
        blocos.append(
            "📥 RECEBIMENTOS\n"
            + "\n".join(
                linha_evento(e)
                for e in recebimentos
            )
        )

    return "\n\n".join(blocos)


def dividir_eventos_em_lotes(eventos):
    """
    Divide apenas quando a atualização ultrapassa o limite.
    Cada lote concluído pode ser marcado individualmente,
    reduzindo risco de duplicidade em uma nova execução.
    """
    lotes = []
    lote_atual = []

    for evento in eventos:
        candidato = lote_atual + [evento]
        detalhes = montar_detalhes(candidato)

        if (
            lote_atual
            and len(detalhes) > LIMITE_DETALHES
        ):
            lotes.append(lote_atual)
            lote_atual = [evento]
        else:
            lote_atual = candidato

    if lote_atual:
        lotes.append(lote_atual)

    return lotes


# ============================================================
# WHATSAPP CLOUD API
# ============================================================

def enviar_whatsapp(
    telefone,
    fabrica,
    data_movimentacao,
    detalhes,
):
    url = (
        f"https://graph.facebook.com/"
        f"{API_VERSION}/{PHONE_NUMBER_ID}/messages"
    )

    headers = {
        "Authorization": f"Bearer {TOKEN}",
        "Content-Type": "application/json",
    }

    data_formatada = formatar_data(data_movimentacao)

    if MODO == "template":
        payload = {
            "messaging_product": "whatsapp",
            "recipient_type": "individual",
            "to": telefone,
            "type": "template",
            "template": {
                "name": TEMPLATE,
                "language": {
                    "code": IDIOMA_TEMPLATE,
                },
                "components": [
                    {
                        "type": "body",
                        "parameters": [
                            {
                                "type": "text",
                                "text": fabrica,
                            },
                            {
                                "type": "text",
                                "text": data_formatada,
                            },
                            {
                                "type": "text",
                                "text": detalhes,
                            },
                        ],
                    }
                ],
            },
        }
    else:
        mensagem = (
            f"📦 *Movimentações da fábrica*\n"
            f"Fábrica: *{fabrica}*\n"
            f"Data: *{data_formatada}*\n\n"
            f"{detalhes}"
        )

        payload = {
            "messaging_product": "whatsapp",
            "recipient_type": "individual",
            "to": telefone,
            "type": "text",
            "text": {
                "preview_url": False,
                "body": mensagem,
            },
        }

    resposta = requests.post(
        url,
        headers=headers,
        json=payload,
        timeout=30,
    )

    try:
        dados = resposta.json()
    except ValueError:
        dados = {"resposta": resposta.text}

    if not resposta.ok:
        raise RuntimeError(
            f"Erro HTTP {resposta.status_code}: {dados}"
        )

    return dados


# ============================================================
# PROCESSAMENTO DA PLANILHA
# ============================================================

def coletar_eventos(ws, colunas, colunas_controle):
    """
    Cria um evento independente para cada:
    - envio ainda não avisado;
    - recebimento ainda não avisado.

    Agrupamento:
    fábrica + telefone + data da movimentação.
    """
    grupos = defaultdict(list)

    col_aviso_envio = colunas_controle["AVISO_ENVIO"]
    col_aviso_recebimento = (
        colunas_controle["AVISO_RECEBIMENTO"]
    )
    col_retorno = colunas_controle["RETORNO_WHATSAPP"]

    for linha in range(2, ws.max_row + 1):
        fabrica = ws.cell(
            linha,
            colunas["FABRICA"],
        ).value

        telefone_original = ws.cell(
            linha,
            colunas["TELEFONE"],
        ).value

        if not fabrica or not telefone_original:
            continue

        try:
            telefone = limpar_telefone(telefone_original)
        except ValueError as erro:
            ws.cell(
                linha,
                col_retorno,
                value=f"ERRO: {erro}",
            )
            continue

        identificador = ws.cell(
            linha,
            colunas["ID"],
        ).value

        item = ws.cell(
            linha,
            colunas["ITEM"],
        ).value

        quantidade = ws.cell(
            linha,
            colunas["QUANTIDADE"],
        ).value

        aviso_envio = ws.cell(
            linha,
            col_aviso_envio,
        ).value

        aviso_recebimento = ws.cell(
            linha,
            col_aviso_recebimento,
        ).value

        # Um envio é um evento independente.
        if not status_ja_enviado(aviso_envio):
            valor_data_envio = ws.cell(
                linha,
                colunas["DATA_ENVIO"],
            ).value

            try:
                data_envio = converter_data(valor_data_envio)
            except ValueError as erro:
                ws.cell(
                    linha,
                    col_retorno,
                    value=f"ERRO DATA_ENVIO: {erro}",
                )
                data_envio = None

            if data_envio:
                chave = (
                    texto_limpo(fabrica),
                    telefone,
                    data_envio,
                )

                grupos[chave].append(
                    {
                        "linha": linha,
                        "tipo": "ENVIO",
                        "id": identificador,
                        "item": item,
                        "quantidade": quantidade,
                    }
                )

        # Um recebimento é outro evento independente.
        if not status_ja_enviado(aviso_recebimento):
            valor_data_recebimento = ws.cell(
                linha,
                colunas["DATA_RECEBIMENTO"],
            ).value

            try:
                data_recebimento = converter_data(
                    valor_data_recebimento
                )
            except ValueError as erro:
                ws.cell(
                    linha,
                    col_retorno,
                    value=f"ERRO DATA_RECEBIMENTO: {erro}",
                )
                data_recebimento = None

            if data_recebimento:
                chave = (
                    texto_limpo(fabrica),
                    telefone,
                    data_recebimento,
                )

                grupos[chave].append(
                    {
                        "linha": linha,
                        "tipo": "RECEBIMENTO",
                        "id": identificador,
                        "item": item,
                        "quantidade": quantidade,
                    }
                )

    return grupos


def marcar_lote_como_enviado(
    ws,
    eventos,
    colunas_controle,
    message_id,
):
    agora = datetime.now()

    for evento in eventos:
        linha = evento["linha"]

        if evento["tipo"] == "ENVIO":
            ws.cell(
                linha,
                colunas_controle["AVISO_ENVIO"],
                value="SIM",
            )
        else:
            ws.cell(
                linha,
                colunas_controle["AVISO_RECEBIMENTO"],
                value="SIM",
            )

        ws.cell(
            linha,
            colunas_controle["DATA_ULTIMO_AVISO"],
            value=agora,
        )

        retorno_atual = texto_limpo(
            ws.cell(
                linha,
                colunas_controle["RETORNO_WHATSAPP"],
            ).value
        )

        novo_retorno = f"SUCESSO: {message_id}"

        if retorno_atual.startswith("SUCESSO:"):
            novo_retorno = (
                retorno_atual
                + " | "
                + novo_retorno
            )

        ws.cell(
            linha,
            colunas_controle["RETORNO_WHATSAPP"],
            value=novo_retorno,
        )


def registrar_erro(
    ws,
    eventos,
    col_retorno,
    erro,
):
    for evento in eventos:
        linha = evento["linha"]
        ws.cell(
            linha,
            col_retorno,
            value=f"ERRO: {erro}",
        )


def processar_planilha():
    if not ARQUIVO.exists():
        raise FileNotFoundError(
            f"Arquivo não encontrado: {ARQUIVO.resolve()}"
        )

    if not TOKEN:
        raise RuntimeError(
            "WHATSAPP_TOKEN não configurado no arquivo .env."
        )

    if not PHONE_NUMBER_ID:
        raise RuntimeError(
            "WHATSAPP_PHONE_NUMBER_ID não configurado no .env."
        )

    if MODO not in ("texto", "template"):
        raise RuntimeError(
            "WHATSAPP_MODE deve ser 'texto' ou 'template'."
        )

    # Backup antes de qualquer alteração.
    pasta_backup = ARQUIVO.parent / "backup"
    pasta_backup.mkdir(exist_ok=True)

    horario = datetime.now().strftime("%Y%m%d_%H%M%S")
    arquivo_backup = (
        pasta_backup
        / f"{ARQUIVO.stem}_{horario}.xlsx"
    )
    shutil.copy2(ARQUIVO, arquivo_backup)

    wb = load_workbook(ARQUIVO)
    ws = localizar_aba(wb)
    colunas = mapear_colunas(ws)

    faltantes = (
        CABECALHOS_OBRIGATORIOS
        - set(colunas)
    )

    if faltantes:
        raise RuntimeError(
            "Colunas obrigatórias ausentes: "
            + ", ".join(sorted(faltantes))
        )

    colunas_controle = {
        "AVISO_ENVIO": criar_coluna_se_nao_existir(
            ws,
            colunas,
            "AVISO_ENVIO",
        ),
        "AVISO_RECEBIMENTO": criar_coluna_se_nao_existir(
            ws,
            colunas,
            "AVISO_RECEBIMENTO",
        ),
        "DATA_ULTIMO_AVISO": criar_coluna_se_nao_existir(
            ws,
            colunas,
            "DATA_ULTIMO_AVISO",
        ),
        "RETORNO_WHATSAPP": criar_coluna_se_nao_existir(
            ws,
            colunas,
            "RETORNO_WHATSAPP",
        ),
    }

    grupos = coletar_eventos(
        ws,
        colunas,
        colunas_controle,
    )

    # Salva também eventuais erros de validação.
    wb.save(ARQUIVO)

    if not grupos:
        print(
            "Nenhum novo envio ou recebimento "
            "para comunicar."
        )
        return

    # Ordenação deixa o processamento previsível:
    # fábrica, data e telefone.
    chaves_ordenadas = sorted(
        grupos,
        key=lambda chave: (
            chave[0].casefold(),
            chave[2],
            chave[1],
        ),
    )

    for chave in chaves_ordenadas:
        fabrica, telefone, data_movimentacao = chave
        eventos = grupos[chave]

        lotes = dividir_eventos_em_lotes(eventos)
        total_lotes = len(lotes)

        for numero_lote, lote in enumerate(
            lotes,
            start=1,
        ):
            detalhes = montar_detalhes(lote)

            if total_lotes > 1:
                detalhes = (
                    f"Parte {numero_lote}/{total_lotes}\n\n"
                    f"{detalhes}"
                )

            try:
                retorno = enviar_whatsapp(
                    telefone=telefone,
                    fabrica=fabrica,
                    data_movimentacao=data_movimentacao,
                    detalhes=detalhes,
                )

                message_id = extrair_message_id(retorno)

                marcar_lote_como_enviado(
                    ws=ws,
                    eventos=lote,
                    colunas_controle=colunas_controle,
                    message_id=message_id,
                )

                # Salva após cada lote enviado.
                wb.save(ARQUIVO)

                print(
                    "OK: "
                    f"{fabrica} | "
                    f"{formatar_data(data_movimentacao)} | "
                    f"lote {numero_lote}/{total_lotes} | "
                    f"{len(lote)} movimentação(ões)."
                )

            except Exception as erro:
                registrar_erro(
                    ws=ws,
                    eventos=lote,
                    col_retorno=colunas_controle[
                        "RETORNO_WHATSAPP"
                    ],
                    erro=erro,
                )

                wb.save(ARQUIVO)

                print(
                    "FALHA: "
                    f"{fabrica} | "
                    f"{formatar_data(data_movimentacao)} | "
                    f"lote {numero_lote}/{total_lotes}: "
                    f"{erro}"
                )

    print("Processamento concluído.")


if __name__ == "__main__":
    processar_planilha()
